"""
e-CALLISTO FITS Analyzer
Version 2.0
Sahan S Liyanage (sahanslst@gmail.com)
Astronomical and Space Science Unit, University of Colombo, Sri Lanka.
"""
import sys
import io
import os
import tempfile
import re
import gc
import requests
from PySide6.QtWidgets import (
    QMainWindow, QSlider, QDialog, QMenuBar, QMessageBox, QLabel, QFormLayout, QGroupBox,
    QStatusBar, QProgressBar, QApplication, QMenu, QCheckBox, QRadioButton, QButtonGroup, QComboBox, QToolBar,
    QLineEdit, QSpinBox, QScrollArea, QFrame, QVBoxLayout, QWidget, QFileDialog, QHBoxLayout, QSizePolicy, QLayout,
    QStackedLayout,
    QProgressDialog,
    QInputDialog,
)

from PySide6.QtGui import QAction, QPixmap, QImage, QGuiApplication, QIcon, QFontDatabase, QActionGroup, QPalette, QPainter, QPdfWriter
from PySide6.QtCore import Qt, QTimer, QSize, QObject, QEvent, QThread, Signal, Slot
from src.UI.callisto_downloader import CallistoDownloaderApp
from src.UI.goes_xrs_gui import MainWindow as GoesXrsWindow
#from soho_lasco_viewer import CMEViewer as CMEViewerWindow
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.widgets import LassoSelector
from matplotlib.widgets import RectangleSelector
from matplotlib.path import Path
from astropy.io import fits
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.colors as mcolors
from mpl_toolkits.axes_grid1 import make_axes_locatable
from matplotlib.ticker import FuncFormatter, ScalarFormatter
import csv
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from src.UI.mpl_style import style_axes
from src.UI.accelerated_plot_widget import AcceleratedPlotWidget
from src.Backend.project_session import ProjectFormatError, read_project, write_project
from src.Backend.fits_io import build_combined_header, extract_ut_start_sec, load_callisto_fits
from src.UI.fits_header_viewer import FitsHeaderViewerDialog
#from PySide6.QtCore import QObject, QEvent
#from PySide6.QtWidgets import QLayout

#LINUX Specific Fixes for messageboxes
IS_LINUX = sys.platform.startswith("linux")

_linux_msgbox_fixer = None

class _LinuxMessageBoxFixer(QObject):
    def eventFilter(self, obj, event):
        if IS_LINUX and event.type() == QEvent.Show and isinstance(obj, QMessageBox):
            # Make the main text label wrap and give it room
            label = obj.findChild(QLabel, "qt_msgbox_label")
            if label:
                label.setWordWrap(True)
                label.setTextInteractionFlags(Qt.TextSelectableByMouse)
                label.setMinimumWidth(520)

            # Allow the dialog to grow to its content
            obj.setSizeGripEnabled(True)
            obj.setMinimumWidth(560)

            lay = obj.layout()
            if lay:
                lay.setSizeConstraint(QLayout.SetMinimumSize)

            QTimer.singleShot(0, obj.adjustSize)

        return super().eventFilter(obj, event)

def _install_linux_msgbox_fixer():
    global _linux_msgbox_fixer
    if not IS_LINUX:
        return
    app = QApplication.instance()
    if app is None or _linux_msgbox_fixer is not None:
        return
    _linux_msgbox_fixer = _LinuxMessageBoxFixer(app)
    app.installEventFilter(_linux_msgbox_fixer)




def start_combine(self):
    QTimer.singleShot(100, self.combine_files)  # delays execution and avoids UI freeze

#Uncomment for windows build

def resource_path(relative_path: str) -> str:
    if hasattr(sys, "_MEIPASS"):
        # Packaged app
        return os.path.join(sys._MEIPASS, relative_path)
    # Development mode
    base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


#Uncomment for Linux build
"""
def resource_path(relative_path: str) -> str:
    # PyInstaller sets sys._MEIPASS
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
        return os.path.join(base_path, relative_path)

    # Development
    base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)
"""

"""
#Uncomment for MacOS build
def resource_path(relative_path: str) -> str:
   # py2app
    if getattr(sys, "frozen", False):
        base_path = os.path.abspath(
            os.path.join(os.path.dirname(sys.executable), "..", "Resources")
        )
        return os.path.join(base_path, relative_path)
    # Development
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    return os.path.join(base_path, relative_path)
"""


#Fix for figure saving issue on Linux
def _ext_from_filter(name_filter: str) -> str:
    m = re.search(r"\*\.(\w+)", name_filter or "")
    return m.group(1).lower() if m else ""

def pick_export_path(parent, caption: str, default_name: str, filters: str, default_filter: str = None):
    """
    Returns (path, ext).
    Linux uses a QFileDialog instance (non-native) so selectedNameFilter is reliable.
    Windows/macOS keep using getSaveFileName.
    """
    if IS_LINUX:
        dlg = QFileDialog(parent, caption)
        dlg.setAcceptMode(QFileDialog.AcceptSave)
        dlg.setNameFilters(filters.split(";;"))
        if default_filter:
            dlg.selectNameFilter(default_filter)
        dlg.selectFile(default_name)

        # Important for Linux reliability
        dlg.setOption(QFileDialog.DontUseNativeDialog, True)

        if not dlg.exec():
            return "", ""
        path = dlg.selectedFiles()[0]
        chosen_filter = dlg.selectedNameFilter()
    else:
        path, chosen_filter = QFileDialog.getSaveFileName(parent, caption, default_name, filters)
        if not path:
            return "", ""

    ext = os.path.splitext(path)[1].lstrip(".").lower()

    # If user didn’t type an extension, take it from the selected filter
    if not ext:
        ext = _ext_from_filter(chosen_filter) or "png"
        path = f"{path}.{ext}"

    return path, ext


class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=10, height=6, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.updateGeometry()


class DownloaderImportWorker(QObject):
    progress_text = Signal(str)
    progress_range = Signal(int, int)
    progress_value = Signal(int)
    finished = Signal(object)
    failed = Signal(str)

    def __init__(self, urls):
        super().__init__()
        self.urls = list(urls or [])

    @Slot()
    def run(self):
        if not self.urls:
            self.failed.emit("No files were received from the downloader.")
            return

        local_files = []
        self.progress_text.emit("Downloading selected FITS files...")
        self.progress_range.emit(0, len(self.urls))
        self.progress_value.emit(0)

        try:
            for i, url in enumerate(self.urls, start=1):
                r = requests.get(url, timeout=25)
                r.raise_for_status()

                original_name = str(url).split("/")[-1] or f"import_{i}.fit"
                temp_dir = tempfile.gettempdir()
                local_path = os.path.join(temp_dir, original_name)

                with open(local_path, "wb") as f:
                    f.write(r.content)

                local_files.append(local_path)
                self.progress_value.emit(i)
        except Exception as e:
            self.failed.emit(f"Failed to download one or more FITS files:\n{e}")
            return

        if len(local_files) == 1:
            try:
                res = load_callisto_fits(local_files[0], memmap=False)
                payload = {
                    "kind": "single",
                    "filename": os.path.basename(local_files[0]),
                    "source_path": local_files[0],
                    "data": res.data,
                    "freqs": res.freqs,
                    "time": res.time,
                    "header0": res.header0.copy(),
                    "ut_start_sec": extract_ut_start_sec(res.header0),
                }
                self.finished.emit(payload)
            except Exception as e:
                self.failed.emit(f"Could not load FITS file:\n{e}")
            return

        from src.Backend.burst_processor import (
            are_time_combinable,
            are_frequency_combinable,
            combine_time,
            combine_frequency,
        )

        try:
            self.progress_text.emit("Checking file compatibility...")
            if are_time_combinable(local_files):
                self.progress_text.emit("Combining files (time mode)...")
                combined = combine_time(local_files)
                self.finished.emit({"kind": "combined", "combined": combined})
                return

            if are_frequency_combinable(local_files):
                self.progress_text.emit("Combining files (frequency mode)...")
                combined = combine_frequency(local_files)
                self.finished.emit({"kind": "combined", "combined": combined})
                return

            self.finished.emit({"kind": "invalid"})
        except Exception as e:
            self.failed.emit(f"An error occurred while combining files:\n{e}")


class MainWindow(QMainWindow):
    DB_SCALE = 2500.0 / 255.0 / 25.4
    HW_DEFAULT_TICK_FONT_PX = 14
    HW_DEFAULT_AXIS_FONT_PX = 16
    HW_DEFAULT_TITLE_FONT_PX = 22

    def __init__(self, theme=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.theme = QApplication.instance().property("theme_manager") if QApplication.instance() else None
        if self.theme and hasattr(self.theme, "themeChanged"):
            self.theme.themeChanged.connect(self._on_theme_changed)

        #Linux Messagebox Fix
        _install_linux_msgbox_fixer()

        self.setWindowTitle("e-CALLISTO FITS Analyzer 2.0")
        #self.resize(1000, 700)
        self.setMinimumSize(1000, 700)

        self.use_utc = False
        self.ut_start_sec = None
        self.use_db = False  # False = Digits (default), True = dB

        # --- Undo / Redo ---
        self._undo_stack = []
        self._redo_stack = []
        # Limit only applies to full data-state snapshots (zoom/pan history is lightweight)
        self._max_undo = 30  # prevent memory blow-up
        self._max_history_entries = 5000  # guardrail for view-history spam

        # --- View (zoom/pan) "home" + history ---
        self._home_view = None
        self._pan_start_view = None

        # --- Graph Properties (non-colormap) ---
        self.graph_title_override = ""  # empty = use context-aware default title
        self.graph_font_family = ""  # empty = use Matplotlib default

        self.tick_font_px = 11
        self.axis_label_font_px = 12
        self.title_font_px = 14
        self._hw_default_font_sizes_active = True

        self._colorbar_label_text = ""

        self.current_cmap_name = "Custom"
        self.lasso_active = False

        self.noise_vmin = None
        self.noise_vmax = None

        self.current_display_data = None

        # --- Graph Properties: style flags ---
        self.title_bold = False
        self.title_italic = False

        self.axis_bold = False
        self.axis_italic = False

        self.ticks_bold = False
        self.ticks_italic = False

        self.remove_titles = False

        self._build_toolbar()
        self._refresh_toolbar_icons()

        # Debounce timer for smooth slider updates
        self.noise_smooth_timer = QTimer()
        self.noise_smooth_timer.setInterval(12)
        self.noise_smooth_timer.setSingleShot(True)
        self.noise_smooth_timer.timeout.connect(self.update_noise_live)
        self.noise_commit_timer = QTimer()
        self.noise_commit_timer.setInterval(80)
        self.noise_commit_timer.setSingleShot(True)
        self.noise_commit_timer.timeout.connect(self._commit_noise_live_update)

        self._noise_undo_pending = False
        self._noise_slider_drag_active = False
        self._noise_preview_active = False
        self._noise_base_data = None
        self._noise_base_source_id = None

        self._import_thread = None
        self._import_worker = None
        self._import_progress_dialog = None

        # Canvas
        self.canvas = MplCanvas(self, width=10, height=6)
        style_axes(self.canvas.ax)

        # Hardware-accelerated plotting canvas (full graph area when enabled)
        self.accel_canvas = AcceleratedPlotWidget(self)
        self.use_hw_live_preview = bool(self.accel_canvas.is_available)
        if self.accel_canvas.is_available:
            self.accel_canvas.mousePositionChanged.connect(self.on_accel_mouse_motion_status)
            self.accel_canvas.viewInteractionFinished.connect(self._on_accel_view_interaction_finished)
            self.accel_canvas.rectZoomFinished.connect(self._on_accel_rect_zoom_finished)
            self.accel_canvas.lassoFinished.connect(self._on_accel_lasso_finished)
            self.accel_canvas.driftPointAdded.connect(self._on_accel_drift_point_added)
            self.accel_canvas.driftCaptureFinished.connect(self._on_accel_drift_capture_finished)

        self.canvas.mpl_connect("scroll_event", self.on_scroll_zoom)
        self._cid_press = self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self._cid_release = self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self._cid_motion = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
        self._cid_motion_status = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_motion_status)

        self._apply_mpl_theme()
        self._apply_accel_theme()

        self._panning = False
        self._last_pan_xy = None

        # --- Navigation state (pan/scroll lock + rectangle zoom) ---
        self.nav_locked = False  # False = normal pan + scroll zoom
        self.rect_zoom_active = False  # True only while RectangleSelector is active
        self._rect_selector = None  # RectangleSelector instance

        # Colorbar
        self.current_colorbar = None
        self.current_cax = None

        # Statusbar
        self.setStatusBar(QStatusBar())
        # Permanent label on right side for cursor coordinates
        self.cursor_label = QLabel("")
        self.cursor_label.setStyleSheet("padding-right: 8px;")
        self.statusBar().addPermanentWidget(self.cursor_label)

        # =========================
        # LEFT SIDEBAR (CROSS-PLATFORM SAFE)
        # Put this whole block where you currently build:
        #   slider_group, units_group_box, graph_group, main_layout, container
        #
        # Required imports (add if missing):
        #   from PySide6.QtWidgets import QScrollArea, QFrame
        # =========================

        # -------------------------
        # Noise clipping sliders
        # -------------------------
        self.lower_slider = QSlider(Qt.Horizontal)
        self.lower_slider.setRange(-100, 100)
        self.lower_slider.setValue(0)

        self.upper_slider = QSlider(Qt.Horizontal)
        self.upper_slider.setRange(-100, 100)
        self.upper_slider.setValue(0)

        slider_group = QGroupBox("Noise Clipping Thresholds")
        slider_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        slider_layout = QVBoxLayout(slider_group)
        slider_layout.setContentsMargins(10, 10, 10, 10)
        slider_layout.setSpacing(6)

        lbl_low = QLabel("Lower Threshold")
        lbl_low.setAlignment(Qt.AlignLeft)
        slider_layout.addWidget(lbl_low)
        slider_layout.addWidget(self.lower_slider)

        lbl_high = QLabel("Upper Threshold")
        lbl_high.setAlignment(Qt.AlignLeft)
        slider_layout.addWidget(lbl_high)
        slider_layout.addWidget(self.upper_slider)

        # -------------------------
        # Units Group (Intensity + Time in one row)
        # -------------------------
        self.units_group_box = QGroupBox("Units")
        self.units_group_box.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        units_layout = QVBoxLayout(self.units_group_box)
        units_layout.setContentsMargins(10, 10, 10, 10)
        units_layout.setSpacing(8)

        # ---- Horizontal container for Intensity + Time ----
        units_row = QHBoxLayout()
        units_row.setSpacing(20)

        # ===== Intensity column =====
        intensity_col = QVBoxLayout()
        intensity_col.setSpacing(8)

        intensity_label = QLabel("Intensity")
        intensity_label.setProperty("section", True)

        self.units_digits_radio = QRadioButton("Digits")
        self.units_db_radio = QRadioButton("dB")
        self.units_digits_radio.setChecked(True)

        self.units_group = QButtonGroup(self)
        self.units_group.addButton(self.units_digits_radio)
        self.units_group.addButton(self.units_db_radio)

        intensity_col.addWidget(intensity_label)
        intensity_col.addWidget(self.units_digits_radio)
        intensity_col.addWidget(self.units_db_radio)
        intensity_col.addStretch(1)

        # ===== Time column =====
        time_col = QVBoxLayout()
        time_col.setSpacing(6)

        time_label = QLabel("Time")
        time_label.setProperty("section", True)

        self.time_sec_radio = QRadioButton("Seconds")
        self.time_ut_radio = QRadioButton("UT")
        self.time_sec_radio.setChecked(True)

        self.time_group = QButtonGroup(self)
        self.time_group.addButton(self.time_sec_radio)
        self.time_group.addButton(self.time_ut_radio)

        time_col.addWidget(time_label)
        time_col.addWidget(self.time_sec_radio)
        time_col.addWidget(self.time_ut_radio)
        time_col.addStretch(1)

        # ---- Add both columns to the row ----
        units_row.addLayout(intensity_col, 1)
        units_row.addLayout(time_col, 1)

        # ---- Add row to Units group ----
        units_layout.addLayout(units_row)
        units_layout.addStretch(1)

        # -------------------------
        # Graph Properties Group
        # -------------------------
        def _section_label(text: str) -> QLabel:
            lbl = QLabel(text)
            lbl.setObjectName("SectionLabel")
            return lbl

        def _spin_row(label_text: str, spin: QSpinBox) -> QWidget:
            w = QWidget()
            row = QHBoxLayout(w)
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)

            label = QLabel(label_text)
            label.setWordWrap(False)
            label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

            spin.setMinimumWidth(90)
            spin.setMaximumWidth(110)
            spin.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            row.addWidget(label, 1)
            row.addWidget(spin, 0, Qt.AlignRight)
            return w

        def _style_row(label_text: str, cb_bold: QCheckBox, cb_italic: QCheckBox) -> QWidget:
            w = QWidget()
            row = QHBoxLayout(w)
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)

            label = QLabel(label_text)
            label.setWordWrap(False)
            label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

            row.addWidget(label, 1)
            row.addWidget(cb_bold, 0)
            row.addWidget(cb_italic, 0)
            return w

        self.graph_group = QGroupBox("Graph Properties")
        self.graph_group.setEnabled(False)
        self.graph_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        graph_layout = QVBoxLayout(self.graph_group)
        graph_layout.setContentsMargins(10, 10, 10, 10)
        graph_layout.setSpacing(6)

        # Appearance
        graph_layout.addWidget(_section_label("Appearance"))

        self.cmap_combo = QComboBox()
        self.cmap_combo.addItems([
            "Custom", "viridis", "plasma", "inferno", "magma",
            "cividis", "turbo", "RdYlBu", "jet", "cubehelix",
        ])
        graph_layout.addWidget(QLabel("Colormap"))
        graph_layout.addWidget(self.cmap_combo)

        self.font_combo = QComboBox()
        self.font_combo.addItem("Default")
        for f in sorted(QFontDatabase.families()):
            self.font_combo.addItem(f)

        graph_layout.addWidget(QLabel("Font family"))
        graph_layout.addWidget(self.font_combo)

        # Text
        graph_layout.addWidget(_section_label("Text"))

        self.title_edit = QLineEdit()
        self.title_edit.setPlaceholderText("Custom title (leave empty for default)")
        self.title_edit.setMinimumHeight(20)
        graph_layout.addWidget(QLabel("Graph title"))
        graph_layout.addWidget(self.title_edit)

        self.remove_titles_chk = QCheckBox("Remove Titles")
        graph_layout.addWidget(self.remove_titles_chk)

        # Font sizes
        graph_layout.addWidget(_section_label("Font sizes"))

        self.tick_font_spin = QSpinBox()
        self.tick_font_spin.setRange(6, 60)
        self.tick_font_spin.setValue(self.tick_font_px)
        graph_layout.addWidget(_spin_row("Tick labels (px)", self.tick_font_spin))

        self.axis_font_spin = QSpinBox()
        self.axis_font_spin.setRange(6, 60)
        self.axis_font_spin.setValue(self.axis_label_font_px)
        graph_layout.addWidget(_spin_row("Axis labels (px)", self.axis_font_spin))

        self.title_font_spin = QSpinBox()
        self.title_font_spin.setRange(6, 80)
        self.title_font_spin.setValue(self.title_font_px)
        graph_layout.addWidget(_spin_row("Title (px)", self.title_font_spin))

        # Text style
        graph_layout.addWidget(_section_label("Text style"))

        self.title_bold_chk = QCheckBox("Bold")
        self.title_italic_chk = QCheckBox("Italic")
        graph_layout.addWidget(_style_row("Title", self.title_bold_chk, self.title_italic_chk))

        self.axis_bold_chk = QCheckBox("Bold")
        self.axis_italic_chk = QCheckBox("Italic")
        graph_layout.addWidget(_style_row("Axis labels", self.axis_bold_chk, self.axis_italic_chk))

        self.ticks_bold_chk = QCheckBox("Bold")
        self.ticks_italic_chk = QCheckBox("Italic")
        graph_layout.addWidget(_style_row("Tick labels", self.ticks_bold_chk, self.ticks_italic_chk))

        graph_layout.addStretch(1)

        # -------------------------
        # Build the LEFT PANEL as a widget, then put it in a ScrollArea
        # This is the key fix for Windows (no overlaps, no clipping).
        # -------------------------
        side_panel_widget = QWidget()
        side_panel_layout = QVBoxLayout(side_panel_widget)
        side_panel_layout.setContentsMargins(8, 8, 8, 8)
        side_panel_layout.setSpacing(10)

        side_panel_layout.addWidget(slider_group)
        side_panel_layout.addWidget(self.units_group_box)
        side_panel_layout.addWidget(self.graph_group)
        side_panel_layout.addStretch(1)

        # Consistent width for all groups (better on Windows DPI scaling)
        SIDEBAR_W = 250
        slider_group.setMaximumWidth(SIDEBAR_W)
        self.units_group_box.setMaximumWidth(SIDEBAR_W)
        self.graph_group.setMaximumWidth(SIDEBAR_W)

        self.side_scroll = QScrollArea()
        self.side_scroll.setWidgetResizable(True)
        self.side_scroll.setFrameShape(QFrame.NoFrame)
        self.side_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.side_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.side_scroll.setMinimumWidth(SIDEBAR_W + 16)  # room for scrollbar
        self.side_scroll.setMaximumWidth(SIDEBAR_W + 28)
        self.side_scroll.setWidget(side_panel_widget)

        self.sidebar_toggle_btn = QPushButton("◀")
        self.sidebar_toggle_btn.setObjectName("SidebarToggleButton")
        self.sidebar_toggle_btn.setToolTip("Collapse sidebar")
        self.sidebar_toggle_btn.setFixedSize(12, 22)
        self.sidebar_toggle_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.sidebar_toggle_btn.setFocusPolicy(Qt.NoFocus)
        self.sidebar_toggle_btn.setStyleSheet("QPushButton#SidebarToggleButton { font-size: 10px; padding: 0px; }")
        self.sidebar_toggle_btn.clicked.connect(self.toggle_left_sidebar)

        self.sidebar_toggle_strip = QWidget()
        self.sidebar_toggle_strip.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        self.sidebar_toggle_strip.setFixedWidth(12)
        sidebar_toggle_layout = QVBoxLayout(self.sidebar_toggle_strip)
        sidebar_toggle_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_toggle_layout.setSpacing(0)
        sidebar_toggle_layout.addStretch(1)
        sidebar_toggle_layout.addWidget(self.sidebar_toggle_btn, 0, Qt.AlignHCenter)
        sidebar_toggle_layout.addStretch(1)
        self._sidebar_collapsed = False

        # -------------------------
        # Style (safe sizes, no tiny max-heights)
        # -------------------------
        sidebar_style = """
        QGroupBox {
            font-weight: bold;
        }
        QLabel {
            font-size: 12px;
        }
        QLabel[section="true"] {
            font-weight: bold;
            color: #444;
            margin-top: 6px;
        }
        QLabel#SectionLabel {
            font-weight: bold;
            color: #555;
            margin-top: 8px;
            margin-bottom: 4px;
        }
        
        QLineEdit, QComboBox {
            min-height: 24px;
            padding: 4px 6px;
            font-size: 12px;
        }

        QSpinBox {
            min-height: 32px;      /* REQUIRED for Windows */
            min-width: 90px;
            padding-left: 6px;
            font-size: 12px;
        }

        QCheckBox {
            spacing: 6px;
            font-size: 12px;
            margin-top: 10px;
            margin-bottom: 10px;
            margin-right: 10px;
            
        }
        """
        slider_group.setStyleSheet(sidebar_style)
        self.units_group_box.setStyleSheet(sidebar_style)
        self.graph_group.setStyleSheet(sidebar_style)

        # -------------------------
        # Main layout with scrollable sidebar + canvas
        # -------------------------
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(0)

        self.plot_stack_host = QWidget()
        self.plot_stack = QStackedLayout(self.plot_stack_host)
        self.plot_stack.setContentsMargins(0, 0, 0, 0)
        self.plot_stack.setSpacing(0)
        self.plot_stack.addWidget(self.canvas)
        self.plot_stack.addWidget(self.accel_canvas)
        self.plot_stack.setCurrentWidget(self.accel_canvas if self.use_hw_live_preview else self.canvas)

        main_layout.addWidget(self.side_scroll, 0)
        main_layout.addWidget(self.sidebar_toggle_strip, 0)
        main_layout.addWidget(self.plot_stack_host, 1)
        self._main_layout = main_layout

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)
        self._set_sidebar_collapsed(False)

        # ----- Menu Bar -----
        menubar = self.menuBar()

        # File Menu
        file_menu = menubar.addMenu("File")

        # --- Open ---
        self.open_action = QAction("Open", self)
        file_menu.addAction(self.open_action)

        file_menu.addSeparator()

        # --- Project (session) ---
        self.open_project_action = QAction("Open Project...", self)
        self.open_project_action.setShortcut("Ctrl+Shift+O")
        file_menu.addAction(self.open_project_action)

        self.save_project_action = QAction("Save Project", self)
        self.save_project_action.setShortcut("Ctrl+S")
        file_menu.addAction(self.save_project_action)

        self.save_project_as_action = QAction("Save Project As...", self)
        self.save_project_as_action.setShortcut("Ctrl+Shift+S")
        file_menu.addAction(self.save_project_as_action)

        file_menu.addSeparator()

        # --- Save As (disabled for main window) ---
        self.save_action = QAction("Save As", self)
        self.save_action.setEnabled(False)
        file_menu.addAction(self.save_action)

        # --- Export As submenu ---
        export_menu = QMenu("Export As", self)
        file_menu.addMenu(export_menu)

        self.export_figure_action = QAction("Export Figure", self)
        export_menu.addAction(self.export_figure_action)
        self.export_figure_action.triggered.connect(self.export_figure)

        self.export_fits_action = QAction("Export to FIT", self)
        export_menu.addAction(self.export_fits_action)
        self.export_fits_action.triggered.connect(self.export_to_fits)

        # Edit Menu
        edit_menu = menubar.addMenu("Edit")

        self.undo_action = QAction("Undo", self)
        self.undo_action.setShortcut("Ctrl+Z")
        edit_menu.addAction(self.undo_action)

        self.redo_action = QAction("Redo", self)
        self.redo_action.setShortcut("Ctrl+Shift+Z")
        edit_menu.addAction(self.redo_action)

        edit_menu.addSeparator()

        self.reset_to_raw_action = QAction("Reset to Raw", self)
        self.reset_to_raw_action.setEnabled(False)
        edit_menu.addAction(self.reset_to_raw_action)
        self.reset_to_raw_action.triggered.connect(self.reset_to_raw)

        edit_menu.addSeparator()

        reset_action = QAction("Reset All", self)
        edit_menu.addAction(reset_action)
        reset_action.triggered.connect(self.reset_all)

        self.undo_action.triggered.connect(self.undo)
        self.redo_action.triggered.connect(self.redo)

        # Download Menu
        download_menu = menubar.addMenu("Download")

        # --- Launch Downloader ---
        launch_downloader_action = QAction("Launch FITS Downloader", self)
        download_menu.addAction(launch_downloader_action)
        launch_downloader_action.triggered.connect(self.launch_downloader)

        # Solar Events Menu
        solar_events_menu = menubar.addMenu("Solar Events")

        # CMEs submenu
        cmes_submenu = solar_events_menu.addMenu("CMEs")
        soho_lasco_action = QAction("SOHO/LASCO CME Catalog", self)
        soho_lasco_action.triggered.connect(self.open_cme_viewer)
        cmes_submenu.addAction(soho_lasco_action)

        # Flares submenu
        flares_submenu = solar_events_menu.addMenu("Flares")
        goes_flux_action = QAction("GOES X-Ray Flux", self)
        goes_flux_action.triggered.connect(self.open_goes_xrs_window)
        flares_submenu.addAction(goes_flux_action)

        # Radio submenu
        radio_submenu = solar_events_menu.addMenu("Radio Bursts")
        radio_action = QAction("e-CALLISTO", self)
        radio_action.triggered.connect(self.launch_downloader)
        radio_submenu.addAction(radio_action)

        # FITS View Menu
        fits_view_menu = menubar.addMenu("FITS View")
        self.view_fits_header_action = QAction("View FITS Header", self)
        self.view_fits_header_action.setEnabled(False)
        fits_view_menu.addAction(self.view_fits_header_action)
        self.view_fits_header_action.triggered.connect(self.open_fits_header_viewer)


        # View Menu
        view_menu = menubar.addMenu("View")
        theme_menu = view_menu.addMenu("Theme")

        self.theme_action_system = QAction("System", self, checkable=True)
        self.theme_action_light = QAction("Light", self, checkable=True)
        self.theme_action_dark = QAction("Dark", self, checkable=True)

        theme_group = QActionGroup(self)
        theme_group.setExclusive(True)
        for a in (self.theme_action_system, self.theme_action_light, self.theme_action_dark):
            theme_group.addAction(a)
            theme_menu.addAction(a)

        processing_menu = menubar.addMenu("Processing")
        hw_menu = processing_menu.addMenu("Hardware Acceleration")
        self.hw_live_preview_action = QAction("Enable", self, checkable=True)
        self.hw_live_preview_action.setChecked(self.use_hw_live_preview)
        self.hw_live_preview_action.setEnabled(bool(self.accel_canvas.is_available))
        hw_menu.addAction(self.hw_live_preview_action)

        # Set initial checks from saved mode
        if self.theme:
            m = self.theme.mode()
            self.theme_action_system.setChecked(m == "system")
            self.theme_action_light.setChecked(m == "light")
            self.theme_action_dark.setChecked(m == "dark")

        # Connect changes
        self.theme_action_system.triggered.connect(lambda: self.theme.set_mode("system"))
        self.theme_action_light.triggered.connect(lambda: self.theme.set_mode("light"))
        self.theme_action_dark.triggered.connect(lambda: self.theme.set_mode("dark"))
        self.hw_live_preview_action.toggled.connect(self.set_hardware_live_preview_enabled)

        # About Menu
        about_menu = menubar.addMenu("About")
        about_action = QAction("About", self)
        about_action.setMenuRole(QAction.NoRole)
        about_menu.addAction(about_action)
        about_action.triggered.connect(self.show_about_dialog)

        # (OPTIONAL) Connect them later like:
        # open_action.triggered.connect(self.open_file)

        self.setCentralWidget(container)

        # Signals
        self.lower_slider.valueChanged.connect(self.schedule_noise_update)
        self.upper_slider.valueChanged.connect(self.schedule_noise_update)
        self.lower_slider.sliderPressed.connect(self._on_noise_slider_pressed)
        self.upper_slider.sliderPressed.connect(self._on_noise_slider_pressed)
        self.lower_slider.sliderReleased.connect(self._on_noise_slider_released)
        self.upper_slider.sliderReleased.connect(self._on_noise_slider_released)
        self.cmap_combo.currentTextChanged.connect(self.change_cmap)
        self.open_action.triggered.connect(self.load_file)
        self.open_project_action.triggered.connect(self.open_project)
        self.save_project_action.triggered.connect(self.save_project)
        self.save_project_as_action.triggered.connect(self.save_project_as)
        self.units_digits_radio.toggled.connect(
            lambda checked: checked and self.set_units_mode(False)
        )
        self.units_db_radio.toggled.connect(
            lambda checked: checked and self.set_units_mode(True)
        )

        self.time_sec_radio.toggled.connect(
            lambda checked: checked and self.set_axis_to_seconds()
        )
        self.time_ut_radio.toggled.connect(
            lambda checked: checked and self.set_axis_to_utc()
        )

        # Keep existing colormap live behavior
        self.cmap_combo.currentTextChanged.connect(self.change_cmap)

        # Real-time graph properties (non-colormap)
        self.title_edit.textChanged.connect(self.apply_graph_properties_live)
        self.font_combo.currentTextChanged.connect(self.apply_graph_properties_live)
        self.tick_font_spin.valueChanged.connect(self._on_tick_font_spin_changed)
        self.axis_font_spin.valueChanged.connect(self._on_axis_font_spin_changed)
        self.title_font_spin.valueChanged.connect(self._on_title_font_spin_changed)

        # Real-time style toggles
        self.remove_titles_chk.toggled.connect(self.apply_graph_properties_live)

        self.title_bold_chk.toggled.connect(self.apply_graph_properties_live)
        self.title_italic_chk.toggled.connect(self.apply_graph_properties_live)

        self.axis_bold_chk.toggled.connect(self.apply_graph_properties_live)
        self.axis_italic_chk.toggled.connect(self.apply_graph_properties_live)

        self.ticks_bold_chk.toggled.connect(self.apply_graph_properties_live)
        self.ticks_italic_chk.toggled.connect(self.apply_graph_properties_live)

        # Data placeholders
        self.raw_data = None
        self.freqs = None
        self.time = None
        self.filename = ""
        self.current_plot_type = "Raw"  # or "NoiseReduced" or "Isolated"

        # ----- Project/session save state -----
        self._project_path = None
        self._project_dirty = False
        self._loading_project = False
        self._max_intensity_state = None  # populated after Max-Intensity dialog closes

        # FITS export metadata
        self._fits_header0 = None  # primary header template
        self._fits_source_path = None  # original single-file path (if any)

        self._is_combined = False
        self._combined_mode = None  # "time" or "frequency"
        self._combined_sources = []  # list of source files used to combine

        self.lasso = None
        self.lasso_mask = None
        self.noise_reduced_data = None

        self.setStyleSheet("""
            QLabel {
                font-size: 13px;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
            }
        """)

        self.noise_reduced_original = None  # backup before lasso

        # Ensure project actions reflect initial state
        self._sync_project_actions()
        self.set_hardware_live_preview_enabled(self.use_hw_live_preview)

    def _is_dark_ui(self) -> bool:
        # Prefer theme manager if available
        theme = getattr(self, "theme", None)
        if theme is not None:
            flag = getattr(theme, "is_dark", None)
            try:
                if callable(flag):
                    return bool(flag())
                return bool(flag)
            except Exception:
                pass

        # Fallback: infer from palette
        app = QApplication.instance()
        if not app:
            return False
        return app.palette().color(QPalette.Window).lightness() < 128

    def _icon(self, filename: str) -> QIcon:
        folder = "icons_dark" if self._is_dark_ui() else "icons"

        rels = [
            os.path.join("assets", folder, filename),
            os.path.join("assets", "icons", filename),  # fallback to light icons
        ]

        bases = []

        if getattr(sys, "frozen", False):
            bases.append(os.path.abspath(os.path.join(os.path.dirname(sys.executable), "..", "Resources")))
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            bases.append(os.path.abspath(meipass))

        here = os.path.abspath(os.path.dirname(__file__))
        bases.extend([
            os.path.abspath(os.path.join(here, "..", "..")),  # project root (src/UI -> root)
            os.path.abspath(os.path.join(here, "..", "..", "..")),  # one more up (safe)
            os.path.abspath(os.getcwd()),
            os.path.abspath(os.path.join(os.getcwd(), "..")),
        ])

        seen = set()
        uniq_bases = []
        for b in bases:
            if b and b not in seen:
                seen.add(b)
                uniq_bases.append(b)

        for b in uniq_bases:
            for rel in rels:
                p = os.path.normpath(os.path.join(b, rel))
                if os.path.exists(p):
                    icon = self._load_icon_file(p)
                    if not icon.isNull():
                        return icon

        for rel in rels:
            try:
                p = resource_path(rel)
                if os.path.exists(p):
                    icon = self._load_icon_file(p)
                    if not icon.isNull():
                        return icon
            except Exception:
                pass

        # Avoid spamming the console with the same missing icon message
        if not hasattr(self, "_missing_icons"):
            self._missing_icons = set()
        if filename not in self._missing_icons:
            self._missing_icons.add(filename)
            print(f"⚠️ Icon not found: {filename}")

        return QIcon()

    def _load_icon_file(self, path: str) -> QIcon:
        icon = QIcon(path)
        if not icon.isNull():
            return icon

        if not str(path).lower().endswith(".svg"):
            return QIcon()

        try:
            from PySide6.QtSvg import QSvgRenderer
        except Exception:
            return QIcon()

        try:
            renderer = QSvgRenderer(path)
            if not renderer.isValid():
                return QIcon()

            size = renderer.defaultSize()
            w = max(48, int(size.width())) if size.isValid() else 64
            h = max(48, int(size.height())) if size.isValid() else 64

            pixmap = QPixmap(w, h)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            renderer.render(painter)
            painter.end()

            icon = QIcon(pixmap)
            return icon if not icon.isNull() else QIcon()
        except Exception:
            return QIcon()

    def _build_toolbar(self):
        tb = QToolBar("Main Toolbar", self)
        tb.setMovable(False)
        tb.setIconSize(QSize(36, 36))
        self.addToolBar(tb)

        # --- Actions (toolbar) ---
        self.tb_open = QAction(self._icon("open.svg"), "Open / Load", self)
        self.tb_open.setShortcut("Ctrl+O")
        self.tb_open.triggered.connect(self.load_file)
        tb.addAction(self.tb_open)

        self.tb_download = QAction(self._icon("download.svg"), "Download", self)
        self.tb_download.triggered.connect(self.launch_downloader)
        tb.addAction(self.tb_download)

        self.tb_export = QAction(self._icon("export.svg"), "Export", self)
        self.tb_export.setShortcut("Ctrl+E")
        self.tb_export.triggered.connect(self.export_figure)
        tb.addAction(self.tb_export)

        self.tb_export_fits = QAction(self._icon("export_fits.svg"), "Export as FITS", self)
        self.tb_export_fits.setShortcut("Ctrl+F")
        self.tb_export_fits.triggered.connect(self.export_to_fits)
        tb.addAction(self.tb_export_fits)

        tb.addSeparator()

        self.tb_undo = QAction(self._icon("undo.svg"), "Undo", self)
        self.tb_undo.triggered.connect(self.undo)
        tb.addAction(self.tb_undo)

        self.tb_redo = QAction(self._icon("redo.svg"), "Redo", self)
        self.tb_redo.triggered.connect(self.redo)
        tb.addAction(self.tb_redo)

        tb.addSeparator()

        self.tb_drift = QAction(self._icon("drift.svg"), "Estimate Drift Rate", self)
        self.tb_drift.triggered.connect(self.activate_drift_tool)
        tb.addAction(self.tb_drift)

        self.tb_isolate = QAction(self._icon("isolate.svg"), "Isolate Burst", self)
        self.tb_isolate.triggered.connect(self.activate_lasso)
        tb.addAction(self.tb_isolate)

        self.tb_max = QAction(self._icon("max.svg"), "Plot Maximum Intensities", self)
        self.tb_max.triggered.connect(self.plot_max_intensities)
        tb.addAction(self.tb_max)

        self.tb_zoom = QAction(self._icon("zoom.svg"), "Rectangular Zooming", self)
        self.tb_zoom.triggered.connect(self.rectangular_zoom)
        tb.addAction(self.tb_zoom)

        self.tb_lock = QAction(self._icon("lock.svg"), "Lock zooming and panning", self)
        self.tb_lock.triggered.connect(self.lock)
        tb.addAction(self.tb_lock)

        self.tb_unlock = QAction(self._icon("unlock.svg"), "Unlock zooming and panning", self)
        self.tb_unlock.triggered.connect(self.unlock)
        tb.addAction(self.tb_unlock)

        tb.addSeparator()

        self.tb_reset_sel = QAction(self._icon("reset_selection.svg"), "Reset Selection", self)
        self.tb_reset_sel.triggered.connect(self.reset_selection)
        tb.addAction(self.tb_reset_sel)

        self.tb_reset_all = QAction(self._icon("reset_all.svg"), "Reset All", self)
        self.tb_reset_all.triggered.connect(self.reset_all)
        tb.addAction(self.tb_reset_all)

        # Initial enable/disable states
        self._sync_toolbar_enabled_states()
        self._sync_nav_actions()

    def _sync_toolbar_enabled_states(self):
        has_file = getattr(self, "raw_data", None) is not None
        has_noise = getattr(self, "noise_reduced_data", None) is not None
        has_undo = len(getattr(self, "_undo_stack", [])) > 0
        has_redo = len(getattr(self, "_redo_stack", [])) > 0
        filename = getattr(self, "filename", "")

        can_reset_view = False
        try:
            cur_view = self._capture_view()
            home_view = getattr(self, "_home_view", None)
            if cur_view and home_view:
                can_reset_view = not self._views_close(cur_view, home_view)
        except Exception:
            can_reset_view = False

        # Always allowed
        self.tb_open.setEnabled(True)
        self.tb_download.setEnabled(True)

        # Needs a plot / filename
        self.tb_export.setEnabled(bool(filename))
        self.tb_export_fits.setEnabled(bool(filename))

        # Undo/redo availability
        self.tb_undo.setEnabled(has_undo)
        self.tb_redo.setEnabled(has_redo)
        act = getattr(self, "undo_action", None)
        if act is not None:
            act.setEnabled(has_undo)
        act = getattr(self, "redo_action", None)
        if act is not None:
            act.setEnabled(has_redo)
        act = getattr(self, "reset_to_raw_action", None)
        if act is not None:
            act.setEnabled(has_file)

        # Tools that require processed data
        self.tb_drift.setEnabled(has_noise)
        self.tb_isolate.setEnabled(has_noise)
        self.tb_max.setEnabled(has_noise)
        self.tb_reset_sel.setEnabled(has_noise or can_reset_view)
        self.tb_reset_all.setEnabled(has_file)
        self._sync_fits_view_actions()
        self._sync_nav_actions()

    def _show_plot_canvas(self):
        stack = getattr(self, "plot_stack", None)
        if stack is not None:
            stack.setCurrentWidget(self.canvas)
        self._noise_preview_active = False

    def _hardware_mode_enabled(self) -> bool:
        return bool(
            getattr(self, "use_hw_live_preview", False)
            and getattr(self, "accel_canvas", None) is not None
            and self.accel_canvas.is_available
        )

    def _show_accel_canvas(self):
        if not self._hardware_mode_enabled():
            return False
        accel = getattr(self, "accel_canvas", None)
        stack = getattr(self, "plot_stack", None)
        if accel is None or stack is None or not bool(accel.is_available):
            return False
        stack.setCurrentWidget(accel)
        self._noise_preview_active = True
        return True

    def _apply_accel_theme(self):
        accel = getattr(self, "accel_canvas", None)
        if accel is None or not bool(accel.is_available):
            return
        accel.set_dark(self._is_dark_ui())
        accel.set_time_mode(self.use_utc, self.ut_start_sec)

    def set_hardware_live_preview_enabled(self, enabled: bool):
        mpl_view = None
        try:
            ax = self.canvas.ax
            if ax is not None and ax.images is not None and len(ax.images) > 0:
                mpl_view = {"xlim": ax.get_xlim(), "ylim": ax.get_ylim()}
        except Exception:
            mpl_view = None

        accel_available = bool(getattr(self, "accel_canvas", None) and self.accel_canvas.is_available)
        self.use_hw_live_preview = bool(enabled) and accel_available

        act = getattr(self, "hw_live_preview_action", None)
        if act is not None and act.isChecked() != self.use_hw_live_preview:
            was_blocked = act.blockSignals(True)
            act.setChecked(self.use_hw_live_preview)
            act.blockSignals(was_blocked)

        if not self.use_hw_live_preview:
            self.noise_commit_timer.stop()
            self._noise_undo_pending = False
            self._noise_slider_drag_active = False
            self._sync_mpl_view_from_accel()
            self._show_plot_canvas()
            self.accel_canvas.stop_interaction_capture()
        else:
            self._apply_accel_theme()
            self.accel_canvas.set_navigation_locked(self.nav_locked)
            self._refresh_accel_plot(view=mpl_view, preserve_view=False)
            self._show_accel_canvas()
        self._sync_toolbar_enabled_states()

    def _sync_mpl_view_from_accel(self):
        if not self._hardware_mode_enabled():
            return
        try:
            view = self.accel_canvas.get_view()
            if not view:
                return
            self.canvas.ax.set_xlim(view["xlim"])
            self.canvas.ax.set_ylim(view["ylim"])
            self.canvas.draw_idle()
        except Exception:
            pass

    def _refresh_accel_plot(self, data=None, title=None, view=None, preserve_view=True):
        if not bool(getattr(self, "accel_canvas", None) and self.accel_canvas.is_available):
            return False
        if self.time is None or self.freqs is None or len(self.time) == 0 or len(self.freqs) == 0:
            return False

        if data is None:
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
        if data is None:
            return False

        if view is None and preserve_view:
            view = self.accel_canvas.get_view()

        display_data = self._intensity_for_display(data)
        self.current_display_data = display_data

        plot_type = self._normalize_plot_type(title if title is not None else self.current_plot_type)
        if self.remove_titles:
            plot_title = ""
            x_label = ""
            y_label = ""
        else:
            plot_title = self.graph_title_override or self._default_graph_title(plot_type)
            x_label = "Time [UT]" if (self.use_utc and self.ut_start_sec is not None) else "Time [s]"
            y_label = "Frequency [MHz]"

        extent = [0, self.time[-1], self.freqs[-1], self.freqs[0]]
        cbar_label = "" if self.remove_titles else ("Intensity [Digits]" if not self.use_db else "Intensity [dB]")
        tick_font_px = self.tick_font_px
        axis_label_font_px = self.axis_label_font_px
        title_font_px = self.title_font_px
        if getattr(self, "_hw_default_font_sizes_active", False):
            tick_font_px = self.HW_DEFAULT_TICK_FONT_PX
            axis_label_font_px = self.HW_DEFAULT_AXIS_FONT_PX
            title_font_px = self.HW_DEFAULT_TITLE_FONT_PX
        self.accel_canvas.set_text_style(
            font_family=self.graph_font_family,
            tick_font_px=tick_font_px,
            axis_label_font_px=axis_label_font_px,
            title_font_px=title_font_px,
            title_bold=self.title_bold,
            title_italic=self.title_italic,
            axis_bold=self.axis_bold,
            axis_italic=self.axis_italic,
            ticks_bold=self.ticks_bold,
            ticks_italic=self.ticks_italic,
        )
        self.accel_canvas.update_image(
            display_data,
            extent=extent,
            cmap=self.get_current_cmap(),
            title=plot_title,
            x_label=x_label,
            y_label=y_label,
            colorbar_label=cbar_label,
            view=view,
        )
        self.accel_canvas.set_time_mode(self.use_utc, self.ut_start_sec)
        return True

    def _apply_mpl_theme(self):
        """
        Ensure Matplotlib canvas (figure, axes, and colorbar) matches the current Qt theme.
        Call this AFTER you finish setting titles/labels/ticks.
        """
        if not hasattr(self, "canvas"):
            return

        fig = getattr(self.canvas, "figure", None) or getattr(self.canvas, "fig", None)
        ax = getattr(self.canvas, "ax", None)
        if fig is None or ax is None:
            return

        # 1) Preferred: use your theme manager if available
        if self.theme and hasattr(self.theme, "apply_mpl"):
            try:
                self.theme.apply_mpl(fig, ax, getattr(self, "current_colorbar", None))
            except Exception:
                pass

        # 2) Fallback: enforce readable colors from Qt palette (covers white bg / black text issues)
        app = QApplication.instance()
        if app:
            pal = app.palette()
            win_bg = pal.color(QPalette.Window).name()
            base_bg = pal.color(QPalette.Base).name()
            fg = pal.color(QPalette.WindowText).name()
            mid = pal.color(QPalette.Mid).name()

            fig.set_facecolor(win_bg)
            ax.set_facecolor(base_bg)

            for spine in ax.spines.values():
                spine.set_color(fg)

            ax.tick_params(axis="both", colors=fg, which="both")
            ax.xaxis.label.set_color(fg)
            ax.yaxis.label.set_color(fg)
            ax.title.set_color(fg)

            # If you use grids elsewhere, keep their color readable without forcing grid on
            ax.grid(False)
            ax.set_axisbelow(True)

            cbar = getattr(self, "current_colorbar", None)
            if cbar is not None:
                cax = cbar.ax
                cax.set_facecolor(base_bg)
                cax.tick_params(colors=fg)
                cax.yaxis.label.set_color(fg)
                for spine in cax.spines.values():
                    spine.set_color(fg)
                try:
                    cbar.outline.set_edgecolor(mid)
                except Exception:
                    pass

        self.canvas.draw_idle()

    def _refresh_toolbar_icons(self):
        # Only run after toolbar actions exist
        for attr, fname in (
                ("tb_open", "open.svg"),
                ("tb_export", "export.svg"),
                ("tb_export_fits","export_fits.svg"),
                ("tb_undo", "undo.svg"),
                ("tb_redo", "redo.svg"),
                ("tb_download", "download.svg"),
                ("tb_drift", "drift.svg"),
                ("tb_isolate", "isolate.svg"),
                ("tb_max", "max.svg"),
                ("tb_zoom", "zoom.svg"),
                ("tb_lock", "lock.svg"),
                ("tb_unlock", "unlock.svg"),
                ("tb_reset_sel", "reset_selection.svg"),
                ("tb_reset_all", "reset_all.svg"),
        ):
            act = getattr(self, attr, None)
            if act is not None:
                act.setIcon(self._icon(fname))

    def _on_theme_changed(self, dark: bool):
        self._refresh_toolbar_icons()

        # If you already added MPL theme syncing earlier, keep it too:
        if hasattr(self, "_apply_mpl_theme"):
            self._apply_mpl_theme()
        self._apply_accel_theme()

    def toggle_left_sidebar(self):
        self._set_sidebar_collapsed(not bool(getattr(self, "_sidebar_collapsed", False)))

    def _set_sidebar_collapsed(self, collapsed: bool):
        self._sidebar_collapsed = bool(collapsed)
        if getattr(self, "side_scroll", None) is None or getattr(self, "sidebar_toggle_btn", None) is None:
            return

        self.side_scroll.setVisible(not self._sidebar_collapsed)
        if self._sidebar_collapsed:
            self.sidebar_toggle_btn.setText("▶")
            self.sidebar_toggle_btn.setToolTip("Expand sidebar")
        else:
            self.sidebar_toggle_btn.setText("◀")
            self.sidebar_toggle_btn.setToolTip("Collapse sidebar")

        layout = getattr(self, "_main_layout", None)
        if layout is not None:
            layout.setSpacing(0)

    def _normalize_plot_type(self, title: str | None) -> str:
        txt = str(title or "").strip()
        if not txt:
            return "Raw"

        lowered = txt.lower()
        raw_aliases = {
            "raw",
            "raw data",
            "dynamic spectrum",
            "combined time",
            "combined time plot",
            "raw data (combined frequency)",
        }
        if lowered in raw_aliases:
            return "Raw"

        if lowered in {"background subtracted", "background-subtracted"}:
            return "Background Subtracted"

        return txt

    def _default_graph_title(self, plot_type: str | None = None) -> str:
        normalized = self._normalize_plot_type(plot_type if plot_type is not None else self.current_plot_type)
        base = str(getattr(self, "filename", "") or "").strip() or "Untitled"

        return f"{base}-{normalized}"

    def _current_graph_title_for_export(self) -> str:
        ax = getattr(getattr(self, "canvas", None), "ax", None)
        if ax is not None:
            try:
                live_title = str(ax.get_title() or "").strip()
                if live_title:
                    return live_title
            except Exception:
                pass

        override = str(getattr(self, "graph_title_override", "") or "").strip()
        if override:
            return override

        return self._default_graph_title(self.current_plot_type)

    def _sanitize_export_stem(self, text: str) -> str:
        stem = str(text or "").strip()
        if not stem:
            stem = "export"

        stem = re.sub(r"\s+", " ", stem)
        stem = re.sub(r"[\\\\/:*?\"<>|]+", "_", stem)
        stem = stem.strip(" .")
        return stem or "export"

    def apply_graph_properties_live(self, *_):
        """
        Apply graph styling changes (title, font family, font sizes, bold/italic flags)
        to the CURRENT plot without rebuilding the image.
        """

        ax = getattr(self.canvas, "ax", None)
        if ax is None:
            return

        # Must have an image already
        if not ax.images or len(ax.images) == 0:
            return

        # -----------------------------
        # 1) READ UI STATE FIRST
        # -----------------------------
        self.remove_titles = self.remove_titles_chk.isChecked()

        self.title_bold = self.title_bold_chk.isChecked()
        self.title_italic = self.title_italic_chk.isChecked()

        self.axis_bold = self.axis_bold_chk.isChecked()
        self.axis_italic = self.axis_italic_chk.isChecked()

        self.ticks_bold = self.ticks_bold_chk.isChecked()
        self.ticks_italic = self.ticks_italic_chk.isChecked()

        self.graph_title_override = self.title_edit.text().strip()

        font_choice = self.font_combo.currentText()
        self.graph_font_family = "" if font_choice == "Default" else font_choice

        self.tick_font_px = int(self.tick_font_spin.value())
        self.axis_label_font_px = int(self.axis_font_spin.value())
        self.title_font_px = int(self.title_font_spin.value())

        # Helpers
        def _weight(bold: bool) -> str:
            return "bold" if bold else "normal"

        def _style(italic: bool) -> str:
            return "italic" if italic else "normal"

        fontfam = self.graph_font_family if self.graph_font_family else None

        # -----------------------------
        # 2) APPLY TITLES / LABELS
        # -----------------------------
        if self.remove_titles:
            ax.set_title("")
            ax.set_xlabel("")
            ax.set_ylabel("")
        else:
            # Default title if custom is empty
            title_text = self.graph_title_override or self._default_graph_title(self.current_plot_type)

            ax.set_title(
                title_text,
                fontsize=self.title_font_px,
                fontfamily=fontfam,
                fontweight=_weight(self.title_bold),
                fontstyle=_style(self.title_italic),
            )

            # Always set axis labels explicitly so toggles apply immediately
            xlab = "Time [UT]" if (self.use_utc and self.ut_start_sec is not None) else "Time [s]"
            ylab = "Frequency [MHz]"

            ax.set_xlabel(
                xlab,
                fontsize=self.axis_label_font_px,
                fontfamily=fontfam,
                fontweight=_weight(self.axis_bold),
                fontstyle=_style(self.axis_italic),
            )
            ax.set_ylabel(
                ylab,
                fontsize=self.axis_label_font_px,
                fontfamily=fontfam,
                fontweight=_weight(self.axis_bold),
                fontstyle=_style(self.axis_italic),
            )

        # -----------------------------
        # 3) APPLY TICK LABEL STYLE
        # -----------------------------
        for lbl in ax.get_xticklabels():
            lbl.set_fontsize(self.tick_font_px)
            lbl.set_fontweight(_weight(self.ticks_bold))
            lbl.set_fontstyle(_style(self.ticks_italic))
            if fontfam:
                lbl.set_fontfamily(fontfam)

        for lbl in ax.get_yticklabels():
            lbl.set_fontsize(self.tick_font_px)
            lbl.set_fontweight(_weight(self.ticks_bold))
            lbl.set_fontstyle(_style(self.ticks_italic))
            if fontfam:
                lbl.set_fontfamily(fontfam)

        # Colorbar style (ticks + title)
        if self.current_colorbar is not None:
            try:
                cax = self.current_colorbar.ax
                fontfam = (self.graph_font_family if self.graph_font_family else None)

                # ---- ticks ----
                cax.tick_params(labelsize=self.tick_font_px)
                for lbl in cax.get_yticklabels():
                    lbl.set_fontsize(self.tick_font_px)
                    lbl.set_fontweight(_weight(self.ticks_bold))
                    lbl.set_fontstyle(_style(self.ticks_italic))
                    if fontfam:
                        lbl.set_fontfamily(fontfam)

                # ---- title/label ----
                if self.remove_titles:
                    # Hide the colorbar label, but DO NOT erase the stored text
                    self.current_colorbar.set_label("")
                    cax.set_ylabel("")  # extra safety for some backends
                    cax.yaxis.label.set_text("")
                else:
                    # Restore label text + style
                    label_text = getattr(self, "_colorbar_label_text", "")
                    if not label_text:
                        label_text = cax.get_ylabel()  # fallback

                    self.current_colorbar.set_label(
                        label_text,
                        fontsize=self.axis_label_font_px,
                        fontfamily=fontfam,
                        fontweight=_weight(self.axis_bold),
                        fontstyle=_style(self.axis_italic),
                    )

                    # Force styling on the underlying label object too
                    ylab = cax.yaxis.label
                    ylab.set_fontsize(self.axis_label_font_px)
                    ylab.set_fontweight(_weight(self.axis_bold))
                    ylab.set_fontstyle(_style(self.axis_italic))
                    if fontfam:
                        ylab.set_fontfamily(fontfam)

            except Exception:
                pass

        # -----------------------------
        # 4) UI: disable title inputs when Remove Titles is on
        # -----------------------------
        disable = self.remove_titles
        self.title_edit.setEnabled(not disable)
        self.title_bold_chk.setEnabled(not disable)
        self.title_italic_chk.setEnabled(not disable)
        self.axis_bold_chk.setEnabled(not disable)
        self.axis_italic_chk.setEnabled(not disable)

        self.canvas.draw_idle()
        self._refresh_accel_plot(preserve_view=True)

    def _on_tick_font_spin_changed(self, _value):
        self._hw_default_font_sizes_active = False
        self.apply_graph_properties_live()

    def _on_axis_font_spin_changed(self, _value):
        self._hw_default_font_sizes_active = False
        self.apply_graph_properties_live()

    def _on_title_font_spin_changed(self, _value):
        self._hw_default_font_sizes_active = False
        self.apply_graph_properties_live()

    def load_file(self):
        if not self._maybe_prompt_save_dirty():
            return
        initial_dir = os.path.dirname(self.filename) if self.filename else ""
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFiles)
        dialog.setNameFilter("FITS files (*.fit *.fits *.fit.gz *.fits.gz)")
        dialog.setOption(QFileDialog.DontUseNativeDialog, True)

        if dialog.exec():
            file_paths = dialog.selectedFiles()
        else:
            return

        if not file_paths:
            return

        if len(file_paths) == 1:
            file_path = file_paths[0]
            self.filename = os.path.basename(file_path)

            res = load_callisto_fits(file_path, memmap=False)
            self.raw_data = res.data
            self._invalidate_noise_cache()
            self.freqs = res.freqs
            self.time = res.time

            # Store header template for Export to FITS / viewing
            self._fits_header0 = res.header0.copy()
            self._fits_source_path = file_path

            self._is_combined = False
            self._combined_mode = None
            self._combined_sources = []
            self.ut_start_sec = extract_ut_start_sec(res.header0)

            # Reset derived state for a fresh start
            self.noise_reduced_data = None
            self.noise_reduced_original = None
            self.lasso_mask = None
            self.noise_vmin = None
            self.noise_vmax = None
            self.current_display_data = None
            self._undo_stack.clear()
            self._redo_stack.clear()

            self.plot_data(self.raw_data, title="Raw")
            self._project_path = None
            self._max_intensity_state = None
            self._mark_project_dirty()
            return

        from src.Backend.burst_processor import (
            are_time_combinable,
            are_frequency_combinable,
            combine_time,
            combine_frequency,
        )

        try:
            if are_time_combinable(file_paths):
                combined = combine_time(file_paths)
            elif are_frequency_combinable(file_paths):
                combined = combine_frequency(file_paths)
            else:
                error_msg = (
                    "The selected FITS files cannot be combined.\n\n"
                    "Valid combinations are:\n"
                    "1. Frequency Combine:\n"
                    "   • Same station\n"
                    "   • Same date\n"
                    "   • Same timestamp (HHMMSS)\n"
                    "   • Different receiver IDs\n"
                    "   • Matching time arrays\n\n"
                    "2. Time Combine:\n"
                    "   • Same station\n"
                    "   • Same receiver ID\n"
                    "   • Same date\n"
                    "   • Different timestamps (continuous time segments)\n"
                    "   • Matching frequency arrays\n\n"
                    "Your selection does not match either rule.\n"
                    "Please choose files that follow one of the above patterns."
                )
                QMessageBox.warning(self, "Invalid Combination Selection", error_msg)
                return

            self.load_combined_into_main(combined)
            self._project_path = None
            self._max_intensity_state = None
            self._mark_project_dirty()
            self.statusBar().showMessage(f"Loaded {len(file_paths)} files (combined)", 5000)
        except Exception as e:
            QMessageBox.critical(self, "Combine Error", f"An error occurred while combining files:\n{e}")
            return

    def _intensity_for_display(self, data):
        if data is None:
            return None
        return data * self.DB_SCALE if self.use_db else data

    def _intensity_range_for_display(self, vmin, vmax):
        if vmin is None or vmax is None:
            return vmin, vmax
        return (vmin * self.DB_SCALE, vmax * self.DB_SCALE) if self.use_db else (vmin, vmax)

    def update_units(self):
        self.set_units_mode(bool(self.units_db_radio.isChecked()))

    def set_units_mode(self, use_db: bool):
        self.use_db = bool(use_db)

        if self.raw_data is None:
            return

        self._mark_project_dirty()

        # Choose which data to replot
        data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data

        # Replot with new unit selection
        self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def combine_frequency_files(self, file_paths):
        file_paths = sorted(file_paths)
        data_list = []
        freq_list = []
        time_array = None

        for path in file_paths:
            res = load_callisto_fits(path, memmap=False)
            data_list.append(res.data)
            freq_list.append(res.freqs)
            time_array = res.time

        combined_data = np.concatenate(data_list, axis=0)
        combined_freqs = np.concatenate(freq_list)
        combined_time = time_array

        return combined_data, combined_freqs, combined_time

    def combine_time_files(self, file_paths):
        file_paths = sorted(file_paths)
        data_list = []
        time_list = []
        freqs = None

        for path in file_paths:
            res = load_callisto_fits(path, memmap=False)
            data_list.append(res.data)
            time_list.append(res.time)
            freqs = res.freqs

        combined_data = np.concatenate(data_list, axis=1)

        fixed_time = []
        offset = 0

        for t in time_list:
            t = np.array(t)
            fixed_time.append(t + offset)
            offset += t[-1]

        combined_time = np.concatenate(fixed_time)

        combined_freqs = freqs
        return combined_data, combined_freqs, combined_time

    def load_fits_into_main(self, file_path):
        res = load_callisto_fits(file_path, memmap=False)
        self.raw_data = res.data
        self._invalidate_noise_cache()
        self.freqs = res.freqs
        self.time = res.time
        self.filename = os.path.basename(file_path)

        # header template
        self._fits_header0 = res.header0.copy()
        self._fits_source_path = file_path

        self._is_combined = False
        self._combined_mode = None
        self._combined_sources = []

        self.ut_start_sec = extract_ut_start_sec(res.header0)

        # Reset derived state for a fresh start
        self.noise_reduced_data = None
        self.noise_reduced_original = None
        self.lasso_mask = None
        self.noise_vmin = None
        self.noise_vmax = None
        self.current_display_data = None
        self._undo_stack.clear()
        self._redo_stack.clear()

        self.plot_data(self.raw_data, title="Raw")
        self._project_path = None
        self._max_intensity_state = None
        self._mark_project_dirty()

    def load_combined_into_main(self, combined):
        self.raw_data = combined["data"]
        self._invalidate_noise_cache()
        self.freqs = combined["freqs"]
        self.time = combined["time"]
        self.filename = combined.get("filename", "Combined")
        self.ut_start_sec = combined.get("ut_start_sec", None)

        # Reset derived state for a fresh start
        self.noise_reduced_data = None
        self.noise_reduced_original = None
        self.lasso_mask = None
        self.noise_vmin = None
        self.noise_vmax = None
        self.current_display_data = None
        self._undo_stack.clear()
        self._redo_stack.clear()

        # metadata for Export to FITS
        self._is_combined = True
        self._combined_mode = combined.get("combine_type", None)
        self._combined_sources = combined.get("sources", [])

        hdr0 = combined.get("header0", None)
        if hdr0 is None:
            hdr0 = build_combined_header(
                None,
                mode=combined.get("combine_type", "combined"),
                sources=combined.get("sources", []) or self._combined_sources,
                data_shape=getattr(self.raw_data, "shape", (0, 0)),
                freqs=self.freqs,
                time=self.time,
            )
        self._fits_header0 = hdr0.copy() if hdr0 is not None else None
        self._fits_source_path = None

        self.plot_data(self.raw_data)
        self._project_path = None
        self._max_intensity_state = None
        self._mark_project_dirty()

    def schedule_noise_update(self):
        if self.raw_data is None:
            return
        self.noise_smooth_timer.start()

    def _on_noise_slider_pressed(self):
        self._noise_slider_drag_active = True

    def _on_noise_slider_released(self):
        self._noise_slider_drag_active = False
        if self.raw_data is None:
            return
        self.noise_smooth_timer.stop()
        self.update_noise_live()

    def _invalidate_noise_cache(self):
        self._noise_base_data = None
        self._noise_base_source_id = None

    def _ensure_noise_base_data(self):
        if self.raw_data is None:
            return None

        source_id = id(self.raw_data)
        if self._noise_base_data is not None and self._noise_base_source_id == source_id:
            return self._noise_base_data

        arr = np.asarray(self.raw_data, dtype=np.float32)
        row_mean = arr.mean(axis=1, keepdims=True, dtype=np.float32)
        self._noise_base_data = arr - row_mean
        self._noise_base_source_id = source_id
        return self._noise_base_data

    def _compute_noise_reduced(self, low: float, high: float):
        base = self._ensure_noise_base_data()
        if base is None:
            return None
        return np.clip(base, low, high).astype(np.float32, copy=False)

    def _update_live_preview_canvas(self, data):
        if not self._hardware_mode_enabled():
            return False
        ok = self._refresh_accel_plot(data=data, title="Background Subtracted", preserve_view=True)
        if ok:
            self._show_accel_canvas()
        return ok

    def _commit_noise_live_update(self):
        self.noise_commit_timer.stop()

        if self._noise_slider_drag_active:
            return

        if self.raw_data is None or self.noise_reduced_data is None:
            self._noise_undo_pending = False
            if self._hardware_mode_enabled():
                self._show_accel_canvas()
            else:
                self._show_plot_canvas()
            return

        self.plot_data(self.noise_reduced_data, title="Background Subtracted")
        self._noise_undo_pending = False

    def update_noise_live(self):
        if self.raw_data is None:
            return

        if not self._noise_undo_pending:
            self._push_undo_state()
            self._noise_undo_pending = True

        low = self.lower_slider.value()
        high = self.upper_slider.value()
        if low > high:
            low, high = high, low

        data = self._compute_noise_reduced(low, high)
        if data is None:
            return

        self.noise_reduced_data = data
        self.noise_reduced_original = data.copy()

        self.noise_vmin = float(np.nanmin(data)) if data.size else None
        self.noise_vmax = float(np.nanmax(data)) if data.size else None
        self.current_plot_type = "Background Subtracted"

        if self._update_live_preview_canvas(data):
            self.noise_commit_timer.start()
            if not self._noise_slider_drag_active:
                self._commit_noise_live_update()
        else:
            self.plot_data(data, title="Background Subtracted")
            self._noise_undo_pending = False

        # enable tools
        self._sync_toolbar_enabled_states()

    def plot_data(self, data, title="Raw", keep_view=False, restore_view=None):
        view = restore_view if restore_view is not None else (self._capture_view() if keep_view else None)
        QTimer.singleShot(0, lambda: self._plot_data_internal(data, title, view))

    def _capture_view(self):
        """Save current zoom/pan limits (only if a plot exists)."""
        if self._hardware_mode_enabled():
            try:
                return self.accel_canvas.get_view()
            except Exception:
                return None

        try:
            ax = self.canvas.ax
            if ax is None or ax.images is None or len(ax.images) == 0:
                return None
            return {
                "xlim": ax.get_xlim(),
                "ylim": ax.get_ylim(),
            }
        except Exception:
            return None

    def _restore_view(self, view):
        """Restore zoom/pan limits safely."""
        if not view:
            return

        if self._hardware_mode_enabled():
            try:
                self.accel_canvas.set_view(view)
            except Exception:
                pass
            return

        try:
            ax = self.canvas.ax
            ax.set_xlim(view["xlim"])
            ax.set_ylim(view["ylim"])
        except Exception:
            pass

    def _plot_data_internal(self, data, title="Raw", view=None):

        self._stop_rect_zoom()

        if self.time is None or self.freqs is None:
            print("Time or frequency data not loaded. Skipping plot.")
            return

        if not hasattr(self.canvas, 'ax') or self.canvas.ax is None:
            print("Canvas not ready yet")
            return

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        # Remove old colorbar axis safely
        try:
            if self.current_cax:
                self.current_cax.remove()
                self.current_cax = None
            if self.current_colorbar:
                self.current_colorbar.remove()
                self.current_colorbar = None
        except Exception as e:
            print("Error removing previous colorbar:", e)

        # Define colormap
        # Choose cmap
        if self.current_cmap_name == "Custom":
            colors = [(0.0, 'blue'), (0.5, 'red'), (1.0, 'yellow')]
            cmap = mcolors.LinearSegmentedColormap.from_list('custom_RdYlBu', colors)
        else:
            cmap = plt.get_cmap(self.current_cmap_name)

        # x-axis always in seconds, UT formatting handled separately
        extent = [0, self.time[-1], self.freqs[-1], self.freqs[0]]

        # Prepare colorbar axis
        divider = make_axes_locatable(self.canvas.ax)
        cax = divider.append_axes("right", size="5%", pad=0.1)
        self.current_cax = cax

        # Show image (convert units for display if needed)
        display_data = self._intensity_for_display(data)

        self.current_display_data = display_data

        im = self.canvas.ax.imshow(display_data, aspect='auto', extent=extent, cmap=cmap)

        self.current_colorbar = self.canvas.figure.colorbar(im, cax=cax)

        label = "Intensity [Digits]" if not self.use_db else "Intensity [dB]"
        self.current_colorbar.set_label(label)

        # Store label string so apply_graph_properties_live can re-style it
        self._colorbar_label_text = label

        plot_type = self._normalize_plot_type(title)
        self.canvas.ax.set_ylabel("Frequency [MHz]")
        self.canvas.ax.set_title(self._default_graph_title(plot_type), fontsize=14)

        # Save full-extent limits as the "home" view (used for Reset Selection after zoom/pan)
        self._home_view = {
            "xlim": self.canvas.ax.get_xlim(),
            "ylim": self.canvas.ax.get_ylim(),
        }

        self.format_axes()  # Format x-axis based on user selection (seconds/UT)
        self._restore_view(view)

        # Apply graph properties (title/font/sizes) after plot rebuild
        self.apply_graph_properties_live()

        # Now force the MPL colors/background to match the theme (dark/light)
        self._apply_mpl_theme()
        style_axes(self.canvas.ax)

        self.graph_group.setEnabled(True)
        self.canvas.draw_idle()

        try:
            self._refresh_accel_plot(data=data, title=plot_type, view=view, preserve_view=(view is not None))
        except Exception:
            pass

        self.current_plot_type = plot_type
        if self._hardware_mode_enabled():
            self._show_accel_canvas()
        else:
            self._show_plot_canvas()
        self._sync_toolbar_enabled_states()
        self.statusBar().showMessage(f"Loaded: {self.filename}", 5000)

    def on_mouse_motion_status(self, event):
        """Show time, frequency and intensity under cursor in status bar."""
        in_axes = event.inaxes == self.canvas.ax and event.xdata is not None and event.ydata is not None
        x = float(event.xdata) if in_axes else 0.0
        y = float(event.ydata) if in_axes else 0.0
        self._update_cursor_label_from_xy(x, y, in_axes)

    def on_accel_mouse_motion_status(self, x: float, y: float, inside: bool):
        self._update_cursor_label_from_xy(float(x), float(y), bool(inside))

    def _update_cursor_label_from_xy(self, x: float, y: float, inside: bool):
        if not inside:
            self.cursor_label.setText("t = 0.00   |   f = 0.00 MHz   |   I = 0.00")
            return

        if self.current_display_data is None or self.time is None or self.freqs is None:
            return

        time_arr = np.array(self.time)
        freq_arr = np.array(self.freqs)
        if time_arr.size == 0 or freq_arr.size == 0:
            return

        idx_x = int(np.argmin(np.abs(time_arr - float(x))))
        idx_y = int(np.argmin(np.abs(freq_arr - float(y))))

        ny, nx = self.current_display_data.shape
        if idx_x < 0 or idx_x >= nx or idx_y < 0 or idx_y >= ny:
            return

        t_val = time_arr[idx_x]
        f_val = freq_arr[idx_y]
        intensity = self.current_display_data[idx_y, idx_x]

        if self.use_utc and self.ut_start_sec is not None:
            total_seconds = self.ut_start_sec + t_val
            hours = int(total_seconds // 3600) % 24
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d} UT"
        else:
            time_str = f"{t_val:.2f} s"

        unit_label = "dB" if self.use_db else "Digits"
        msg = (
            f"t = {time_str}   |   "
            f"f = {f_val:.2f} MHz   |   "
            f"I = {intensity:.2f} {unit_label}"
        )
        self.cursor_label.setText(msg)

    def change_cmap(self, name):
        self._push_undo_state()
        self.current_cmap_name = name
        if self.raw_data is None:
            return
        data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
        self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def get_current_cmap(self):
        if self.current_cmap_name == "Custom":
            colors = [(0.0, 'blue'), (0.5, 'red'), (1.0, 'yellow')]
            return mcolors.LinearSegmentedColormap.from_list('custom_RdYlBu', colors)
        else:
            return plt.get_cmap(self.current_cmap_name)

    def _on_accel_view_interaction_finished(self, prev_view, new_view):
        if not self._hardware_mode_enabled():
            return
        if prev_view and new_view and (not self._views_close(prev_view, new_view)):
            self._push_undo_view(prev_view)
        self._sync_toolbar_enabled_states()

    def _on_accel_rect_zoom_finished(self, prev_view, new_view):
        if not self._hardware_mode_enabled():
            return
        self.rect_zoom_active = False
        if prev_view and new_view and (not self._views_close(prev_view, new_view)):
            self._push_undo_view(prev_view)
        self.statusBar().showMessage("Zoomed to selected region (still locked).", 2500)
        self._sync_toolbar_enabled_states()

    def _on_accel_lasso_finished(self, verts):
        if not self._hardware_mode_enabled():
            return
        if not self.lasso_active:
            return
        self.on_lasso_select(list(verts))

    def _on_accel_drift_point_added(self, x: float, y: float):
        self.drift_points.append((float(x), float(y)))
        self.accel_canvas.show_drift_points(self.drift_points, with_segments=False)

    def _on_accel_drift_capture_finished(self, points):
        self.drift_points = [(float(x), float(y)) for (x, y) in (points or [])]
        self.finish_drift_estimation()

    def on_scroll_zoom(self, event):
        """Smooth zoom using mouse scroll wheel."""
        if self._hardware_mode_enabled():
            return

        if self.lasso_active:
            return

        if getattr(self, "nav_locked", False) or getattr(self, "rect_zoom_active", False):
            return

        ax = self.canvas.ax

        # Mouse pointer must be inside the plot
        if event.inaxes != ax:
            return
        if event.xdata is None or event.ydata is None:
            return

        prev_view = self._capture_view()

        # Zoom factor
        base_scale = 1.15  # smooth and gentle zoom

        # Zoom IN
        if event.button == "up":
            scale_factor = 1 / base_scale
        # Zoom OUT
        elif event.button == "down":
            scale_factor = base_scale
        else:
            return

        # Current axis limits
        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()

        x_range = (x_max - x_min)
        y_range = (y_max - y_min)

        # Mouse location in data coords
        xdata = event.xdata
        ydata = event.ydata

        # Compute new ranges
        new_x_range = x_range * scale_factor
        new_y_range = y_range * scale_factor

        # Shift around mouse cursor
        x_shift = (xdata - x_min) * (1 - scale_factor)
        y_shift = (ydata - y_min) * (1 - scale_factor)

        new_x_min = x_min + x_shift
        new_x_max = new_x_min + new_x_range

        new_y_min = y_min + y_shift
        new_y_max = new_y_min + new_y_range

        # Apply limits
        ax.set_xlim(new_x_min, new_x_max)
        ax.set_ylim(new_y_min, new_y_max)

        if prev_view:
            self._push_undo_view(prev_view)

        self.canvas.draw_idle()
        self._sync_toolbar_enabled_states()

    def on_mouse_press(self, event):
        """
        Start panning with LEFT mouse button inside the main axes.
        (No modifier keys needed.)
        """
        if self._hardware_mode_enabled():
            return

        if getattr(self, "nav_locked", False) or getattr(self, "rect_zoom_active", False):
            return

        if self.lasso_active:
            return

        # Only react if we click inside the image axes
        if event.inaxes != self.canvas.ax:
            return

        # Left button = start pan
        if event.button == 1 and event.xdata is not None and event.ydata is not None:
            self._panning = True
            self._last_pan_xy = (event.xdata, event.ydata)
            self._pan_start_view = self._capture_view()

    def on_mouse_move(self, event):
        """
        Perform the pan movement while the left mouse button is held.
        """
        if self._hardware_mode_enabled():
            return

        if getattr(self, "nav_locked", False) or getattr(self, "rect_zoom_active", False):
            return

        if self.lasso_active:
            return

        # Only act if we are in panning mode and pointer is on data
        if not self._panning or event.xdata is None or event.ydata is None:
            return

        ax = self.canvas.ax

        # Previous mouse position in data coordinates
        x_prev, y_prev = self._last_pan_xy

        # Current mouse position
        x_curr, y_curr = event.xdata, event.ydata

        # Difference
        dx = x_prev - x_curr
        dy = y_prev - y_curr

        # Current limits
        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()

        # Shift both axes
        ax.set_xlim(x_min + dx, x_max + dx)
        ax.set_ylim(y_min + dy, y_max + dy)

        self.canvas.draw_idle()

        # Update last mouse position
        self._last_pan_xy = (x_curr, y_curr)

    def on_mouse_release(self, event):
        """
        Stop panning when the mouse button is released.
        """
        if self._hardware_mode_enabled():
            return

        self._panning = False
        self._last_pan_xy = None
        if getattr(self, "_pan_start_view", None):
            try:
                cur_view = self._capture_view()
                if cur_view and (not self._views_close(cur_view, self._pan_start_view)):
                    self._push_undo_view(self._pan_start_view)
            except Exception:
                pass
            finally:
                self._pan_start_view = None
                self._sync_toolbar_enabled_states()


    def activate_drift_tool(self):
        self.statusBar().showMessage("Click multiple points along the burst. Right-click or double-click to finish.",
                                     8000)
        self.drift_points = []
        if self._hardware_mode_enabled():
            self.accel_canvas.begin_drift_capture()
            self.accel_canvas.show_drift_points([], with_segments=False)
            return
        self.drift_click_cid = self.canvas.mpl_connect("button_press_event", self.on_drift_point_click)

    def on_drift_point_click(self, event):
        if not event.inaxes:
            return

        # Right-click or double-click to finish
        if event.button == 3 or event.dblclick:
            self.finish_drift_estimation()
            return

        self.drift_points.append((event.xdata, event.ydata))
        self.canvas.ax.plot(event.xdata, event.ydata, 'w*')
        self.canvas.draw()

    def finish_drift_estimation(self):
        if self._hardware_mode_enabled():
            self.accel_canvas.stop_interaction_capture()
        else:
            self.canvas.mpl_disconnect(self.drift_click_cid)

        if len(self.drift_points) < 2:
            self.statusBar().showMessage("Need at least two points to estimate drift.", 4000)
            return

        drift_rates = []
        for i in range(len(self.drift_points) - 1):
            x1, y1 = self.drift_points[i]
            x2, y2 = self.drift_points[i + 1]
            if abs(x2 - x1) < 1e-12:
                continue
            drift = (y2 - y1) / (x2 - x1)
            drift_rates.append(drift)
            # Draw line between points
            if not self._hardware_mode_enabled():
                self.canvas.ax.plot([x1, x2], [y1, y2], linestyle='--', color='lime')

        if not drift_rates:
            self.statusBar().showMessage("Need points with different time values to estimate drift.", 4000)
            return

        avg_drift = np.mean(drift_rates)
        if self._hardware_mode_enabled():
            self.accel_canvas.show_drift_points(self.drift_points, with_segments=True)
        else:
            self.canvas.ax.legend(["Drift Segments"])
            self.canvas.draw()

        self.statusBar().showMessage(
            f"Average Drift Rate: {avg_drift:.4f} MHz/s, Start Frequency: {y1: .3f}, End Frequency: {y2: .3f}, Duration: {x2 - x1: .3f} s",
            0)

    def activate_lasso(self):
        if self.noise_reduced_data is None:
            QMessageBox.warning(self, "Error", "Please apply background substraction before isolating a burst.")
            return

        if self._hardware_mode_enabled():
            self.lasso_active = True
            self.accel_canvas.begin_lasso_capture()
            self.statusBar().showMessage("Press, drag around the burst, and release to isolate.", 5000)
            return

        self.canvas.mpl_disconnect(self._cid_press)
        self.canvas.mpl_disconnect(self._cid_motion)
        self.canvas.mpl_disconnect(self._cid_release)

        self.lasso_active = True

        # Disconnect old lasso
        if self.lasso:
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None

        self.canvas.ax.set_title("Draw around the burst")
        self.canvas.draw()

        self.lasso = LassoSelector(self.canvas.ax, onselect=self.on_lasso_select)

    def on_lasso_select(self, verts):

        if self.noise_reduced_data is None:
            print("Lasso used before data was prepared. Ignoring.")
            return

        verts_arr = np.asarray(verts, dtype=float)
        if verts_arr.ndim != 2 or verts_arr.shape[0] < 3 or verts_arr.shape[1] != 2:
            print("Invalid lasso selection. Ignoring.")
            return
        if not np.allclose(verts_arr[0], verts_arr[-1]):
            verts_arr = np.vstack([verts_arr, verts_arr[0]])

        path = Path(verts_arr, closed=True)

        ny, nx = self.noise_reduced_data.shape
        # Hardware view uses an inverted Y-axis transform; map rows accordingly
        # so the lasso mask aligns with what the user actually drew.
        if self._hardware_mode_enabled():
            y = np.linspace(self.freqs[-1], self.freqs[0], ny)
        else:
            y = np.linspace(self.freqs[0], self.freqs[-1], ny)
        x = np.linspace(0, self.time[-1], nx)
        X, Y = np.meshgrid(x, y)

        coords = np.column_stack((X.flatten(), Y.flatten()))
        mask = path.contains_points(coords).reshape(ny, nx)

        self.lasso_mask = mask  # store for use later

        # Safely disconnect the lasso tool
        if self.lasso:
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None

        if not self._hardware_mode_enabled():
            self._cid_press = self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
            self._cid_motion = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
            self._cid_release = self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self.lasso_active = False
        # Defer drawing to avoid crash during event handling
        QTimer.singleShot(0, lambda: self._plot_isolated_burst(mask))

    def _plot_isolated_burst(self, mask):
        self._push_undo_state()
        # Build isolated data array
        burst_isolated = np.zeros_like(self.noise_reduced_data)
        burst_isolated[mask] = self.noise_reduced_data[mask]

        # Clear figure
        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)

        # Remove old colorbar safely
        try:
            if self.current_colorbar:
                self.current_colorbar.remove()
        except Exception:
            pass
        self.current_colorbar = None

        try:
            if self.current_cax:
                self.current_cax.remove()
        except Exception:
            pass
        self.current_cax = None

        # Restore extent
        extent = [0, self.time[-1], self.freqs[-1], self.freqs[0]]

        # Get SAME colormap user selected
        cmap = self.get_current_cmap()

        # === CREATE COLORBAR AXIS HERE (must exist before colorbar) ===
        divider = make_axes_locatable(self.canvas.ax)
        cax = divider.append_axes("right", size="5%", pad=0.1)
        self.current_cax = cax

        # Plot with fixed vmin/vmax from noise reduction (converted to display units if needed)
        display_burst = self._intensity_for_display(burst_isolated)
        vmin, vmax = self._intensity_range_for_display(self.noise_vmin, self.noise_vmax)

        im = self.canvas.ax.imshow(
            display_burst,
            aspect='auto',
            extent=extent,
            cmap=cmap,
            vmin=vmin,
            vmax=vmax,
        )

        self.current_display_data = display_burst

        # Create new colorbar
        self.current_colorbar = self.canvas.figure.colorbar(im, cax=cax)
        self._colorbar_label_text = "Intensity [Digits]" if not self.use_db else "Intensity [dB]"
        self.current_colorbar.set_label(self._colorbar_label_text)

        # Labels
        self.canvas.ax.set_title("Isolated Burst")
        self.canvas.ax.set_ylabel("Frequency [MHz]")

        # Save full-extent limits as the "home" view (used for Reset Selection after zoom/pan)
        self._home_view = {
            "xlim": self.canvas.ax.get_xlim(),
            "ylim": self.canvas.ax.get_ylim(),
        }

        self.format_axes()

        # Keep graph styling consistent (fonts) then apply theme colors
        self.apply_graph_properties_live()
        self._apply_mpl_theme()

        self.canvas.draw_idle()
        self._refresh_accel_plot(data=burst_isolated, title="Isolated Burst", preserve_view=False)
        if self._hardware_mode_enabled():
            self._show_accel_canvas()
        else:
            self._show_plot_canvas()

        # Replace display data with isolated data
        self.noise_reduced_data = burst_isolated

        self.statusBar().showMessage("Burst isolated using lasso", 4000)

    def plot_max_intensities(self):
        # Ensure any active lasso from the main plot is fully disconnected
        if getattr(self, "lasso", None):
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None

        if self.noise_reduced_data is None:
            print("No burst-isolated data available.")
            return

        # Diagnostics
        print("🖥️ Screens:", QGuiApplication.screens())
        print("🎯 Creating MaxIntensityPlotDialog...")

        try:
            session = None
            cached = getattr(self, "_max_intensity_state", None)
            if (
                isinstance(cached, dict)
                and cached.get("time_channels") is not None
                and cached.get("freqs") is not None
                and (cached.get("source_filename") in (None, self.filename))
            ):
                time_channel_number = np.asarray(cached["time_channels"], dtype=float)
                max_intensity_freqs = np.asarray(cached["freqs"], dtype=float)
                session = cached
            else:
                data = self.noise_reduced_data
                _ny, nx = data.shape
                time_channel_number = np.linspace(0, nx, nx)
                max_intensity_freqs = self.freqs[np.argmax(data, axis=0)]

            # Safely create the dialog
            dialog = MaxIntensityPlotDialog(
                time_channel_number,
                max_intensity_freqs,
                self.filename,
                parent=self,
                session=session,
            )
            dialog.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)

            # Connect to GC after close
            dialog.finished.connect(lambda: gc.collect())

            dialog.exec()

            try:
                self._max_intensity_state = dialog.session_state()
                if isinstance(self._max_intensity_state, dict):
                    self._max_intensity_state["source_filename"] = self.filename
            except Exception:
                pass

            self._mark_project_dirty()
            gc.collect()

        except Exception as e:
            print(f"❌ Error showing MaxIntensityPlotDialog: {e}")

    def _pixmap_to_rgba_array(self, pixmap: QPixmap) -> np.ndarray:
        image = pixmap.toImage().convertToFormat(QImage.Format_RGBA8888)
        return self._qimage_to_rgba_array(image)

    def _qimage_to_rgba_array(self, image: QImage) -> np.ndarray:
        width = image.width()
        height = image.height()
        ptr = image.bits()
        ptr.setsize(image.sizeInBytes())
        arr = np.frombuffer(ptr, dtype=np.uint8).reshape((height, image.bytesPerLine()))
        return arr[:, : width * 4].reshape((height, width, 4)).copy()

    def _export_hardware_visible_plot(self, file_path: str, ext_final: str) -> None:
        ext = str(ext_final or "").lower()
        plot_item = self.accel_canvas.export_plot_item()
        if plot_item is None:
            raise RuntimeError("Hardware plot is not available for export.")

        try:
            import pyqtgraph.exporters as pg_exporters
        except Exception:
            pg_exporters = None

        if pg_exporters is not None:
            raster_exts = {"png", "tif", "tiff", "jpg", "jpeg", "bmp", "webp"}
            if ext in raster_exts:
                exporter = pg_exporters.ImageExporter(plot_item)
                try:
                    params = exporter.parameters()
                    width = max(1, int(self.accel_canvas.width() * self.accel_canvas.devicePixelRatioF()))
                    params["width"] = max(width, 1400)
                except Exception:
                    pass
                exporter.export(file_path)
                return

            if ext == "svg":
                exporter = pg_exporters.SVGExporter(plot_item)
                exporter.export(file_path)
                return

            if ext in {"pdf", "eps"}:
                temp_png = None
                try:
                    fd, temp_png = tempfile.mkstemp(suffix=".png")
                    os.close(fd)
                    exporter = pg_exporters.ImageExporter(plot_item)
                    try:
                        params = exporter.parameters()
                        width = max(1, int(self.accel_canvas.width() * self.accel_canvas.devicePixelRatioF()))
                        params["width"] = max(width, 1800)
                    except Exception:
                        pass
                    exporter.export(temp_png)

                    img = QImage(temp_png)
                    if img.isNull():
                        raise RuntimeError("Failed to capture hardware plot image.")

                    if ext == "pdf":
                        writer = QPdfWriter(file_path)
                        writer.setResolution(300)
                        painter = QPainter(writer)
                        target = writer.pageLayout().paintRectPixels(writer.resolution())
                        scaled = img.scaled(target.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        x = target.x() + (target.width() - scaled.width()) // 2
                        y = target.y() + (target.height() - scaled.height()) // 2
                        painter.drawImage(x, y, scaled)
                        painter.end()
                        return

                    rgba = self._qimage_to_rgba_array(img.convertToFormat(QImage.Format_RGBA8888))
                    fig = Figure(figsize=(rgba.shape[1] / 300.0, rgba.shape[0] / 300.0), dpi=300)
                    ax = fig.add_axes([0, 0, 1, 1])
                    ax.imshow(rgba)
                    ax.axis("off")
                    fig.savefig(file_path, dpi=300, bbox_inches="tight", pad_inches=0, format="eps")
                    return
                finally:
                    if temp_png:
                        try:
                            os.remove(temp_png)
                        except Exception:
                            pass

        # Last-resort fallback (can be blank on some OpenGL drivers)
        pixmap = self.accel_canvas.grab()
        if pixmap.isNull():
            raise RuntimeError("Could not capture the accelerated plot image.")
        if not pixmap.save(file_path, ext.upper() if ext else "PNG"):
            raise RuntimeError(f"Failed to save image as {ext}.")

    def _pick_export_path_for_figure(self, caption: str, default_name: str, filters: str, default_filter: str = None):
        """
        Hardware-acceleration mode uses native save dialog directly because
        some Linux/OpenGL combinations can hang with the non-native dialog.
        """
        if self._hardware_mode_enabled():
            path, chosen_filter = QFileDialog.getSaveFileName(
                self,
                caption,
                default_name,
                filters,
                default_filter or "",
            )
            if not path:
                return "", ""

            ext = os.path.splitext(path)[1].lstrip(".").lower()
            if not ext:
                ext = _ext_from_filter(chosen_filter) or "png"
                path = f"{path}.{ext}"
            return path, ext

        return pick_export_path(
            self,
            caption,
            default_name,
            filters,
            default_filter=default_filter,
        )

    def export_figure(self):

        if not self.filename:
            QMessageBox.warning(self, "No File Loaded", "Load a FITS file before exporting.")
            return

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        if self._hardware_mode_enabled():
            base_title = str(self.graph_title_override or self._default_graph_title(self.current_plot_type)).strip()
            default_name = self._sanitize_export_stem(base_title)
        else:
            default_name = self._sanitize_export_stem(self._current_graph_title_for_export())

        file_path, ext = self._pick_export_path_for_figure(
            "Export Figure",
            default_name,
            formats,
            default_filter="PNG (*.png)",
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        def normalize_ext(ext_value: str) -> str:
            """
            Accepts values like: 'png', '.png', 'PNG (*.png)'
            Returns: 'png'
            """
            if not ext_value:
                return "png"
            s = str(ext_value).strip().lower()
            if s.startswith("."):
                s = s[1:]
            m = re.search(r"\*\.(\w+)", s)
            if m:
                return m.group(1).lower()
            return s

        try:
            root, current_ext = os.path.splitext(file_path)

            # If user did not type an extension, add one based on returned ext
            if current_ext == "":
                ext_final = normalize_ext(ext)
                file_path = f"{file_path}.{ext_final}"
            else:
                ext_final = current_ext.lower().lstrip(".")

            if self._hardware_mode_enabled():
                self._export_hardware_visible_plot(file_path, ext_final)
            else:
                self.canvas.figure.savefig(
                    file_path,
                    dpi=300,
                    bbox_inches="tight",
                    format=ext_final
                )

            QMessageBox.information(self, "Export Complete", f"Figure saved:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"An error occurred:\n{e}")

    def _recommend_bitpix_for_export(self, data: np.ndarray) -> int:
        """
        Recommend a FITS BITPIX (8/16/32) based on the data range.

        CALLISTO raw files are typically 8-bit, but combined/processed data can exceed that.
        JavaViewer cannot read BITPIX=-64, so we avoid float64 exports.
        """
        try:
            arr = np.asarray(data)
            if arr.size == 0:
                return 16
            finite = np.isfinite(arr)
            if not np.any(finite):
                return 16
            mn = float(np.nanmin(arr))
            mx = float(np.nanmax(arr))
        except Exception:
            return 16

        if 0.0 <= mn and mx <= 255.0:
            return 8
        if -32768.0 <= mn and mx <= 32767.0:
            return 16
        return 32

    def _cast_data_for_bitpix(self, data: np.ndarray, bitpix: int) -> np.ndarray:
        arr = np.asarray(data)
        arr = np.nan_to_num(arr, nan=0.0, posinf=0.0, neginf=0.0)

        if int(bitpix) == 8:
            out = np.rint(arr)
            out = np.clip(out, 0, 255)
            return out.astype(np.uint8, copy=False)

        if int(bitpix) == 16:
            info = np.iinfo(np.int16)
            out = np.rint(arr)
            out = np.clip(out, info.min, info.max)
            return out.astype(np.int16, copy=False)

        if int(bitpix) == 32:
            info = np.iinfo(np.int32)
            out = np.rint(arr)
            out = np.clip(out, info.min, info.max)
            return out.astype(np.int32, copy=False)

        raise ValueError(f"Unsupported BITPIX: {bitpix}")

    def _sanitize_primary_header_for_export(self, hdr: fits.Header) -> fits.Header:
        # Structural/scaling keywords are determined from the data by Astropy.
        for key in ("SIMPLE", "BITPIX", "NAXIS", "NAXIS1", "NAXIS2", "EXTEND", "BSCALE", "BZERO", "BLANK"):
            try:
                hdr.remove(key, ignore_missing=True, remove_all=True)
            except Exception:
                pass
        return hdr

    def _axis_kind_from_name(self, name: str) -> str | None:
        n = str(name or "").strip().lower()
        if n in ("freq", "frequency", "freqs", "frequency_mhz", "freq_mhz"):
            return "freq"
        if n in ("time", "times", "time_s", "time_sec", "seconds", "sec"):
            return "time"
        if "freq" in n:
            return "freq"
        if n.startswith("time"):
            return "time"
        return None

    def _update_axis_table_hdu(self, hdu, freqs: np.ndarray, times: np.ndarray) -> bool:
        data = getattr(hdu, "data", None)
        if data is None:
            return False
        dtype = getattr(data, "dtype", None)
        names = list(getattr(dtype, "names", []) or [])
        if not names:
            return False

        axis_map: dict[str, np.ndarray] = {}
        for name in names:
            kind = self._axis_kind_from_name(name)
            if kind == "freq":
                axis_map[name] = freqs
            elif kind == "time":
                axis_map[name] = times

        if not axis_map:
            extname = str(hdu.header.get("EXTNAME", "")).lower()
            if "freq" in extname and len(names) == 1:
                axis_map[names[0]] = freqs
            elif "time" in extname and len(names) == 1:
                axis_map[names[0]] = times

        if not axis_map:
            return False

        old_rows = int(data.shape[0]) if hasattr(data, "shape") and len(data.shape) > 0 else 0
        row_lengths = []
        for name, axis in axis_map.items():
            field_dtype = dtype.fields[name][0]
            if field_dtype.shape == ():
                row_lengths.append(len(axis))

        if row_lengths:
            nrows = row_lengths[0]
            for length in row_lengths[1:]:
                if length != nrows:
                    nrows = max(row_lengths)
                    break
        else:
            nrows = old_rows if old_rows > 0 else 1

        new_descr = []
        for name in names:
            field_dtype = dtype.fields[name][0]
            base = field_dtype.base
            if name in axis_map:
                if field_dtype.shape == ():
                    new_descr.append((name, base))
                else:
                    new_descr.append((name, base, (len(axis_map[name]),)))
            else:
                if field_dtype.shape == ():
                    new_descr.append((name, field_dtype))
                else:
                    new_descr.append((name, base, field_dtype.shape))

        new_dtype = np.dtype(new_descr)
        new_data = np.zeros(nrows, dtype=new_dtype)

        for name, axis in axis_map.items():
            axis_arr = np.asarray(axis)
            target = new_data[name]
            if target.ndim == 1:
                axis_cast = axis_arr.astype(target.dtype, copy=False)
                if axis_cast.shape[0] < nrows:
                    pad = np.zeros(nrows, dtype=target.dtype)
                    pad[:axis_cast.shape[0]] = axis_cast
                    axis_cast = pad
                elif axis_cast.shape[0] > nrows:
                    axis_cast = axis_cast[:nrows]
                new_data[name] = axis_cast
            else:
                vec_len = target.shape[1]
                axis_cast = axis_arr.astype(target.dtype, copy=False)
                if axis_cast.shape[0] < vec_len:
                    pad = np.zeros(vec_len, dtype=target.dtype)
                    pad[:axis_cast.shape[0]] = axis_cast
                    axis_cast = pad
                elif axis_cast.shape[0] > vec_len:
                    axis_cast = axis_cast[:vec_len]
                new_data[name][:] = axis_cast

        for name in names:
            if name in axis_map:
                continue
            try:
                old_col = data[name]
                if old_col.shape == new_data[name].shape:
                    new_data[name] = old_col
                else:
                    if old_col.size > 0:
                        new_data[name][0] = old_col[0]
                        if new_data[name].shape[0] > 1:
                            new_data[name][1:] = new_data[name][0]
            except Exception:
                pass

        hdu.data = new_data
        try:
            hdu.update_header()
        except Exception:
            pass
        return True

    def _build_export_hdul_from_template(
        self,
        template_hdul: fits.HDUList,
        primary: fits.PrimaryHDU,
        freqs: np.ndarray,
        times: np.ndarray,
    ) -> tuple[fits.HDUList, bool]:
        new_hdus = [primary]
        updated_any = False
        for hdu in template_hdul[1:]:
            new_hdu = hdu.copy()
            if isinstance(new_hdu, (fits.BinTableHDU, fits.TableHDU)):
                if self._update_axis_table_hdu(new_hdu, freqs, times):
                    updated_any = True
            new_hdus.append(new_hdu)
        return fits.HDUList(new_hdus), updated_any

    def export_to_fits(self):
        if self.raw_data is None or self.freqs is None or self.time is None:
            QMessageBox.warning(self, "No Data", "Load a FITS file before exporting.")
            return

        # Pick exactly what is currently shown
        data_to_save = getattr(self, "current_display_data", None)
        if data_to_save is None:
            data_to_save = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data

        # Choose BITPIX for compatibility (JavaViewer does not support BITPIX=-64)
        rec_bitpix = self._recommend_bitpix_for_export(data_to_save)
        bitpix_items = [
            f"Auto (Recommended: {rec_bitpix})",
            "8 (unsigned byte)",
            "16 (signed int16)",
            "32 (signed int32)",
        ]
        chosen, ok = QInputDialog.getItem(
            self,
            "Export FITS - BITPIX",
            "Choose BITPIX for the exported FITS:",
            bitpix_items,
            0,
            False,
        )
        if not ok:
            return

        bitpix = rec_bitpix
        try:
            if isinstance(chosen, str) and chosen.strip().startswith("8"):
                bitpix = 8
            elif isinstance(chosen, str) and chosen.strip().startswith("16"):
                bitpix = 16
            elif isinstance(chosen, str) and chosen.strip().startswith("32"):
                bitpix = 32
        except Exception:
            bitpix = rec_bitpix

        try:
            export_data = self._cast_data_for_bitpix(data_to_save, bitpix)
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not convert data for BITPIX={bitpix}:\n{e}")
            return

        default_name = f"{self._sanitize_export_stem(self._current_graph_title_for_export())}.fit"

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to FITS",
            default_name,
            "FITS files (*.fit *.fits *.fit.gz *.fits.gz)"
        )
        if not save_path:
            return

        # Ensure extension
        lower = save_path.lower()
        if not (lower.endswith(".fit") or lower.endswith(".fits") or lower.endswith(".fit.gz") or lower.endswith(
                ".fits.gz")):
            save_path += ".fit"

        # Header template
        if self._fits_header0 is not None:
            hdr0 = self._fits_header0.copy()
        else:
            hdr0 = fits.Header()

        hdr0 = self._sanitize_primary_header_for_export(hdr0)

        # Mark what this file is
        hdr0["HISTORY"] = "Exported by e-CALLISTO FITS Analyzer"
        hdr0["HISTORY"] = f"Export plot type: {self.current_plot_type}"
        hdr0["HISTORY"] = f"Units shown: {'dB' if getattr(self, 'use_db', False) else 'Digits'}"

        if self._is_combined:
            hdr0["COMBINED"] = True
            if self._combined_mode:
                hdr0["COMBMETH"] = str(self._combined_mode)
            hdr0["NFILES"] = len(self._combined_sources) if self._combined_sources else 0
            if self._combined_sources:
                hdr0["HISTORY"] = f"First source: {os.path.basename(self._combined_sources[0])}"
                hdr0["HISTORY"] = f"Last source: {os.path.basename(self._combined_sources[-1])}"
        else:
            hdr0["COMBINED"] = False
            if self._fits_source_path:
                hdr0["HISTORY"] = f"Source: {os.path.basename(self._fits_source_path)}"

        # Save BUNIT if you want the file to be self-describing
        hdr0["BUNIT"] = "dB" if getattr(self, "use_db", False) else "Digits"

        primary = fits.PrimaryHDU(data=export_data, header=hdr0)
        try:
            primary.header["BSCALE"] = 1
            primary.header["BZERO"] = 0
        except Exception:
            pass
        try:
            primary.header["DATAMIN"] = float(np.nanmin(export_data))
            primary.header["DATAMAX"] = float(np.nanmax(export_data))
        except Exception:
            pass

        freqs = np.asarray(self.freqs, dtype=np.float32)
        times = np.asarray(self.time, dtype=np.float32)

        hdul = None
        updated_any = False
        template_path = self._fits_source_path
        if not template_path and self._combined_sources:
            template_path = self._combined_sources[0]
        if template_path and os.path.exists(template_path):
            try:
                with fits.open(template_path, memmap=False) as tmpl:
                    hdul, updated_any = self._build_export_hdul_from_template(tmpl, primary, freqs, times)
            except Exception:
                hdul = None
                updated_any = False

        if hdul is not None and self._is_combined and not updated_any:
            hdul = None

        if hdul is None:
            try:
                cols = fits.ColDefs([
                    fits.Column(name="FREQUENCY", format=f"{freqs.size}E", array=[freqs]),
                    fits.Column(name="TIME", format=f"{times.size}E", array=[times]),
                ])
                axis_hdu = fits.BinTableHDU.from_columns(cols)
                axis_hdu.header["EXTNAME"] = "AXIS"
                hdul = fits.HDUList([primary, axis_hdu])
            except Exception:
                hdul = fits.HDUList([primary])

        try:
            hdul[0].header["EXTEND"] = True if len(hdul) > 1 else False
        except Exception:
            pass

        try:
            hdul.writeto(save_path, overwrite=True, output_verify="silentfix")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write FITS file:\n{e}")
            return

        self.statusBar().showMessage(f"Exported FITS (BITPIX={bitpix}): {os.path.basename(save_path)}", 5000)

    def reset_all(self):
        self.noise_smooth_timer.stop()
        self.noise_commit_timer.stop()
        self._noise_undo_pending = False
        self._noise_slider_drag_active = False
        if self._hardware_mode_enabled():
            self._show_accel_canvas()
        else:
            self._show_plot_canvas()
        try:
            self.accel_canvas.clear()
        except Exception:
            pass

        # Safely remove colorbar
        try:
            if self.current_colorbar and self.current_colorbar.ax:
                self.current_colorbar.remove()
        except Exception:
            pass
        self.current_colorbar = None

        try:
            if self.current_cax:
                self.current_cax.remove()
        except Exception:
            pass
        self.current_cax = None

        # Clear canvas
        self.canvas.ax.clear()
        self.canvas.draw()

        # Clear data
        self.raw_data = None
        self._invalidate_noise_cache()
        self.freqs = None
        self.time = None
        self.filename = ""
        self.noise_reduced_data = None
        self.noise_reduced_original = None
        self.lasso_mask = None
        self.current_plot_type = "Raw"
        self.current_display_data = None
        self.noise_vmin = None
        self.noise_vmax = None

        # FITS metadata / provenance
        self._fits_header0 = None
        self._fits_source_path = None
        self._is_combined = False
        self._combined_mode = None
        self._combined_sources = []
        self.ut_start_sec = None
        self._home_view = None
        self._pan_start_view = None

        # Reset GUI
        self.statusBar().showMessage("All reset", 4000)

        # Tool bar
        self._sync_toolbar_enabled_states()
        self.graph_group.setEnabled(False)

        if self.canvas.ax:
            self.canvas.ax.set_xlim(0, 1)
            self.canvas.ax.set_ylim(1, 0)

        self._max_intensity_state = None
        self._set_project_clean(None)

        print("Application reset to initial state.")

    def reset_to_raw(self):
        if self.raw_data is None:
            QMessageBox.warning(self, "No Data", "Load a FITS file first.")
            return

        had_processed = (
            self.noise_reduced_data is not None
            or self.noise_reduced_original is not None
            or self.lasso_mask is not None
            or self.lasso_active
            or bool(getattr(self, "lasso", None))
            or self.lower_slider.value() != 0
            or self.upper_slider.value() != 0
        )
        if had_processed:
            self._push_undo_state()

        self.noise_smooth_timer.stop()
        self.noise_commit_timer.stop()
        self._noise_undo_pending = False
        self._noise_slider_drag_active = False
        if self._hardware_mode_enabled():
            self._show_accel_canvas()
        else:
            self._show_plot_canvas()

        if getattr(self, "lasso", None):
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None
        self.lasso_active = False
        try:
            self.accel_canvas.stop_interaction_capture()
            self.accel_canvas.clear_overlays()
        except Exception:
            pass

        # Ensure pan handlers are restored if lasso activation had disconnected them.
        try:
            self.canvas.mpl_disconnect(self._cid_press)
        except Exception:
            pass
        try:
            self.canvas.mpl_disconnect(self._cid_motion)
        except Exception:
            pass
        try:
            self.canvas.mpl_disconnect(self._cid_release)
        except Exception:
            pass
        self._cid_press = self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self._cid_motion = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
        self._cid_release = self.canvas.mpl_connect("button_release_event", self.on_mouse_release)

        self.noise_reduced_data = None
        self.noise_reduced_original = None
        self.lasso_mask = None
        self.noise_vmin = None
        self.noise_vmax = None

        self.lower_slider.blockSignals(True)
        self.upper_slider.blockSignals(True)
        try:
            self.lower_slider.setValue(0)
            self.upper_slider.setValue(0)
        finally:
            self.lower_slider.blockSignals(False)
            self.upper_slider.blockSignals(False)

        self.plot_data(self.raw_data, title="Raw")
        if had_processed:
            self._mark_project_dirty()
        self.statusBar().showMessage("Reset to raw", 3000)
        self._sync_toolbar_enabled_states()

    def _clear_drift_overlays(self, keep_view: bool = True) -> bool:
        had_drift_points = bool(getattr(self, "drift_points", []))
        self.drift_points = []

        if self._hardware_mode_enabled():
            try:
                self.accel_canvas.stop_interaction_capture()
            except Exception:
                pass
            try:
                self.accel_canvas.show_drift_points([], with_segments=False)
                self.accel_canvas.clear_overlays()
            except Exception:
                pass
            return had_drift_points

        # Matplotlib path
        cid = getattr(self, "drift_click_cid", None)
        if cid is not None:
            try:
                self.canvas.mpl_disconnect(cid)
            except Exception:
                pass
            self.drift_click_cid = None

        if had_drift_points and self.raw_data is not None:
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, keep_view=keep_view)
            return True

        try:
            legend = self.canvas.ax.get_legend()
            if legend is not None:
                legend.remove()
                self.canvas.draw_idle()
        except Exception:
            pass
        return had_drift_points

    def show_about_dialog(self):
        QMessageBox.information(
            self,
            "About e-Callisto FITS Analyzer",
            "e-CALLISTO FITS Analyzer version 2.0.\n\n"
            "Developed by Sahan S Liyanage\n\n"
            "Astronomical and Space Science Unit\n"
            "University of Colombo, Sri Lanka\n\n"
            "2026©Copyright, All Rights Reserved."
        )

    def reset_selection(self):
        had_drift_points = self._clear_drift_overlays(keep_view=True)

        if self.noise_reduced_original is not None:
            self._push_undo_state()
            self.noise_reduced_data = self.noise_reduced_original.copy()
            if self.time is not None and self.freqs is not None:
                self.plot_data(self.noise_reduced_data, title="Background Subtracted")
            self.lasso_mask = None
            self.lasso = None
            if had_drift_points:
                self.statusBar().showMessage("Selection reset (drift markers cleared)", 4000)
            else:
                self.statusBar().showMessage("Selection Reset", 4000)
            print("Lasso selection reset. Original noise-reduced data restored.")
            self._sync_toolbar_enabled_states()
        else:
            # If no selection exists, treat this as a "reset view" (home) action after zoom/pan.
            view_reset = False
            try:
                cur_view = self._capture_view()
                home_view = getattr(self, "_home_view", None)
                if cur_view and home_view and (not self._views_close(cur_view, home_view)):
                    self._push_undo_view(cur_view)
                    self._restore_view(home_view)
                    if self._hardware_mode_enabled():
                        self._show_accel_canvas()
                    else:
                        self.canvas.draw_idle()
                    view_reset = True
                    if had_drift_points:
                        self.statusBar().showMessage("View reset (drift markers cleared)", 2500)
                    else:
                        self.statusBar().showMessage("View reset", 2500)
                elif had_drift_points:
                    self.statusBar().showMessage("Drift markers reset", 2500)
                else:
                    print("No noise-reduced backup found. Reset skipped.")
            finally:
                self._sync_toolbar_enabled_states()

    def open_combine_freq_window(self):
        dialog = CombineFrequencyDialog(self)
        dialog.exec()

    def open_combine_time_window(self):
        dialog = CombineTimeDialog(self)
        dialog.exec()

    def _set_checked_if_exists(self, attr_name: str, checked: bool):
        obj = getattr(self, attr_name, None)
        if obj is None:
            return
        try:
            was_blocked = obj.blockSignals(True)
            obj.setChecked(checked)
        except Exception:
            pass
        finally:
            try:
                obj.blockSignals(was_blocked)
            except Exception:
                pass

    def set_axis_to_seconds(self):
        self.use_utc = False

        # Old Graph-menu actions (may not exist anymore)
        self._set_checked_if_exists("xaxis_sec_action", True)
        self._set_checked_if_exists("xaxis_ut_action", False)

        # If you are using radio buttons, store them as these names (optional)
        self._set_checked_if_exists("xaxis_sec_radio", True)
        self._set_checked_if_exists("xaxis_ut_radio", False)
        self._set_checked_if_exists("time_sec_radio", True)
        self._set_checked_if_exists("time_ut_radio", False)

        if self.raw_data is not None:
            self._mark_project_dirty()
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def set_axis_to_utc(self):
        self.use_utc = True

        # Old Graph-menu actions (may not exist anymore)
        self._set_checked_if_exists("xaxis_sec_action", False)
        self._set_checked_if_exists("xaxis_ut_action", True)

        # If you are using radio buttons, store them as these names (optional)
        self._set_checked_if_exists("xaxis_sec_radio", False)
        self._set_checked_if_exists("xaxis_ut_radio", True)
        self._set_checked_if_exists("time_sec_radio", False)
        self._set_checked_if_exists("time_ut_radio", True)

        if self.raw_data is not None:
            self._mark_project_dirty()
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def format_axes(self):
        if self.use_utc and self.ut_start_sec is not None:
            def format_func(x, pos):
                # Show seconds when viewing a short span (e.g. short files or zoomed-in region)
                try:
                    x0, x1 = self.canvas.ax.get_xlim()
                    span = abs(float(x1) - float(x0))
                except Exception:
                    span = None

                show_seconds = (span is not None) and (span <= 5 * 60)

                total_seconds = float(self.ut_start_sec) + float(x)
                total_seconds_i = int(round(total_seconds))

                hours = int(total_seconds_i // 3600) % 24
                minutes = int((total_seconds_i % 3600) // 60)
                seconds = int(total_seconds_i % 60)

                if show_seconds:
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                return f"{hours:02d}:{minutes:02d}"

            self.canvas.ax.xaxis.set_major_formatter(FuncFormatter(format_func))
            self.canvas.ax.set_xlabel("Time [UT]")
        else:
            self.canvas.ax.xaxis.set_major_formatter(ScalarFormatter())
            self.canvas.ax.set_xlabel("Time [s]")

        self.canvas.ax.figure.canvas.draw()
        if getattr(self, "accel_canvas", None) is not None and self.accel_canvas.is_available:
            self.accel_canvas.set_time_mode(self.use_utc, self.ut_start_sec)

    def _show_import_progress_dialog(self, total_steps: int):
        self._close_import_progress_dialog()

        parent = self
        downloader = getattr(self, "downloader_dialog", None)
        if downloader is not None:
            try:
                if downloader.isVisible():
                    parent = downloader
            except Exception:
                pass

        dlg = QProgressDialog("Downloading selected FITS files...", "", 0, max(1, int(total_steps)), parent)
        dlg.setWindowTitle("Importing FITS Files")
        dlg.setWindowModality(Qt.ApplicationModal)
        dlg.setWindowFlag(Qt.WindowStaysOnTopHint, True)
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setValue(0)
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

        self._import_progress_dialog = dlg

    def _close_import_progress_dialog(self):
        dlg = getattr(self, "_import_progress_dialog", None)
        if dlg is None:
            return
        try:
            dlg.close()
            dlg.deleteLater()
        except Exception:
            pass
        self._import_progress_dialog = None

    @Slot(str)
    def _on_import_progress_text(self, text: str):
        dlg = getattr(self, "_import_progress_dialog", None)
        if dlg is None:
            return
        dlg.setLabelText(str(text or "Importing FITS files..."))

    @Slot(int, int)
    def _on_import_progress_range(self, minimum: int, maximum: int):
        dlg = getattr(self, "_import_progress_dialog", None)
        if dlg is None:
            return
        mn = int(minimum)
        mx = max(int(maximum), mn)
        dlg.setRange(mn, mx)
        if dlg.value() < mn:
            dlg.setValue(mn)

    @Slot(int)
    def _on_import_progress_value(self, value: int):
        dlg = getattr(self, "_import_progress_dialog", None)
        if dlg is None:
            return
        v = int(value)
        if v > dlg.maximum():
            dlg.setMaximum(v)
        dlg.setValue(v)

    def _cleanup_import_worker(self):
        thread = getattr(self, "_import_thread", None)
        if thread is not None:
            try:
                thread.deleteLater()
            except Exception:
                pass
        self._import_thread = None
        self._import_worker = None

    def _emit_downloader_import_success(self):
        dlg = getattr(self, "downloader_dialog", None)
        if dlg is None:
            return
        try:
            dlg.import_success.emit()
        except Exception:
            pass

    def _load_single_import_payload(self, payload: dict):
        data = payload.get("data", None)
        freqs = payload.get("freqs", None)
        time = payload.get("time", None)
        if data is None or freqs is None or time is None:
            raise ValueError("Imported FITS payload is incomplete.")

        self.raw_data = data
        self._invalidate_noise_cache()
        self.freqs = freqs
        self.time = time
        self.filename = str(payload.get("filename", "") or "Imported")

        hdr0 = payload.get("header0", None)
        self._fits_header0 = hdr0.copy() if hdr0 is not None else None
        self._fits_source_path = payload.get("source_path", None)

        self._is_combined = False
        self._combined_mode = None
        self._combined_sources = []
        self.ut_start_sec = payload.get("ut_start_sec", None)

        self.noise_reduced_data = None
        self.noise_reduced_original = None
        self.lasso_mask = None
        self.noise_vmin = None
        self.noise_vmax = None
        self.current_display_data = None
        self._undo_stack.clear()
        self._redo_stack.clear()

        self.plot_data(self.raw_data, title="Raw")
        self._project_path = None
        self._max_intensity_state = None
        self._mark_project_dirty()

    @Slot(object)
    def _on_downloader_import_finished(self, payload):
        self._close_import_progress_dialog()

        try:
            kind = payload.get("kind", "") if isinstance(payload, dict) else ""

            if kind == "single":
                self._load_single_import_payload(payload)
                self._emit_downloader_import_success()
                return

            if kind == "combined":
                combined = payload.get("combined", None) if isinstance(payload, dict) else None
                if combined is None:
                    QMessageBox.critical(self, "Import Failed", "Combined FITS payload is missing.")
                    return
                self.load_combined_into_main(combined)
                self._emit_downloader_import_success()
                return

            if kind == "invalid":
                QMessageBox.warning(
                    self,
                    "Invalid Selection",
                    "Selected files cannot be time-combined or frequency-combined.\n"
                    "Please ensure they are consecutive in time or adjacent in frequency."
                )
                return

            QMessageBox.critical(self, "Import Failed", "Unexpected import result from downloader.")
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", f"Could not import downloaded files:\n{e}")

    @Slot(str)
    def _on_downloader_import_failed(self, message: str):
        self._close_import_progress_dialog()
        QMessageBox.critical(self, "Import Failed", str(message or "Import failed."))

    def process_imported_files(self, urls):
        if not urls:
            QMessageBox.warning(self, "No Files", "No files were received from the downloader.")
            return

        if not self._maybe_prompt_save_dirty():
            return

        if self._import_thread is not None and self._import_thread.isRunning():
            QMessageBox.information(self, "Import In Progress", "Another FITS import is already running.")
            return

        self._show_import_progress_dialog(len(urls))

        self._import_thread = QThread(self)
        self._import_worker = DownloaderImportWorker(urls)
        self._import_worker.moveToThread(self._import_thread)

        self._import_thread.started.connect(self._import_worker.run)
        self._import_worker.progress_text.connect(self._on_import_progress_text)
        self._import_worker.progress_range.connect(self._on_import_progress_range)
        self._import_worker.progress_value.connect(self._on_import_progress_value)
        self._import_worker.finished.connect(self._on_downloader_import_finished)
        self._import_worker.failed.connect(self._on_downloader_import_failed)

        self._import_worker.finished.connect(self._import_thread.quit)
        self._import_worker.failed.connect(self._import_thread.quit)
        self._import_worker.finished.connect(self._import_worker.deleteLater)
        self._import_worker.failed.connect(self._import_worker.deleteLater)
        self._import_thread.finished.connect(self._cleanup_import_worker)

        self._import_thread.start()

    def _stop_rect_zoom(self):
        """Remove rectangle zoom selector safely (if active)."""
        sel = getattr(self, "_rect_selector", None)
        if sel is not None:
            try:
                sel.set_active(False)
                sel.disconnect_events()
            except Exception:
                pass
            self._rect_selector = None
        self.rect_zoom_active = False
        if self._hardware_mode_enabled():
            try:
                self.accel_canvas.cancel_rect_zoom()
                self.accel_canvas.set_navigation_locked(self.nav_locked)
            except Exception:
                pass

    def _on_rect_zoom_select(self, eclick, erelease):
        """Callback when the user finishes drawing the rectangle."""
        if eclick.inaxes != self.canvas.ax or erelease.inaxes != self.canvas.ax:
            self._stop_rect_zoom()
            return

        x0, y0 = eclick.xdata, eclick.ydata
        x1, y1 = erelease.xdata, erelease.ydata
        if x0 is None or x1 is None or y0 is None or y1 is None:
            self._stop_rect_zoom()
            return

        xmin, xmax = sorted([x0, x1])
        ymin, ymax = sorted([y0, y1])

        # Ignore tiny rectangles (prevents accidental clicks)
        if abs(xmax - xmin) < 1e-6 or abs(ymax - ymin) < 1e-6:
            self._stop_rect_zoom()
            return

        ax = self.canvas.ax
        prev_view = self._capture_view()
        ax.set_xlim(xmin, xmax)
        ax.set_ylim(ymin, ymax)
        if prev_view:
            self._push_undo_view(prev_view)
        self.canvas.draw_idle()

        self._stop_rect_zoom()
        self.statusBar().showMessage("Zoomed to selected region (still locked).", 2500)
        self._sync_toolbar_enabled_states()

    def _sync_nav_actions(self):
        """Enable/disable Lock/Unlock/Zoom actions to match current state."""
        locked = bool(getattr(self, "nav_locked", False))
        has_plot = getattr(self, "raw_data", None) is not None

        # Zoom rectangle allowed only when locked AND a plot exists
        self.tb_zoom.setEnabled(locked and has_plot)

        # Lock and unlock behave like mutually exclusive buttons
        self.tb_lock.setEnabled((not locked) and has_plot)
        self.tb_unlock.setEnabled(locked and has_plot)

    def lock(self):
        """Disable scroll zoom + panning. Rectangle zoom becomes available."""
        if getattr(self, "raw_data", None) is None:
            self.statusBar().showMessage("Load a FITS file first.", 2500)
            return

        self.nav_locked = True
        self._panning = False
        self._last_pan_xy = None
        self._stop_rect_zoom()
        if self._hardware_mode_enabled():
            self.accel_canvas.set_navigation_locked(True)

        self._sync_nav_actions()
        self.statusBar().showMessage("Navigation locked. Use Rectangle Zoom if needed.", 3000)

    def unlock(self):
        """Enable scroll zoom + panning again."""
        if getattr(self, "raw_data", None) is None:
            self.statusBar().showMessage("Load a FITS file first.", 2500)
            return

        self.nav_locked = False
        self._panning = False
        self._last_pan_xy = None
        self._stop_rect_zoom()
        if self._hardware_mode_enabled():
            self.accel_canvas.set_navigation_locked(False)

        self._sync_nav_actions()
        self.statusBar().showMessage("Navigation unlocked. Pan and scroll zoom enabled.", 3000)

    def rectangular_zoom(self):
        """
        Start rectangle zoom tool.
        Requirement: only allowed when lock is active.
        """
        if getattr(self, "raw_data", None) is None:
            self.statusBar().showMessage("Load a FITS file first.", 2500)
            return

        if not getattr(self, "nav_locked", False):
            self.statusBar().showMessage("Click Lock first to enable Rectangle Zoom.", 3500)
            return

        # Do not conflict with lasso
        if getattr(self, "lasso_active", False):
            self.statusBar().showMessage("Finish the lasso tool first.", 3000)
            return

        # Stop any existing rectangle selector
        self._stop_rect_zoom()

        if self._hardware_mode_enabled():
            self.rect_zoom_active = True
            self.accel_canvas.start_rect_zoom_once()
            self.statusBar().showMessage("Drag a rectangle on the plot to zoom.", 4000)
            return

        ax = self.canvas.ax
        self.rect_zoom_active = True

        # Create rectangle selector
        self._rect_selector = RectangleSelector(
            ax,
            self._on_rect_zoom_select,
            useblit=True,
            button=[1],  # left mouse only
            interactive=False
        )

        self.statusBar().showMessage("Drag a rectangle on the plot to zoom.", 4000)

    def launch_downloader(self):
        self.downloader_dialog = CallistoDownloaderApp()
        self.downloader_dialog.import_request.connect(self.process_imported_files)

        self.import_success_signal = lambda: self.downloader_dialog.accept()
        self.downloader_dialog.import_success.connect(self.import_success_signal)

        self.downloader_dialog.exec()

    def open_goes_xrs_window(self):
        self.goes_window = GoesXrsWindow()
        self.goes_window.show()

    def open_soho_lasco_window(self):
        self.open_cme_viewer()

    def _capture_state(self):
        """Capture the current application state for Undo/Redo."""
        state = {
            "raw_data": None if self.raw_data is None else self.raw_data.copy(),
            "noise_reduced_data": None if self.noise_reduced_data is None else self.noise_reduced_data.copy(),
            "noise_reduced_original": None if self.noise_reduced_original is None else self.noise_reduced_original.copy(),
            "lasso_mask": None if self.lasso_mask is None else self.lasso_mask.copy(),
            "freqs": None if self.freqs is None else self.freqs.copy(),
            "time": None if self.time is None else self.time.copy(),
            "filename": self.filename,
            "current_plot_type": self.current_plot_type,
            "lower_slider": self.lower_slider.value(),
            "upper_slider": self.upper_slider.value(),
            "use_db": self.use_db,
            "use_utc": self.use_utc,
            "cmap": self.current_cmap_name,
            "view": self._capture_view(),
        }
        return state

    def _entry_kind(self, entry) -> str:
        if isinstance(entry, dict) and entry.get("kind") in ("state", "view"):
            return entry["kind"]
        return "state"

    def _entry_state(self, entry):
        if self._entry_kind(entry) != "state":
            return None
        if isinstance(entry, dict) and "state" in entry:
            return entry["state"]
        return entry

    def _entry_view(self, entry):
        kind = self._entry_kind(entry)
        if kind == "view":
            return entry.get("view") if isinstance(entry, dict) else None
        st = self._entry_state(entry)
        if isinstance(st, dict):
            return st.get("view")
        return None

    def _count_state_entries(self, stack) -> int:
        n = 0
        for e in stack:
            if self._entry_kind(e) == "state":
                n += 1
        return n

    def _trim_history(self):
        # Enforce limit on heavy (state) snapshots by dropping the oldest state entry,
        # along with any older view entries before it.
        while self._count_state_entries(self._undo_stack) > self._max_undo:
            drop_to = None
            for i, e in enumerate(self._undo_stack):
                if self._entry_kind(e) == "state":
                    drop_to = i
                    break
            if drop_to is None:
                break
            del self._undo_stack[: drop_to + 1]

        # Guardrail to prevent unbounded growth from view-history spam
        if len(self._undo_stack) > self._max_history_entries:
            extra = len(self._undo_stack) - self._max_history_entries
            del self._undo_stack[:extra]

        if len(self._redo_stack) > self._max_history_entries:
            extra = len(self._redo_stack) - self._max_history_entries
            del self._redo_stack[:extra]

    def _views_close(self, a, b, tol: float = 1e-6) -> bool:
        if not a or not b:
            return False
        for key in ("xlim", "ylim"):
            try:
                a0, a1 = a.get(key)
                b0, b1 = b.get(key)
            except Exception:
                return False
            try:
                a0 = float(a0)
                a1 = float(a1)
                b0 = float(b0)
                b1 = float(b1)
            except Exception:
                return False
            if abs(a0 - b0) > tol or abs(a1 - b1) > tol:
                return False
        return True

    def _push_undo_view(self, view):
        if not view:
            return
        # Avoid pushing duplicate consecutive view entries
        if self._undo_stack and self._entry_kind(self._undo_stack[-1]) == "view":
            last_view = self._entry_view(self._undo_stack[-1])
            if last_view and self._views_close(last_view, view):
                return

        self._undo_stack.append({"kind": "view", "view": view})
        self._redo_stack.clear()
        self._trim_history()

    def _push_undo_state(self):
        self._undo_stack.append({"kind": "state", "state": self._capture_state()})
        self._redo_stack.clear()
        self._trim_history()
        self._mark_project_dirty()

    def _restore_state(self, state):
        """Restore a previously captured application state."""
        self.raw_data = state["raw_data"]
        self._invalidate_noise_cache()
        self.noise_reduced_data = state["noise_reduced_data"]
        self.noise_reduced_original = state["noise_reduced_original"]
        self.lasso_mask = state["lasso_mask"]
        self.freqs = state["freqs"]
        self.time = state["time"]
        self.filename = state["filename"]
        self.current_plot_type = state["current_plot_type"]
        self.use_db = state["use_db"]
        self.use_utc = state["use_utc"]
        self.current_cmap_name = state["cmap"]

        self.lower_slider.blockSignals(True)
        self.upper_slider.blockSignals(True)
        self.lower_slider.setValue(state["lower_slider"])
        self.upper_slider.setValue(state["upper_slider"])
        self.lower_slider.blockSignals(False)
        self.upper_slider.blockSignals(False)

        if self.raw_data is not None:
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, restore_view=state.get("view"))

    def undo(self):
        if not self._undo_stack:
            self.statusBar().showMessage("Nothing to undo", 2000)
            return
        entry = self._undo_stack.pop()
        kind = self._entry_kind(entry)

        if kind == "view":
            cur_view = self._capture_view()
            if cur_view:
                self._redo_stack.append({"kind": "view", "view": cur_view})
                self._trim_history()

            view = self._entry_view(entry)
            if view:
                self._restore_view(view)
                self.canvas.draw_idle()
        else:
            current = {"kind": "state", "state": self._capture_state()}
            self._redo_stack.append(current)
            self._trim_history()

            state = self._entry_state(entry)
            if state:
                self._restore_state(state)

        self.statusBar().showMessage("Undo", 2000)
        self._sync_toolbar_enabled_states()

    def redo(self):
        if not self._redo_stack:
            self.statusBar().showMessage("Nothing to redo", 2000)
            return
        entry = self._redo_stack.pop()
        kind = self._entry_kind(entry)

        if kind == "view":
            cur_view = self._capture_view()
            if cur_view:
                self._undo_stack.append({"kind": "view", "view": cur_view})
                self._trim_history()

            view = self._entry_view(entry)
            if view:
                self._restore_view(view)
                self.canvas.draw_idle()
        else:
            current = {"kind": "state", "state": self._capture_state()}
            self._undo_stack.append(current)
            self._trim_history()

            state = self._entry_state(entry)
            if state:
                self._restore_state(state)

        self.statusBar().showMessage("Redo", 2000)
        self._sync_toolbar_enabled_states()

    # -----------------------------
    # Project/session Save + Load
    # -----------------------------

    def _mark_project_dirty(self):
        if getattr(self, "_loading_project", False):
            return
        if getattr(self, "raw_data", None) is None:
            return
        self._project_dirty = True
        self._sync_project_actions()

    def _set_project_clean(self, path: str | None):
        self._project_path = path
        self._project_dirty = False
        self._sync_project_actions()

    def _sync_project_actions(self):
        has_data = getattr(self, "raw_data", None) is not None

        for name, enabled in (
            ("save_project_action", has_data),
            ("save_project_as_action", has_data),
        ):
            act = getattr(self, name, None)
            if act is not None:
                act.setEnabled(bool(enabled))

    def _sync_fits_view_actions(self):
        has_data = getattr(self, "raw_data", None) is not None
        act = getattr(self, "view_fits_header_action", None)
        if act is not None:
            act.setEnabled(bool(has_data))

    def _maybe_prompt_save_dirty(self) -> bool:
        if not getattr(self, "_project_dirty", False):
            return True

        resp = QMessageBox.question(
            self,
            "Unsaved Project",
            "You have unsaved project changes.\n\nSave before continuing?",
            QMessageBox.StandardButton.Save
            | QMessageBox.StandardButton.Discard
            | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Save,
        )

        if resp == QMessageBox.StandardButton.Save:
            return bool(self.save_project())
        if resp == QMessageBox.StandardButton.Discard:
            return True
        return False

    def _project_default_filename(self) -> str:
        base = "project"
        if getattr(self, "filename", ""):
            base = os.path.splitext(os.path.basename(self.filename))[0] or base
        return f"{base}.efaproj"

    def _capture_project_payload(self):
        state = self._capture_state()

        arrays = {
            "raw_data": state["raw_data"],
            "noise_reduced_data": state["noise_reduced_data"],
            "noise_reduced_original": state["noise_reduced_original"],
            "lasso_mask": state["lasso_mask"],
            "freqs": state["freqs"],
            "time": state["time"],
        }

        header_txt = None
        try:
            if getattr(self, "_fits_header0", None) is not None:
                header_txt = self._fits_header0.tostring(sep="\n", endcard=True, padding=False)
        except Exception:
            header_txt = None

        graph = {
            "remove_titles": bool(getattr(self, "remove_titles", False)),
            "title_bold": bool(getattr(self, "title_bold", False)),
            "title_italic": bool(getattr(self, "title_italic", False)),
            "axis_bold": bool(getattr(self, "axis_bold", False)),
            "axis_italic": bool(getattr(self, "axis_italic", False)),
            "ticks_bold": bool(getattr(self, "ticks_bold", False)),
            "ticks_italic": bool(getattr(self, "ticks_italic", False)),
            "title_override": str(getattr(self, "graph_title_override", "")),
            "font_family": str(getattr(self, "graph_font_family", "")),
            "tick_font_px": int(getattr(self, "tick_font_px", 11)),
            "axis_label_font_px": int(getattr(self, "axis_label_font_px", 12)),
            "title_font_px": int(getattr(self, "title_font_px", 14)),
        }

        meta = {
            "filename": state["filename"],
            "current_plot_type": state["current_plot_type"],
            "lower_slider": int(state["lower_slider"]),
            "upper_slider": int(state["upper_slider"]),
            "use_db": bool(state["use_db"]),
            "use_utc": bool(state["use_utc"]),
            "ut_start_sec": self.ut_start_sec,
            "cmap": state["cmap"],
            "view": state["view"],
            "noise_vmin": self.noise_vmin,
            "noise_vmax": self.noise_vmax,
            "fits_header": header_txt,
            "fits_source_path": getattr(self, "_fits_source_path", None),
            "is_combined": bool(getattr(self, "_is_combined", False)),
            "combined_mode": getattr(self, "_combined_mode", None),
            "combined_sources": list(getattr(self, "_combined_sources", []) or []),
            "graph": graph,
        }

        # Optional derived analysis state (populated after dialogs)
        if getattr(self, "_max_intensity_state", None):
            meta["max_intensity"] = {"present": True}
            arrays.update({
                "max_time_channels": self._max_intensity_state.get("time_channels"),
                "max_freqs": self._max_intensity_state.get("freqs"),
            })
            meta["max_intensity"].update({
                "fundamental": bool(self._max_intensity_state.get("fundamental", True)),
                "harmonic": bool(self._max_intensity_state.get("harmonic", False)),
                "analyzer": self._max_intensity_state.get("analyzer"),
            })

        return meta, arrays

    def _apply_project_payload(self, meta: dict, arrays: dict):
        self._loading_project = True
        try:
            # Clear undo/redo stacks on project load
            self._undo_stack.clear()
            self._redo_stack.clear()

            self.raw_data = arrays.get("raw_data", None)
            self._invalidate_noise_cache()
            self.noise_reduced_data = arrays.get("noise_reduced_data", None)
            self.noise_reduced_original = arrays.get("noise_reduced_original", None)
            self.lasso_mask = arrays.get("lasso_mask", None)
            self.freqs = arrays.get("freqs", None)
            self.time = arrays.get("time", None)

            # Copies avoid read-only arrays from np.load
            for name in ("raw_data", "noise_reduced_data", "noise_reduced_original", "lasso_mask", "freqs", "time"):
                val = getattr(self, name, None)
                if val is not None:
                    try:
                        setattr(self, name, val.copy())
                    except Exception:
                        pass

            self.filename = meta.get("filename", "") or ""
            self.current_plot_type = self._normalize_plot_type(meta.get("current_plot_type", "Raw"))

            self.use_db = bool(meta.get("use_db", False))
            self.use_utc = bool(meta.get("use_utc", False))
            self.ut_start_sec = meta.get("ut_start_sec", None)
            self.current_cmap_name = meta.get("cmap", "Custom") or "Custom"

            self.noise_vmin = meta.get("noise_vmin", None)
            self.noise_vmax = meta.get("noise_vmax", None)

            self._fits_source_path = meta.get("fits_source_path", None)
            self._is_combined = bool(meta.get("is_combined", False))
            self._combined_mode = meta.get("combined_mode", None)
            self._combined_sources = list(meta.get("combined_sources", []) or [])

            header_txt = meta.get("fits_header", None)
            self._fits_header0 = None
            if header_txt:
                try:
                    self._fits_header0 = fits.Header.fromstring(header_txt, sep="\n")
                except Exception:
                    self._fits_header0 = None

            # Restore widgets without triggering live updates
            try:
                self.lower_slider.blockSignals(True)
                self.upper_slider.blockSignals(True)
                self.lower_slider.setValue(int(meta.get("lower_slider", self.lower_slider.value())))
                self.upper_slider.setValue(int(meta.get("upper_slider", self.upper_slider.value())))
            finally:
                self.lower_slider.blockSignals(False)
                self.upper_slider.blockSignals(False)

            # Units radios
            try:
                self.units_digits_radio.blockSignals(True)
                self.units_db_radio.blockSignals(True)
                self.units_db_radio.setChecked(bool(self.use_db))
                self.units_digits_radio.setChecked(not bool(self.use_db))
            finally:
                self.units_digits_radio.blockSignals(False)
                self.units_db_radio.blockSignals(False)

            # Time-axis radios
            try:
                self.time_sec_radio.blockSignals(True)
                self.time_ut_radio.blockSignals(True)
                self.time_ut_radio.setChecked(bool(self.use_utc))
                self.time_sec_radio.setChecked(not bool(self.use_utc))
            finally:
                self.time_sec_radio.blockSignals(False)
                self.time_ut_radio.blockSignals(False)

            # Colormap combo
            try:
                self.cmap_combo.blockSignals(True)
                if self.current_cmap_name:
                    self.cmap_combo.setCurrentText(self.current_cmap_name)
            finally:
                self.cmap_combo.blockSignals(False)

            # Graph properties
            graph = meta.get("graph", {}) or {}
            self._hw_default_font_sizes_active = False
            try:
                self.remove_titles_chk.blockSignals(True)
                self.title_bold_chk.blockSignals(True)
                self.title_italic_chk.blockSignals(True)
                self.axis_bold_chk.blockSignals(True)
                self.axis_italic_chk.blockSignals(True)
                self.ticks_bold_chk.blockSignals(True)
                self.ticks_italic_chk.blockSignals(True)
                self.title_edit.blockSignals(True)
                self.font_combo.blockSignals(True)
                self.tick_font_spin.blockSignals(True)
                self.axis_font_spin.blockSignals(True)
                self.title_font_spin.blockSignals(True)

                self.remove_titles_chk.setChecked(bool(graph.get("remove_titles", False)))
                self.title_bold_chk.setChecked(bool(graph.get("title_bold", False)))
                self.title_italic_chk.setChecked(bool(graph.get("title_italic", False)))
                self.axis_bold_chk.setChecked(bool(graph.get("axis_bold", False)))
                self.axis_italic_chk.setChecked(bool(graph.get("axis_italic", False)))
                self.ticks_bold_chk.setChecked(bool(graph.get("ticks_bold", False)))
                self.ticks_italic_chk.setChecked(bool(graph.get("ticks_italic", False)))

                self.title_edit.setText(str(graph.get("title_override", "")) or "")

                font_family = str(graph.get("font_family", "")) or ""
                self.font_combo.setCurrentText(font_family if font_family else "Default")

                self.tick_font_spin.setValue(int(graph.get("tick_font_px", self.tick_font_spin.value())))
                self.axis_font_spin.setValue(int(graph.get("axis_label_font_px", self.axis_font_spin.value())))
                self.title_font_spin.setValue(int(graph.get("title_font_px", self.title_font_spin.value())))
            finally:
                self.remove_titles_chk.blockSignals(False)
                self.title_bold_chk.blockSignals(False)
                self.title_italic_chk.blockSignals(False)
                self.axis_bold_chk.blockSignals(False)
                self.axis_italic_chk.blockSignals(False)
                self.ticks_bold_chk.blockSignals(False)
                self.ticks_italic_chk.blockSignals(False)
                self.title_edit.blockSignals(False)
                self.font_combo.blockSignals(False)
                self.tick_font_spin.blockSignals(False)
                self.axis_font_spin.blockSignals(False)
                self.title_font_spin.blockSignals(False)

            # Derived analysis state (optional)
            self._max_intensity_state = None
            if (meta.get("max_intensity") or {}).get("present"):
                self._max_intensity_state = {
                    "time_channels": arrays.get("max_time_channels", None),
                    "freqs": arrays.get("max_freqs", None),
                    "fundamental": bool((meta.get("max_intensity") or {}).get("fundamental", True)),
                    "harmonic": bool((meta.get("max_intensity") or {}).get("harmonic", False)),
                    "analyzer": (meta.get("max_intensity") or {}).get("analyzer"),
                    "source_filename": self.filename,
                }

            # Redraw
            if self.raw_data is not None:
                data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
                self.plot_data(data, title=self.current_plot_type, restore_view=meta.get("view"))
                self.graph_group.setEnabled(True)
                self._sync_toolbar_enabled_states()
        finally:
            self._loading_project = False
            self._sync_project_actions()

    def save_project(self) -> bool:
        if getattr(self, "raw_data", None) is None:
            QMessageBox.information(self, "Save Project", "Load a FITS file first.")
            return False

        if not getattr(self, "_project_path", None):
            return bool(self.save_project_as())

        try:
            meta, arrays = self._capture_project_payload()
            write_project(self._project_path, meta=meta, arrays=arrays)
        except Exception as e:
            QMessageBox.critical(self, "Save Project Failed", f"Could not save project:\n{e}")
            return False

        self._set_project_clean(self._project_path)
        self.statusBar().showMessage(f"Project saved: {os.path.basename(self._project_path)}", 5000)
        return True

    def save_project_as(self) -> bool:
        if getattr(self, "raw_data", None) is None:
            QMessageBox.information(self, "Save Project", "Load a FITS file first.")
            return False

        start_dir = ""
        if getattr(self, "_project_path", None):
            start_dir = os.path.dirname(self._project_path)
        elif getattr(self, "_fits_source_path", None):
            start_dir = os.path.dirname(self._fits_source_path)

        default_name = os.path.join(start_dir, self._project_default_filename()) if start_dir else self._project_default_filename()

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Project As",
            default_name,
            "e-CALLISTO Project (*.efaproj)",
        )
        if not path:
            return False
        if not path.lower().endswith(".efaproj"):
            path += ".efaproj"

        try:
            meta, arrays = self._capture_project_payload()
            write_project(path, meta=meta, arrays=arrays)
        except Exception as e:
            QMessageBox.critical(self, "Save Project Failed", f"Could not save project:\n{e}")
            return False

        self._set_project_clean(path)
        self.statusBar().showMessage(f"Project saved: {os.path.basename(path)}", 5000)
        return True

    def open_project(self):
        if not self._maybe_prompt_save_dirty():
            return

        start_dir = ""
        if getattr(self, "_project_path", None):
            start_dir = os.path.dirname(self._project_path)
        elif getattr(self, "_fits_source_path", None):
            start_dir = os.path.dirname(self._fits_source_path)

        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Project",
            start_dir,
            "e-CALLISTO Project (*.efaproj)",
        )
        if not path:
            return

        try:
            payload = read_project(path)
        except ProjectFormatError as e:
            QMessageBox.critical(self, "Open Project Failed", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Open Project Failed", f"Could not open project:\n{e}")
            return

        self._apply_project_payload(payload.meta, payload.arrays)
        self._set_project_clean(path)
        self.statusBar().showMessage(f"Project loaded: {os.path.basename(path)}", 5000)

    def open_fits_header_viewer(self):
        if getattr(self, "raw_data", None) is None:
            QMessageBox.information(self, "FITS Header", "Load a FITS file first.")
            return

        hdr = getattr(self, "_fits_header0", None)
        if hdr is None:
            hdr = fits.Header()

        base = "fits"
        if getattr(self, "filename", ""):
            base = os.path.splitext(os.path.basename(self.filename))[0] or base
        default_name = f"{base}_header.txt"

        title = "FITS Header"
        if getattr(self, "filename", ""):
            title = f"FITS Header — {self.filename}"

        self._fits_header_viewer = FitsHeaderViewerDialog(
            hdr,
            title=title,
            default_name=default_name,
            parent=self,
        )
        self._fits_header_viewer.show()

    def closeEvent(self, event):
        if not self._maybe_prompt_save_dirty():
            event.ignore()
            return
        try:
            self.noise_smooth_timer.stop()
            self.noise_commit_timer.stop()
        except Exception:
            pass
        super().closeEvent(event)


    def open_cme_viewer(self):
        from src.UI.soho_lasco_viewer import CMEViewer  # import here, not at top
        self._cme_viewer = CMEViewer(parent=self)
        self._cme_viewer.show()


class MaxIntensityPlotDialog(QDialog):
    def __init__(self, time_channels, max_freqs, filename, parent=None, session=None):
        super().__init__(parent)
        self.setWindowTitle("Maximum Intensities for Each Time Channel")
        self.resize(1000, 700)
        self.filename = filename
        self.current_plot_type = "MaxIntensityPlot"
        self._analyzer_state = None

        # Data
        self.time_channels = np.array(time_channels)
        self.freqs = np.array(max_freqs)
        self.selected_mask = np.zeros_like(self.time_channels, dtype=bool)
        self.lasso = None

        # Canvas
        self.canvas = MplCanvas(self, width=10, height=6)
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)
        self.canvas.ax.scatter(self.time_channels, self.freqs, marker="o", s=5, color='red')
        self.canvas.ax.set_xlabel("Time Channel Number")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.set_title("Maximum Intensity for Each Time Channel")
        self.canvas.draw()

        # Buttons
        self.select_button = QPushButton("Select Outliers")
        self.remove_button = QPushButton("Remove Outliers")

        self.fundamental_radio = QRadioButton("Fundamental")
        self.harmonic_radio = QRadioButton("Harmonic")
        self.fundamental_radio.setChecked(True)

        self.mode_group = QButtonGroup(self)
        self.mode_group.addButton(self.fundamental_radio)
        self.mode_group.addButton(self.harmonic_radio)

        self.analyze_button = QPushButton("Analyze Burst")
        self.select_button.setToolTip("Use Lasso tool to select points to remove")
        self.remove_button.setToolTip("Remove previously selected outliers")
        self.select_button.setMinimumWidth(150)
        self.remove_button.setMinimumWidth(150)
        self.analyze_button.setMinimumWidth(150)
        self.select_button.clicked.connect(self.activate_lasso)
        self.remove_button.clicked.connect(self.remove_selected_outliers)

        self.analyze_button.clicked.connect(lambda: self.open_analyze_window(
            fundamental=self.fundamental_radio.isChecked(),
            harmonic=self.harmonic_radio.isChecked()
        ))

        # Layouts
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.select_button)
        button_layout.addWidget(self.remove_button)
        button_layout.addWidget(self.fundamental_radio)
        button_layout.addWidget(self.harmonic_radio)
        button_layout.addWidget(self.analyze_button)
        button_layout.addStretch()

        # Status bar
        self.status = QStatusBar()
        self.status.showMessage("Ready")

        # Menubar
        menubar = QMenuBar(self)
        file_menu = menubar.addMenu("File")
        self.save_action = QAction("Save As", self)
        self.export_action = QAction("Export As", self)
        file_menu.addAction(self.save_action)
        file_menu.addAction(self.export_action)
        self.save_action.triggered.connect(self.save_as_csv)
        self.export_action.triggered.connect(self.export_figure)

        edit_menu = menubar.addMenu("Edit")
        reset_action = QAction("Reset All", self)
        edit_menu.addAction(reset_action)
        reset_action.triggered.connect(self.reset_all)

        analyze_menu = menubar.addMenu("Analyze")
        analyze_action = QAction("Open Analyzer", self)
        analyze_menu.addAction(analyze_action)
        analyze_action.triggered.connect(self.open_analyze_window)

        about_menu = menubar.addMenu("About")
        about_action = QAction("About", self)
        about_action.setMenuRole(QAction.NoRole)
        about_menu.addAction(about_action)
        about_action.triggered.connect(self.show_about_dialog)

        # Main Layout
        layout = QVBoxLayout()
        layout.setMenuBar(menubar)
        layout.addLayout(button_layout)
        layout.addWidget(self.canvas)
        layout.addWidget(self.status)
        self.setLayout(layout)

        # Styling
        self.setStyleSheet("""
            QPushButton { font-size: 13px; padding: 6px 12px; }
            QLabel { font-size: 13px; }
        """)

        # Restore optional session state (radio selections + analyzer state)
        if isinstance(session, dict):
            try:
                self._analyzer_state = session.get("analyzer", None)
                harmonic = bool(session.get("harmonic", False))
                if harmonic:
                    self.harmonic_radio.setChecked(True)
                else:
                    self.fundamental_radio.setChecked(True)
            except Exception:
                pass

    def activate_lasso(self):
        self.canvas.ax.set_title("Draw around outliers to remove")
        self.canvas.draw()

        if self.lasso:
            self.lasso.disconnect_events()

        self.lasso = LassoSelector(self.canvas.ax, onselect=self.on_lasso_select)

    def on_lasso_select(self, verts):
        path = Path(verts)
        points = np.column_stack((self.time_channels, self.freqs))
        self.selected_mask = path.contains_points(points)
        if self.lasso:
            self.lasso.disconnect_events()
            self.lasso = None
        self.status.showMessage(f"{np.sum(self.selected_mask)} points selected", 3000)

    def remove_selected_outliers(self):
        if not np.any(self.selected_mask):
            self.status.showMessage("No points selected for removal", 3000)
            return

        self.time_channels = self.time_channels[~self.selected_mask]
        self.freqs = self.freqs[~self.selected_mask]
        self.selected_mask = np.zeros_like(self.time_channels, dtype=bool)
        self._analyzer_state = None

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)

        self.canvas.ax.scatter(self.time_channels, self.freqs, marker="o", s=5, color='red')
        self.canvas.ax.set_xlabel("Time Channel Number")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.set_title("Filtered Max Intensities")
        self.canvas.draw()

        self.status.showMessage("Selected outliers removed", 3000)

    def session_state(self) -> dict:
        return {
            "time_channels": np.asarray(self.time_channels, dtype=float),
            "freqs": np.asarray(self.freqs, dtype=float),
            "fundamental": bool(self.fundamental_radio.isChecked()),
            "harmonic": bool(self.harmonic_radio.isChecked()),
            "analyzer": self._analyzer_state,
        }

    def reset_all(self):
        # Clear canvas
        self.canvas.ax.clear()
        self.canvas.draw()

        # Clear internal variables
        self.raw_data = None
        self.freqs = None
        self.time = None
        self.filename = ""
        self.noise_reduced_data = None
        self.lasso_mask = None
        self.current_plot_type = "Raw"

        self.statusBar().showMessage("All reset", 4000)

        print("Application reset to initial state.")

    def save_as_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save CSV File", "", "CSV files (*.csv)")
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Time Channel", "Frequency (MHz)"])
                for t, fval in zip(self.time_channels * 0.25, self.freqs):
                    writer.writerow([t, fval])
            self.status.showMessage(f"Saved to {file_path}", 3000)
        except Exception as e:
            self.status.showMessage(f"Error: {e}", 3000)

    def export_figure(self):

        if not self.filename:
            QMessageBox.warning(self, "No File Loaded", "Load a FITS file before exporting.")
            return

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        base_name = self.filename.split(".")[0]
        suffix = self.current_plot_type.replace(" ", "")
        default_name = f"{base_name}_{suffix}"

        file_path, ext = pick_export_path(
            self,
            "Export Figure",
            default_name,
            formats,
            default_filter="PNG (*.png)"
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        try:
            # ✅ If user didn't type an extension, add the one from ext
            root, current_ext = os.path.splitext(file_path)
            if current_ext == "":
                ext = ext.lower().lstrip(".")  # ext should be like "png"
                file_path = f"{file_path}.{ext}"
            else:
                ext = current_ext.lower().lstrip(".")  # use what user typed

            self.canvas.figure.savefig(
                file_path,
                dpi=300,
                bbox_inches="tight",
                format=ext
            )

            QMessageBox.information(self, "Export Complete", f"Figure saved:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"An error occurred:\n{e}")

    def show_about_dialog(self):
        QMessageBox.information(
            self,
            "About e-Callisto FITS Analyzer",
            "e-CALLISTO FITS Analyzer version 2.0.\n\n"
            "Developed by Sahan S Liyanage\n\n"
            "Astronomical and Space Science Unit\n"
            "University of Colombo, Sri Lanka\n\n"
            "2026©Copyright, All Rights Reserved."
        )

    def open_analyze_window(self, fundamental=None, harmonic=None):
        if fundamental is None and harmonic is None:
            fundamental = bool(self.fundamental_radio.isChecked())
            harmonic = bool(self.harmonic_radio.isChecked())

        dialog = AnalyzeDialog(
            self.time_channels,
            self.freqs,
            self.filename,
            fundamental=bool(fundamental),
            harmonic=bool(harmonic),
            parent=self,
            session=self._analyzer_state,
        )
        dialog.exec()
        try:
            self._analyzer_state = dialog.session_state()
        except Exception:
            pass

    def closeEvent(self, event):
        try:
            if hasattr(self.canvas, 'ax'):
                self.canvas.ax.clear()
            self.canvas.figure.clf()
            self.canvas.deleteLater()

            if self.lasso:
                self.lasso.disconnect_events()
                self.lasso = None
        except Exception as e:
            print(f"Cleanup error: {e}")
        event.accept()

from PySide6.QtWidgets import (
    QDialog, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QFileDialog, QComboBox, QScrollArea, QWidget, QSizePolicy
)

from matplotlib.figure import Figure
import numpy as np
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score, mean_squared_error



class AnalyzeDialog(QDialog):
    def __init__(self, time_channels, freqs, filename, fundamental=True, harmonic=False, parent=None, session=None):
        super().__init__(parent)
        self.fundamental = fundamental
        self.harmonic = harmonic

        self.setWindowTitle("Analyzer")
        self.resize(1100, 700)

        self.time = np.array(time_channels) * 0.25
        self.freq = np.array(freqs)
        self.filename = filename.split(".")[0]
        self.current_plot_title = f"{self.filename}_Best_Fit"

        # Canvas
        self.canvas = MplCanvas(self, width=8, height=5)
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)

        # Buttons
        self.max_button = QPushButton("Maximum Intensities")
        self.fit_button = QPushButton("Best Fit")
        self.save_plot_button = QPushButton("Save Graph")
        self.save_data_button = QPushButton("Save Data")
        self.existing_excel_checkbox = QCheckBox("Existing Excel File")

        self.extra_plot_label = QLabel("Extra Plots:")
        self.extra_plot_combo = QComboBox()
        self.extra_plot_combo.addItems([
            "Shock Speed vs Shock Height",
            "Shock Speed vs Frequency",
            "Shock Height vs Frequency"
        ])
        self.extra_plot_button = QPushButton("Plot")

        self.max_button.clicked.connect(self.plot_max)
        self.fit_button.clicked.connect(self.plot_fit)
        self.save_plot_button.clicked.connect(self.save_graph)
        self.save_data_button.clicked.connect(self.save_data)
        self.extra_plot_button.clicked.connect(self.plot_extra)

        # Plot control layout
        plot_button_layout = QHBoxLayout()
        plot_button_layout.addWidget(self.max_button)
        plot_button_layout.addWidget(self.fit_button)

        left_layout = QVBoxLayout()
        left_layout.addLayout(plot_button_layout)
        left_layout.addWidget(self.canvas)

        # === Info Panel ===

        # --- Newkirk fold selection (n-fold) ---
        self.fold_label = QLabel("Fold-number:")
        self.fold_combo = QComboBox()
        self.fold_combo.addItems(["1", "2", "3", "4"])
        self.fold_combo.setCurrentIndex(0)

        self.fold_calc_button = QPushButton("Calculate")
        self.fold_calc_button.setEnabled(False)  # enable only after Best Fit
        self.fold_calc_button.clicked.connect(self.recalculate_shock_parameters)

        # Put fold controls into a widget so it can live inside self.labels
        self.fold_row_widget = QWidget()
        fold_row_layout = QHBoxLayout(self.fold_row_widget)
        fold_row_layout.setContentsMargins(0, 0, 0, 0)
        fold_row_layout.addWidget(self.fold_label)
        fold_row_layout.addWidget(self.fold_combo)
        fold_row_layout.addWidget(self.fold_calc_button)

        # Optional: keeps it neat
        self.fold_combo.setFixedWidth(70)

        self.equation_label = QLabel("Best Fit Equation:")
        self.equation_display = QLabel("")
        self.equation_display.setTextFormat(Qt.RichText)
        self.equation_display.setStyleSheet("font-size: 16px; padding: 4px;")

        self.stats_header = QLabel("<b>Fit Metrics:</b>")
        self.r2_display = QLabel("R² = ")
        self.rmse_display = QLabel("RMSE = ")

        self.shock_header = QLabel("<b>Shock Parameters:</b>")
        self.avg_freq_display = QLabel("")
        self.drift_display = QLabel("")
        self.start_freq_display = QLabel("")
        self.initial_shock_speed_display = QLabel("")
        self.initial_shock_height_display = QLabel("")
        self.avg_shock_speed_display = QLabel("")
        self.avg_shock_height_display = QLabel("")

        self.labels = [
            self.fold_row_widget,
            self.equation_label, self.equation_display,
            self.stats_header, self.r2_display, self.rmse_display,
            self.shock_header,
            self.avg_freq_display, self.drift_display, self.start_freq_display,
            self.initial_shock_speed_display, self.initial_shock_height_display,
            self.avg_shock_speed_display, self.avg_shock_height_display,
            self.save_plot_button, self.save_data_button, self.existing_excel_checkbox,
            self.extra_plot_label, self.extra_plot_combo, self.extra_plot_button
        ]

        right_inner = QVBoxLayout()
        for widget in self.labels:
            right_inner.addWidget(widget)
        right_inner.addStretch()

        right_inner.addStretch()

        right_widget = QWidget()
        right_widget.setLayout(right_inner)
        right_scroll = QScrollArea()
        right_scroll.setWidget(right_widget)
        right_scroll.setWidgetResizable(True)
        right_scroll.setMinimumWidth(300)

        self.status = QStatusBar()
        self.status.showMessage("Ready")

        # Main layout
        main_layout = QHBoxLayout()
        main_layout.addLayout(left_layout, stretch=3)
        main_layout.addWidget(right_scroll, stretch=1)
        main_with_status = QVBoxLayout()
        main_with_status.addLayout(main_layout)
        main_with_status.addWidget(self.status)
        self.setLayout(main_with_status)

        # Styling
        self.setStyleSheet("""
            QLabel {
                font-size: 13px;
                padding: 2px;
            }
            QLabel#value {
                font-weight: bold;
            }
            QPushButton {
                padding: 6px;
            }
        """)

        if isinstance(session, dict):
            try:
                self.restore_session(session)
            except Exception:
                pass

    def plot_max(self):
        self.canvas.ax.clear()
        self.canvas.ax.scatter(self.time, self.freq, s=10, color='blue')
        self.canvas.ax.set_title(f"{self.filename}_Maximum_Intensity")
        self.canvas.ax.set_xlabel("Time (s)")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.grid(True)
        self.canvas.draw()
        self.equation_display.setText("")
        self.fold_calc_button.setEnabled(False)
        self.status.showMessage("Max intensities plotted successfully!", 3000)

    def plot_fit(self, _checked=False, params=None, std_errs=None):
        def model_func(t, a, b): return a * t ** (b)

        def drift_rate(t, a_, b_): return a_ * b_ * t ** (b_ - 1)

        if params is None:
            params, cov = curve_fit(model_func, self.time, self.freq, maxfev=10000)
            a, b = params
            std_errs = np.sqrt(np.diag(cov))
        else:
            a, b = params
            if std_errs is None:
                std_errs = np.array([np.nan, np.nan], dtype=float)

        time_fit = np.linspace(self.time.min(), self.time.max(), 400)
        freq_fit = model_func(time_fit, a, b)

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)

        self.canvas.ax.scatter(self.time, self.freq, s=10, color='blue', label="Original Data")
        self.canvas.ax.plot(time_fit, freq_fit, color='red', label=fr"Best Fit: $f = {a:.2f} \cdot t^{{{b:.2f}}}$")
        self.canvas.ax.set_title(f"{self.filename}_Best_Fit")
        self.canvas.ax.set_xlabel("Time (s)")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.legend()
        self.canvas.ax.grid(True)
        self.canvas.draw()
        self.current_plot_title = f"{self.filename}_Best_Fit"

        predicted = model_func(self.time, a, b)
        r2 = r2_score(self.freq, predicted)
        rmse = np.sqrt(mean_squared_error(self.freq, predicted))

        self.equation_display.setText(f"<b>f(t) = {a:.2f} · t<sup>{b:.2f}</sup></b>")
        self.r2_display.setText(f"R² = {r2:.4f}")
        self.rmse_display.setText(f"RMSE = {rmse:.4f}")

        # Cache fit parameters for session persistence
        try:
            self._fit_params = {
                "a": float(a),
                "b": float(b),
                "std_errs": [float(std_errs[0]), float(std_errs[1])],
                "r2": float(r2),
                "rmse": float(rmse),
            }
        except Exception:
            self._fit_params = None

        drift_vals = drift_rate(self.time, a, b)
        residuals = self.freq - predicted
        freq_err = np.std(residuals)
        drift_errs = np.abs(drift_vals) * np.sqrt((std_errs[0] / a) ** 2 + (std_errs[1] / b) ** 2)

        drift_vals = drift_rate(self.time, a, b)
        residuals = self.freq - predicted
        freq_err = np.std(residuals)
        drift_errs = np.abs(drift_vals) * np.sqrt((std_errs[0] / a) ** 2 + (std_errs[1] / b) ** 2)

        # Cache results so we can recompute shock params for different folds
        self._drift_vals = drift_vals
        self._drift_errs = drift_errs
        self.freq_err = freq_err

        # Enable fold recalculation now that Best Fit exists
        self.fold_calc_button.setEnabled(True)

        # Compute and display shock parameters using selected fold-number
        self._update_shock_parameters(self._selected_fold())

        self.status.showMessage("Best fit plotted successfully!", 3000)

    def session_state(self) -> dict:
        state = {
            "fundamental": bool(getattr(self, "fundamental", True)),
            "harmonic": bool(getattr(self, "harmonic", False)),
            "fold": int(self._selected_fold()),
        }
        fit = getattr(self, "_fit_params", None)
        if isinstance(fit, dict):
            state["fit_params"] = fit
        return state

    def restore_session(self, state: dict):
        try:
            fold = int(state.get("fold", 1))
        except Exception:
            fold = 1
        fold = max(1, min(4, fold))
        try:
            self.fold_combo.setCurrentIndex(fold - 1)
        except Exception:
            pass

        fit = state.get("fit_params", None)
        if not isinstance(fit, dict):
            return

        if "a" not in fit or "b" not in fit:
            return

        try:
            a = float(fit["a"])
            b = float(fit["b"])
        except Exception:
            return

        std_errs = fit.get("std_errs", None)
        std_errs_arr = None
        if isinstance(std_errs, (list, tuple)) and len(std_errs) >= 2:
            try:
                std_errs_arr = np.array([float(std_errs[0]), float(std_errs[1])], dtype=float)
            except Exception:
                std_errs_arr = None

        self.plot_fit(params=(a, b), std_errs=std_errs_arr)

    def _selected_fold(self):
        try:
            n = int(self.fold_combo.currentText())
        except Exception:
            n = 1
        return max(1, min(4, n))

    def recalculate_shock_parameters(self):
        if not hasattr(self, "_drift_vals") or not hasattr(self, "_drift_errs"):
            QMessageBox.information(self, "Analyzer", "Please click 'Best Fit' first.")
            return

        n = self._selected_fold()
        self._update_shock_parameters(n)
        self.status.showMessage(f"Updated using Newkirk {n}-fold model.", 3000)

    def _update_shock_parameters(self, n):
        # Your updated n-fold formulas
        denom = n * 3.385
        drift_vals = self._drift_vals
        drift_errs = self._drift_errs

        shock_speed = (13853221.38 * np.abs(drift_vals)) / (
                self.freq * (np.log(self.freq ** 2 / denom) ** 2)
        )
        R_p = 4.32 * np.log(10) / np.log(self.freq ** 2 / denom)

        # Starting frequency (same logic you already use)
        percentile = 90
        start_freq = np.percentile(self.freq, percentile)
        if self.harmonic:
            start_freq = start_freq / 2

        idx = np.abs(self.freq - start_freq).argmin()
        f0 = self.freq[idx]
        drift_err0 = drift_errs[idx]

        start_shock_speed = shock_speed[idx]
        start_height = R_p[idx]

        shock_speed_err = (13853221.38 * drift_err0) / (
                f0 * (np.log(f0 ** 2 / denom) ** 2)
        )

        # Error propagation for R_p based on your n-fold expression
        g0 = np.log(f0 ** 2 / denom)
        dRp_df = 8.64 * np.log(10) / (f0 * (g0 ** 2))
        Rp_err = np.abs(dRp_df * self.freq_err)

        # Averages (drift and freq do not depend on n, speeds/heights do)
        avg_freq = np.mean(self.freq)
        avg_freq_err = np.std(self.freq) / np.sqrt(len(self.freq))
        avg_drift = np.mean(drift_vals)
        avg_drift_err = np.std(drift_vals) / np.sqrt(len(drift_vals))

        avg_speed = np.mean(shock_speed)
        avg_speed_err = np.std(shock_speed) / np.sqrt(len(shock_speed))
        avg_height = np.mean(R_p)
        avg_height_err = np.std(R_p) / np.sqrt(len(R_p))

        # Store arrays for extra plots
        self.shock_speed = shock_speed
        self.R_p = R_p
        self.start_freq = start_freq
        self.start_height = start_height

        # Optional but helpful to show which model is used
        self.shock_header.setText(f"<b>Shock Parameters (Newkirk {n}-fold):</b>")

        # Update the right-panel text
        self.avg_freq_display.setText(f"Average Frequency: <b>{avg_freq:.2f} ± {avg_freq_err:.2f}</b> MHz")
        self.drift_display.setText(f"Average Drift Rate: <b>{avg_drift:.4f} ± {avg_drift_err:.4f}</b> MHz/s")
        self.start_freq_display.setText(f"Starting Frequency: <b>{start_freq:.2f} ± {self.freq_err:.2f}</b> MHz")
        self.initial_shock_speed_display.setText(
            f"Initial Shock Speed: <b>{start_shock_speed:.2f} ± {shock_speed_err:.2f}</b> km/s"
        )
        self.initial_shock_height_display.setText(
            f"Initial Shock Height: <b>{start_height:.3f} ± {Rp_err:.3f}</b> Rₛ"
        )
        self.avg_shock_speed_display.setText(f"Average Shock Speed: <b>{avg_speed:.2f} ± {avg_speed_err:.2f}</b> km/s")
        self.avg_shock_height_display.setText(
            f"Average Shock Height: <b>{avg_height:.3f} ± {avg_height_err:.3f}</b> Rₛ"
        )

    def save_graph(self):
        plot_name = getattr(self, "current_plot_title", None) or f"{self.filename}_Plot"

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        file_path, ext = pick_export_path(
            self,
            "Export Figure",
            plot_name,
            formats,
            default_filter="PNG (*.png)"
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        try:
            # ✅ If user didn't type an extension, add the one from ext
            root, current_ext = os.path.splitext(file_path)
            if current_ext == "":
                ext = ext.lower().lstrip(".")
                file_path = f"{file_path}.{ext}"
            else:
                ext = current_ext.lower().lstrip(".")

            self.canvas.figure.savefig(
                file_path,
                dpi=300,
                bbox_inches="tight",
                format=ext
            )
            QMessageBox.information(self, "Export Complete", f"Plot saved:\n{file_path}")
            self.status.showMessage("Export successful!", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not save file:\n{e}")
            self.status.showMessage("Export failed!", 3000)

    def save_data(self):

        # All known e-Callisto station names
        station_list = [
            'ALASKA-ANCHORAGE', 'ALASKA-COHOE', 'ALASKA-HAARP', 'ALGERIA-CRAAG', 'ALMATY',
            'Arecibo-observatory', 'AUSTRIA-Krumbach', 'AUSTRIA-MICHELBACH', 'AUSTRIA-OE3FLB',
            'AUSTRIA-UNIGRAZ', 'Australia-ASSA', 'BRAZIL', 'BIR', 'Croatia-Visnjan', 'DENMARK',
            'EGYPT-Alexandria', 'EGYPT-SpaceAgency', 'ETHIOPIA', 'FINLAND-Siuntio', 'FINLAND-Kempele',
            'GERMANY-ESSEN', 'GERMANY-DLR', 'GLASGOW', 'GREENLAND', 'HUMAIN', 'HURBANOVO',
            'INDIA-GAURI', 'INDIA-Nashik', 'INDIA-OOTY', 'INDIA-UDAIPUR', 'INDONESIA',
            'ITALY-Strassolt', 'JAPAN-IBARAKI', 'KASI', 'KRIM', 'MEXART',
            'MEXICO-ENSENADA-UNAM', 'MEXICO-FCFM-UANL', 'MEXICO-FCFM-UNACH', 'MEXICO-LANCE-A',
            'MEXICO-LANCE-B', 'MEXICO-UANL-INFIERNILLO', 'MONGOLIA-UB', 'MRO', 'MRT1', 'MRT3',
            'Malaysia_Banting', 'NASA-GSFC', 'NORWAY-EGERSUND', 'NORWAY-NY-AALESUND', 'NORWAY-RANDABERG',
            'PARAGUAY', 'POLAND-BALDY', 'POLAND-Grotniki', 'ROMANIA', 'ROSWELL-NM', 'RWANDA',
            'SOUTHAFRICA-SANSA', 'SPAIN-ALCALA', 'SPAIN-PERALEJOS', 'SPAIN-SIGUENZA', 'SRI-Lanka',
            'SSRT', 'SWISS-CalU', 'SWISS-FM', 'SWISS-HB9SCT', 'SWISS-HEITERSWIL', 'SWISS-IRSOL',
            'SWISS-Landschlacht', 'SWISS-MUHEN', 'TAIWAN-NCU', 'THAILAND-Pathumthani', 'TRIEST',
            'TURKEY', 'UNAM', 'URUGUAY', 'USA-ARIZONA-ERAU', 'USA-BOSTON', 'UZBEKISTAN'
        ]

        # ✅ Extract Station
        station = "UNKNOWN"
        filename_lower = self.filename.lower()
        for s in station_list:
            if filename_lower.startswith(s.lower()):
                station = s
                break

        # ✅ Extract Date
        date_match = re.search(r'_(\d{4})(\d{2})(\d{2})_', self.filename)
        if date_match:
            date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
        else:
            date = "UNKNOWN"

        # ✅ Excel File Handling
        if self.existing_excel_checkbox.isChecked():
            path, _ = QFileDialog.getOpenFileName(self, "Select Existing Excel File", "", "Excel Files (*.xlsx)")
            if not path:
                return
            try:
                wb = load_workbook(path)
                ws = wb.active
            except Exception as e:
                QMessageBox.critical(self, "Load Error", f"Could not open Excel file:\n{str(e)}")
                return
        else:
            path, _ = QFileDialog.getSaveFileName(self, "Save as Excel", f"{self.filename}_data.xlsx",
                                                  "Excel Files (*.xlsx)")
            if not path:
                return
            try:
                wb = Workbook()
                ws = wb.active
                headers = [
                    "Date", "Station", "Best_fit", "R_sq", "RMSE",
                    "avg_freq", "avg_freq_err", "Avg_drift", "avg_drift_err",
                    "start_freq", "start_freq_err", "initial_shock_speed", "initial_shock_speed_err",
                    "initial_shock_height", "initial_shock_height_err", "avg_shock_speed", "avg_shock_speed_err",
                    "avg_shock_height", "avg_shock_height_err", "avg_drift_abs"
                ]
                ws.append(headers)
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Could not create Excel file:\n{str(e)}")
                return

        # ✅ Extract and clean text
        def extract_val_err(label):
            # Remove HTML tags
            clean_text = re.sub(r'<[^>]+>', '', label.text())
            # Remove units and stray characters
            clean_text = re.sub(r'(MHz|km/s|Rₛ|s|/)', '', clean_text)
            # Clean spaces
            clean_text = clean_text.strip()
            # Extract value ± error
            value_text = clean_text.split(":")[-1].strip()
            if "±" in value_text:
                value, err = value_text.split("±")
                return value.strip(), err.strip()
            else:
                return value_text.strip(), ""

        # ✅ Read values
        try:
            best_fit = re.sub(r'<[^>]+>', '', self.equation_display.text()).replace("<sup>", "^").replace("</sup>", "")
            r2 = self.r2_display.text().split("=")[-1].strip()
            rmse = self.rmse_display.text().split("=")[-1].strip()

            avg_freq, avg_freq_err = extract_val_err(self.avg_freq_display)
            avg_drift, avg_drift_err = extract_val_err(self.drift_display)

            try:
                avg_drift_abs = abs(float(avg_drift))
            except ValueError:
                avg_drift_abs = ""

            start_freq, start_freq_err = extract_val_err(self.start_freq_display)
            init_speed, init_speed_err = extract_val_err(self.initial_shock_speed_display)
            init_height, init_height_err = extract_val_err(self.initial_shock_height_display)
            avg_speed, avg_speed_err = extract_val_err(self.avg_shock_speed_display)
            avg_height, avg_height_err = extract_val_err(self.avg_shock_height_display)

            row = [
                date, station, best_fit, r2, rmse,
                avg_freq, avg_freq_err, avg_drift, avg_drift_err,
                start_freq, start_freq_err, init_speed, init_speed_err,
                init_height, init_height_err, avg_speed, avg_speed_err,
                avg_height, avg_height_err, avg_drift_abs
            ]

            ws.append(row)
            wb.save(path)
            self.status.showMessage("✅ Data saved to Excel successfully!", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Write Error", f"Could not write to Excel file:\n{str(e)}")
            self.status.showMessage("❌ Failed to save data to Excel.", 3000)

    def plot_extra(self):
        choice = self.extra_plot_combo.currentText()
        self.canvas.ax.clear()
        if choice == "Shock Speed vs Shock Height":
            self.canvas.ax.scatter(self.R_p, self.shock_speed, color='green', s=10)
            self.canvas.ax.set_xlabel("Shock Height (Rₛ)")
            self.canvas.ax.set_ylabel("Shock Speed (km/s)")
            self.canvas.ax.set_title(f"{self.filename}_Shock_Speed_vs_Shock_Height")
            self.current_plot_title = f"{self.filename}_Shock_Speed_vs_Shock_Height"
            self.status.showMessage("Shock Speed vs Shock Height plotted successfully!", 3000)

        elif choice == "Shock Speed vs Frequency":
            self.canvas.ax.scatter(self.freq, self.shock_speed, color='purple', s=10)
            self.canvas.ax.set_xlabel("Frequency (MHz)")
            self.canvas.ax.set_ylabel("Shock Speed (km/s)")
            self.canvas.ax.set_title(f"{self.filename}_Shock_Speed_vs_Frequency")
            self.current_plot_title = f"{self.filename}_Shock_Speed_vs_Frequency"
            self.status.showMessage("Shock Speed vs Frequency plotted successfully!", 3000)

        elif choice == "Shock Height vs Frequency":
            self.canvas.ax.scatter(self.R_p, self.freq, color='red', s=10)
            self.canvas.ax.set_xlabel("Shock Height (Rₛ)")
            self.canvas.ax.set_ylabel("Frequency (MHz)")
            self.canvas.ax.set_title(f"{self.filename}_Rs_vs_Freq")
            self.current_plot_title = f"{self.filename}_Rs_vs_Freq"
            self.status.showMessage("Shock Height vs Frequency plotted successfully!", 3000)
        self.canvas.ax.grid(True)
        self.canvas.draw()


class CombineFrequencyDialog(QDialog):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.setWindowTitle("Combine Frequency Ranges")
        self.setMinimumWidth(600)

        self.file_paths = []

        self.load_button = QPushButton("Import FITS Files")
        self.load_button.clicked.connect(self.load_files)

        self.combine_button = QPushButton("Combine")
        self.combine_button.clicked.connect(self.combine_files)
        self.combine_button.setEnabled(False)

        self.import_button = QPushButton("Import to Analyzer")
        self.import_button.clicked.connect(self.import_to_main)
        self.import_button.setEnabled(False)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)

        self.image_label = QLabel("Combined output will appear here.")
        self.image_label.setAlignment(Qt.AlignCenter)

        layout = QVBoxLayout()
        layout.addWidget(self.load_button)
        layout.addWidget(self.combine_button)
        layout.addWidget(self.import_button)
        layout.addWidget(self.image_label)
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)

        self.combined_data = None
        self.combined_freqs = None
        self.combined_time = None
        self.combined_filename = "Combined_Frequency"
        self.combined_header0 = None

    def load_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select FITS Files to Combine",
            "",
            "FITS files (*.fit *.fits *.fit.gz *.fits.gz)"
        )
        if len(files) != 2:
            QMessageBox.warning(self, "Error", "Please select exactly TWO files.")
            return

        station1 = files[0].split("/")[-1].split("_")[0]
        station2 = files[1].split("/")[-1].split("_")[0]

        if station1 != station2:
            QMessageBox.critical(self, "Error",
                                 "You must select consecutive frequency data files from the same station!")
            return

        self.file_paths = files
        self.combine_button.setEnabled(True)

    def combine_files(self):
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(10)
        QApplication.processEvents()

        try:
            res1 = load_callisto_fits(self.file_paths[0], memmap=False)
            data1, freqs1, time1, hdr1 = res1.data, res1.freqs, res1.time, res1.header0
            self.progress_bar.setValue(30)
            QApplication.processEvents()

            res2 = load_callisto_fits(self.file_paths[1], memmap=False)
            data2, freqs2, time2 = res2.data, res2.freqs, res2.time
            self.progress_bar.setValue(60)
            QApplication.processEvents()

            if not np.allclose(time1, time2, rtol=1e-2):
                QMessageBox.critical(self, "Error", "Time arrays must match to combine frequencies.")
                self.progress_bar.setVisible(False)
                return

            self.combined_data = np.vstack([data1, data2])
            self.combined_freqs = np.concatenate([freqs1, freqs2])
            self.combined_time = time1
            self.combined_header0 = build_combined_header(
                hdr1,
                mode="frequency",
                sources=self.file_paths,
                data_shape=self.combined_data.shape,
                freqs=self.combined_freqs,
                time=self.combined_time,
            )
            self.progress_bar.setValue(80)
            QApplication.processEvents()

            # Plot image
            fig, ax = plt.subplots(figsize=(6, 4))
            style_axes(ax)
            extent = [0, self.combined_time[-1], self.combined_freqs[-1], self.combined_freqs[0]]
            cmap = mcolors.LinearSegmentedColormap.from_list("custom", [(0.0, 'blue'), (0.5, 'red'), (1.0, 'yellow')])
            ax.imshow(self.combined_data, aspect='auto', extent=extent, cmap=cmap)
            ax.set_xlabel("Time [s]")
            ax.set_ylabel("Frequency [MHz]")
            # Extract base filenames (e.g., 'BIR_20240720_123000_123000_00.fit.gz')
            fname1 = os.path.basename(self.file_paths[0])
            fname2 = os.path.basename(self.file_paths[1])

            # Extract focus codes (last 2 digits before .fit.gz, assuming filename ends with _00.fit.gz or _01.fit.gz etc.)
            focus1 = fname1.split("_")[-1].split(".")[0]
            focus2 = fname2.split("_")[-1].split(".")[0]

            # Extract common base (e.g., remove focus code and extension)
            base_name = "_".join(fname1.split("_")[:-1])

            # Set title with base + both focus codes
            ax.set_title(f"{base_name}_{focus1}+{focus2} (Combined Frequency)")

            self.combined_title = f"{base_name}_{focus1}+{focus2} (Combined Frequency)"
            ax.set_title(self.combined_title)

            buf = io.BytesIO()
            fig.savefig(buf, format='png')
            buf.seek(0)
            img = QImage()
            img.loadFromData(buf.read())
            self.image_label.setPixmap(QPixmap.fromImage(img).scaledToWidth(550))
            buf.close()
            plt.close(fig)

            self.progress_bar.setValue(100)
            QApplication.processEvents()
            self.import_button.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            self.progress_bar.setVisible(False)

    def import_to_main(self):
        if self.combined_data is None or self.combined_freqs is None or self.combined_time is None:
            QMessageBox.warning(self, "No Data", "Please combine the files first.")
            return
        self.main_window.raw_data = self.combined_data
        self.main_window.freqs = self.combined_freqs
        self.main_window.time = self.combined_time
        self.main_window.filename = self.combined_title  # ✅ update filename as the title

        # Mark as combined so Export-to-FITS can record provenance correctly
        self.main_window._is_combined = True
        self.main_window._combined_mode = "frequency"
        self.main_window._combined_sources = list(self.file_paths)
        self.main_window._fits_header0 = self.combined_header0.copy() if self.combined_header0 is not None else None
        self.main_window._fits_source_path = None

        # UT start for UT-axis formatting
        self.main_window.ut_start_sec = extract_ut_start_sec(self.main_window._fits_header0)

        # Reset derived state for a fresh start
        self.main_window.noise_reduced_data = None
        self.main_window.noise_reduced_original = None
        self.main_window.lasso_mask = None
        self.main_window.noise_vmin = None
        self.main_window.noise_vmax = None
        self.main_window.current_display_data = None
        self.main_window._undo_stack.clear()
        self.main_window._redo_stack.clear()

        self.main_window._project_path = None
        self.main_window._max_intensity_state = None
        self.main_window._mark_project_dirty()

        self.main_window.plot_data(self.combined_data, title="Raw")
        self.close()


class CombineTimeDialog(QDialog):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.setWindowTitle("Combine Time Ranges")
        self.setMinimumWidth(600)

        self.file_paths = []
        self.combined_data = None
        self.combined_header0 = None

        # Buttons
        self.load_button = QPushButton("Import FITS Files")
        self.load_button.clicked.connect(self.load_files)

        self.combine_button = QPushButton("Combine")
        self.combine_button.clicked.connect(self.combine_files)
        self.combine_button.setEnabled(False)

        self.import_button = QPushButton("Import to Analyzer")
        self.import_button.clicked.connect(self.import_to_main)
        self.import_button.setEnabled(False)

        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)

        # Output Image Preview
        self.image_label = QLabel("Combined output will appear here.")
        self.image_label.setAlignment(Qt.AlignCenter)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.load_button)
        layout.addWidget(self.combine_button)
        layout.addWidget(self.import_button)
        layout.addWidget(self.image_label)
        layout.addWidget(self.progress_bar)
        self.setLayout(layout)

    def load_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select FITS Files to Combine",
            "",
            "FITS files (*.fit *.fits *.fit.gz *.fits.gz)"
        )

        if len(files) < 2:
            QMessageBox.warning(self, "Error", "Please select at least 2 FITS files.")
            return

        try:
            from datetime import datetime

            # Sort files by timestamp
            self.file_paths = sorted(files, key=lambda f: os.path.basename(f).split("_")[2])

            # Check station/date and time continuity
            parts_ref = os.path.basename(self.file_paths[0]).split("_")
            t_ref = datetime.strptime(parts_ref[2], "%H%M%S")

            for f in self.file_paths[1:]:
                parts = os.path.basename(f).split("_")
                if parts[0] != parts_ref[0] or parts[1] != parts_ref[1]:
                    raise ValueError("Different station or date")

                t_next = datetime.strptime(parts[2], "%H%M%S")
                diff = abs((t_next - t_ref).total_seconds())
                if not (800 <= diff <= 1000):  # ~15min ±1.5min
                    raise ValueError(f"File {f} is not consecutive")
                t_ref = t_next

            self.combine_button.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Invalid Selection", f"Error while validating files:\n{str(e)}")

    def combine_files(self):
        if len(self.file_paths) < 2:
            QMessageBox.warning(self, "Error", "Please load at least 2 valid FITS files to combine.")
            return

        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(10)

        try:
            combined_data = None
            combined_time = None
            reference_freqs = None
            header0 = None

            for idx, file_path in enumerate(self.file_paths):
                res = load_callisto_fits(file_path, memmap=False)
                data, freqs, time = res.data, res.freqs, res.time
                if header0 is None:
                    header0 = res.header0

                if reference_freqs is None:
                    reference_freqs = freqs
                elif not np.allclose(freqs, reference_freqs):
                    raise ValueError("Frequency mismatch in file: " + os.path.basename(file_path))

                # Compute dt and shift time
                if idx == 0:
                    dt = time[1] - time[0]
                    adjusted_time = time
                    combined_data = data
                    combined_time = adjusted_time
                else:
                    dt = time[1] - time[0]
                    shift = combined_time[-1] + dt
                    adjusted_time = time + shift
                    combined_data = np.concatenate((combined_data, data), axis=1)
                    combined_time = np.concatenate((combined_time, adjusted_time))

            self.combined_data = combined_data
            self.combined_time = combined_time
            self.main_window.freqs = reference_freqs
            self.main_window.time = combined_time
            self.combined_header0 = build_combined_header(
                header0,
                mode="time",
                sources=self.file_paths,
                data_shape=combined_data.shape,
                freqs=reference_freqs,
                time=combined_time,
            )

            self.progress_bar.setValue(80)

            # Plot preview
            fig, ax = plt.subplots(figsize=(6, 4))
            style_axes(ax)
            extent = [combined_time[0], combined_time[-1], reference_freqs[-1], reference_freqs[0]]
            cmap = LinearSegmentedColormap.from_list('custom_cmap', [(0, 'darkblue'), (1, 'orange')])
            im = ax.imshow(combined_data, aspect='auto', extent=extent, cmap=cmap)
            ax.set_xlabel("Time [s]")
            ax.set_ylabel("Frequency [MHz]")
            ax.set_title("Combined Time Plot")
            fig.tight_layout()

            temp_dir = tempfile.gettempdir()
            preview_path = os.path.join(temp_dir, "preview_combined_time.png")
            fig.savefig(preview_path, dpi=100)
            plt.close(fig)

            self.image_label.setPixmap(QPixmap(preview_path).scaled(550, 350, Qt.KeepAspectRatio))
            self.progress_bar.setValue(100)
            self.import_button.setEnabled(True)

            # Set filename
            base1 = os.path.basename(self.file_paths[0]).split(".")[0]
            self.main_window.filename = base1 + "_combined_time"

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to combine:\n{str(e)}")

    def import_to_main(self):
        if self.combined_data is not None and self.combined_time is not None:
            self.main_window.raw_data = self.combined_data
            self.main_window.freqs = self.main_window.freqs  # already set earlier
            self.main_window.time = self.combined_time
            self.main_window.filename = self.main_window.filename  # already set earlier

            # Mark as combined so Export-to-FITS can record provenance correctly
            self.main_window._is_combined = True
            self.main_window._combined_mode = "time"
            self.main_window._combined_sources = list(self.file_paths)
            self.main_window._fits_header0 = self.combined_header0.copy() if self.combined_header0 is not None else None
            self.main_window._fits_source_path = None

            # Calculate UT start from FITS header of first file
            self.main_window.ut_start_sec = extract_ut_start_sec(self.main_window._fits_header0)

            # Reset derived state for a fresh start
            self.main_window.noise_reduced_data = None
            self.main_window.noise_reduced_original = None
            self.main_window.lasso_mask = None
            self.main_window.noise_vmin = None
            self.main_window.noise_vmax = None
            self.main_window.current_display_data = None
            self.main_window._undo_stack.clear()
            self.main_window._redo_stack.clear()

            self.main_window._project_path = None
            self.main_window._max_intensity_state = None
            self.main_window._mark_project_dirty()

            self.main_window.plot_data(self.combined_data, title="Raw")
            self.close()
