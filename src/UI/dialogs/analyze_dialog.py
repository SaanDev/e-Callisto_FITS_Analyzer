"""
e-CALLISTO FITS Analyzer
Version 2.2-dev
Sahan S Liyanage (sahanslst@gmail.com)
Astronomical and Space Science Unit, University of Colombo, Sri Lanka.
"""

from __future__ import annotations

import os
import re
import sys

import numpy as np
from openpyxl import Workbook, load_workbook
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QStatusBar,
    QVBoxLayout,
    QWidget,
)
from scipy.optimize import curve_fit
from sklearn.metrics import mean_squared_error, r2_score

from src.UI.gui_shared import MplCanvas, pick_export_path
from src.UI.mpl_style import style_axes

class AnalyzeDialog(QDialog):
    sessionChanged = Signal(dict)

    def __init__(self, time_channels, freqs, filename, fundamental=True, harmonic=False, parent=None, session=None):
        super().__init__(parent)
        self.fundamental = fundamental
        self.harmonic = harmonic

        self.setWindowTitle("Analyzer")
        self.resize(1100, 700)

        self.time_channels = np.asarray(time_channels, dtype=float).reshape(-1)
        self.time = self.time_channels * 0.25
        self.freq = np.asarray(freqs, dtype=float).reshape(-1)
        self.filename = filename.split(".")[0]
        self.current_plot_title = f"{self.filename}_Best_Fit"
        self._fit_params = None
        self._shock_summary = {}
        self._suppress_emit = False

        # Canvas
        self.canvas = MplCanvas(self, width=8, height=5)
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)

        # Buttons
        self.max_button = QPushButton("Maximum Intensities")
        self.fit_button = QPushButton("Best Fit")
        self.save_plot_button = QPushButton("Save Graph")
        self.save_data_button = QPushButton("Save Data")
        self.existing_excel_checkbox = QCheckBox("Existing Excel File")

        self.extra_plot_label = QLabel("Extra Plots:")
        self.extra_plot_combo = QComboBox()
        self.extra_plot_combo.addItems([
            "Shock Speed vs Shock Height",
            "Shock Speed vs Frequency",
            "Shock Height vs Frequency"
        ])
        self.extra_plot_button = QPushButton("Plot")

        self.max_button.clicked.connect(self.plot_max)
        self.fit_button.clicked.connect(self.plot_fit)
        self.save_plot_button.clicked.connect(self.save_graph)
        self.save_data_button.clicked.connect(self.save_data)
        self.extra_plot_button.clicked.connect(self.plot_extra)

        # Plot control layout
        plot_button_layout = QHBoxLayout()
        plot_button_layout.addWidget(self.max_button)
        plot_button_layout.addWidget(self.fit_button)

        left_layout = QVBoxLayout()
        left_layout.addLayout(plot_button_layout)
        left_layout.addWidget(self.canvas)

        # === Info Panel ===

        # --- Newkirk fold selection (n-fold) ---
        self.fold_label = QLabel("Fold-number:")
        self.fold_combo = QComboBox()
        self.fold_combo.addItems(["1", "2", "3", "4"])
        self.fold_combo.setCurrentIndex(0)

        self.fold_calc_button = QPushButton("Calculate")
        self.fold_calc_button.setEnabled(False)  # enable only after Best Fit
        self.fold_calc_button.clicked.connect(self.recalculate_shock_parameters)

        # Put fold controls into a widget so it can live inside self.labels
        self.fold_row_widget = QWidget()
        fold_row_layout = QHBoxLayout(self.fold_row_widget)
        fold_row_layout.setContentsMargins(0, 0, 0, 0)
        fold_row_layout.addWidget(self.fold_label)
        fold_row_layout.addWidget(self.fold_combo)
        fold_row_layout.addWidget(self.fold_calc_button)

        # Reserve enough room for the selected fold value across Qt styles/themes.
        self.fold_combo.setMinimumContentsLength(2)
        self.fold_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon)
        self.fold_combo.setMinimumWidth(max(70, self.fold_combo.sizeHint().width()))

        self.equation_label = QLabel("Best Fit Equation:")
        self.equation_display = QLabel("")
        self.equation_display.setTextFormat(Qt.RichText)
        self.equation_display.setStyleSheet("font-size: 16px; padding: 4px;")

        self.stats_header = QLabel("<b>Fit Metrics:</b>")
        self.r2_display = QLabel("R² = ")
        self.rmse_display = QLabel("RMSE = ")

        self.shock_header = QLabel("<b>Shock Parameters:</b>")
        self.avg_freq_display = QLabel("")
        self.drift_display = QLabel("")
        self.start_freq_display = QLabel("")
        self.initial_shock_speed_display = QLabel("")
        self.initial_shock_height_display = QLabel("")
        self.avg_shock_speed_display = QLabel("")
        self.avg_shock_height_display = QLabel("")

        self.labels = [
            self.fold_row_widget,
            self.equation_label, self.equation_display,
            self.stats_header, self.r2_display, self.rmse_display,
            self.shock_header,
            self.avg_freq_display, self.drift_display, self.start_freq_display,
            self.initial_shock_speed_display, self.initial_shock_height_display,
            self.avg_shock_speed_display, self.avg_shock_height_display,
            self.save_plot_button, self.save_data_button, self.existing_excel_checkbox,
            self.extra_plot_label, self.extra_plot_combo, self.extra_plot_button
        ]

        right_inner = QVBoxLayout()
        for widget in self.labels:
            right_inner.addWidget(widget)
        right_inner.addStretch()

        right_inner.addStretch()

        right_widget = QWidget()
        right_widget.setLayout(right_inner)
        right_scroll = QScrollArea()
        right_scroll.setWidget(right_widget)
        right_scroll.setWidgetResizable(True)
        right_scroll.setMinimumWidth(300)

        self.status = QStatusBar()
        self.status.showMessage("Ready")

        # Main layout
        main_layout = QHBoxLayout()
        main_layout.addLayout(left_layout, stretch=3)
        main_layout.addWidget(right_scroll, stretch=1)
        main_with_status = QVBoxLayout()
        main_with_status.addLayout(main_layout)
        main_with_status.addWidget(self.status)
        self.setLayout(main_with_status)

        if isinstance(session, dict):
            try:
                self.restore_session(session, emit_change=False)
            except Exception:
                pass

    def _emit_session_changed(self):
        if self._suppress_emit:
            return
        try:
            self.sessionChanged.emit(self.session_state())
        except Exception:
            pass

    def _set_summary_labels_from_dict(self, summary: dict):
        if not isinstance(summary, dict):
            return
        fold = int(summary.get("fold", self._selected_fold()) or self._selected_fold())
        self.shock_header.setText(f"<b>Shock Parameters (Newkirk {fold}-fold):</b>")

        def _f(v, digits=2):
            if v is None:
                return ""
            try:
                return f"{float(v):.{digits}f}"
            except Exception:
                return ""

        self.avg_freq_display.setText(
            f"Average Frequency: <b>{_f(summary.get('avg_freq_mhz'), 2)} ± {_f(summary.get('avg_freq_err_mhz'), 2)}</b> MHz"
        )
        self.drift_display.setText(
            f"Average Drift Rate: <b>{_f(summary.get('avg_drift_mhz_s'), 4)} ± {_f(summary.get('avg_drift_err_mhz_s'), 4)}</b> MHz/s"
        )
        self.start_freq_display.setText(
            f"Starting Frequency: <b>{_f(summary.get('start_freq_mhz'), 2)} ± {_f(summary.get('start_freq_err_mhz'), 2)}</b> MHz"
        )
        self.initial_shock_speed_display.setText(
            f"Initial Shock Speed: <b>{_f(summary.get('initial_shock_speed_km_s'), 2)} ± {_f(summary.get('initial_shock_speed_err_km_s'), 2)}</b> km/s"
        )
        self.initial_shock_height_display.setText(
            f"Initial Shock Height: <b>{_f(summary.get('initial_shock_height_rs'), 3)} ± {_f(summary.get('initial_shock_height_err_rs'), 3)}</b> Rₛ"
        )
        self.avg_shock_speed_display.setText(
            f"Average Shock Speed: <b>{_f(summary.get('avg_shock_speed_km_s'), 2)} ± {_f(summary.get('avg_shock_speed_err_km_s'), 2)}</b> km/s"
        )
        self.avg_shock_height_display.setText(
            f"Average Shock Height: <b>{_f(summary.get('avg_shock_height_rs'), 3)} ± {_f(summary.get('avg_shock_height_err_rs'), 3)}</b> Rₛ"
        )

    def plot_max(self):
        self.canvas.ax.clear()
        self.canvas.ax.scatter(self.time, self.freq, s=10, color='blue')
        self.canvas.ax.set_title(f"{self.filename}_Maximum_Intensity")
        self.canvas.ax.set_xlabel("Time (s)")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.grid(True)
        self.canvas.draw()
        self.equation_display.setText("")
        self.fold_calc_button.setEnabled(False)
        self.status.showMessage("Max intensities plotted successfully!", 3000)

    def plot_fit(self, _checked=False, params=None, std_errs=None):
        def model_func(t, a, b): return a * t ** (b)

        def drift_rate(t, a_, b_): return a_ * b_ * t ** (b_ - 1)

        if params is None:
            params, cov = curve_fit(model_func, self.time, self.freq, maxfev=10000)
            a, b = params
            std_errs = np.sqrt(np.diag(cov))
        else:
            a, b = params
            if std_errs is None:
                std_errs = np.array([np.nan, np.nan], dtype=float)

        time_fit = np.linspace(self.time.min(), self.time.max(), 400)
        freq_fit = model_func(time_fit, a, b)

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        style_axes(self.canvas.ax)

        self.canvas.ax.scatter(self.time, self.freq, s=10, color='blue', label="Original Data")
        self.canvas.ax.plot(time_fit, freq_fit, color='red', label=fr"Best Fit: $f = {a:.2f} \cdot t^{{{b:.2f}}}$")
        self.canvas.ax.set_title(f"{self.filename}_Best_Fit")
        self.canvas.ax.set_xlabel("Time (s)")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.legend()
        self.canvas.ax.grid(True)
        self.canvas.draw()
        self.current_plot_title = f"{self.filename}_Best_Fit"

        predicted = model_func(self.time, a, b)
        r2 = r2_score(self.freq, predicted)
        rmse = np.sqrt(mean_squared_error(self.freq, predicted))

        self.equation_display.setText(f"<b>f(t) = {a:.2f} · t<sup>{b:.2f}</sup></b>")
        self.r2_display.setText(f"R² = {r2:.4f}")
        self.rmse_display.setText(f"RMSE = {rmse:.4f}")

        # Cache fit parameters for session persistence
        try:
            self._fit_params = {
                "a": float(a),
                "b": float(b),
                "std_errs": [float(std_errs[0]), float(std_errs[1])],
                "r2": float(r2),
                "rmse": float(rmse),
            }
        except Exception:
            self._fit_params = None

        drift_vals = drift_rate(self.time, a, b)
        residuals = self.freq - predicted
        freq_err = np.std(residuals)
        drift_errs = np.abs(drift_vals) * np.sqrt((std_errs[0] / a) ** 2 + (std_errs[1] / b) ** 2)

        drift_vals = drift_rate(self.time, a, b)
        residuals = self.freq - predicted
        freq_err = np.std(residuals)
        drift_errs = np.abs(drift_vals) * np.sqrt((std_errs[0] / a) ** 2 + (std_errs[1] / b) ** 2)

        # Cache results so we can recompute shock params for different folds
        self._drift_vals = drift_vals
        self._drift_errs = drift_errs
        self.freq_err = freq_err

        # Enable fold recalculation now that Best Fit exists
        self.fold_calc_button.setEnabled(True)

        # Compute and display shock parameters using selected fold-number
        self._update_shock_parameters(self._selected_fold())

        self.status.showMessage("Best fit plotted successfully!", 3000)
        self._emit_session_changed()

    def session_state(self) -> dict:
        fold = int(self._selected_fold())
        fit = dict(getattr(self, "_fit_params", None) or {})
        shock = dict(getattr(self, "_shock_summary", {}) or {})
        if shock:
            shock["fold"] = int(shock.get("fold", fold) or fold)
            shock["fundamental"] = bool(shock.get("fundamental", self.fundamental))
            shock["harmonic"] = bool(shock.get("harmonic", self.harmonic))

        return {
            "source": {"filename": str(self.filename or "")},
            "max_intensity": {
                "time_channels": np.asarray(self.time_channels, dtype=float),
                "freqs": np.asarray(self.freq, dtype=float),
                "fundamental": bool(self.fundamental),
                "harmonic": bool(self.harmonic),
            },
            "analyzer": {
                "fit_params": fit,
                "fold": fold,
                "shock_summary": shock,
            },
            "ui": {
                "restore_max_window": True,
                "restore_analyzer_window": True,
            },
        }

    def restore_session(self, state: dict, *, emit_change: bool = True):
        self._suppress_emit = True
        try:
            max_block = dict(state.get("max_intensity") or state) if isinstance(state, dict) else {}
            analyzer = dict(state.get("analyzer") or state) if isinstance(state, dict) else {}
            fit = analyzer.get("fit_params", None)
            shock = analyzer.get("shock_summary", None)

            if max_block.get("time_channels") is not None and max_block.get("freqs") is not None:
                try:
                    t_arr = np.asarray(max_block.get("time_channels"), dtype=float).reshape(-1)
                    f_arr = np.asarray(max_block.get("freqs"), dtype=float).reshape(-1)
                    if len(t_arr) == len(f_arr) and len(t_arr) > 0:
                        self.time_channels = t_arr
                        self.time = self.time_channels * 0.25
                        self.freq = f_arr
                except Exception:
                    pass

            self.fundamental = bool(max_block.get("fundamental", self.fundamental))
            self.harmonic = bool(max_block.get("harmonic", self.harmonic))

            try:
                fold = int(analyzer.get("fold", 1))
            except Exception:
                fold = 1
            fold = max(1, min(4, fold))
            try:
                self.fold_combo.setCurrentIndex(fold - 1)
            except Exception:
                pass

            if not isinstance(fit, dict):
                fit = None

            if fit is not None and ("a" in fit and "b" in fit):
                try:
                    a = float(fit["a"])
                    b = float(fit["b"])
                except Exception:
                    a = None
                    b = None

                if a is not None and b is not None:
                    std_errs = fit.get("std_errs", None)
                    std_errs_arr = None
                    if isinstance(std_errs, (list, tuple)) and len(std_errs) >= 2:
                        try:
                            std_errs_arr = np.array([float(std_errs[0]), float(std_errs[1])], dtype=float)
                        except Exception:
                            std_errs_arr = None
                    self.plot_fit(params=(a, b), std_errs=std_errs_arr)
            elif isinstance(shock, dict):
                self._shock_summary = dict(shock)
                self._set_summary_labels_from_dict(self._shock_summary)
        finally:
            self._suppress_emit = False
        if emit_change:
            self._emit_session_changed()

    def _selected_fold(self):
        try:
            n = int(self.fold_combo.currentText())
        except Exception:
            n = 1
        return max(1, min(4, n))

    def recalculate_shock_parameters(self):
        if not hasattr(self, "_drift_vals") or not hasattr(self, "_drift_errs"):
            QMessageBox.information(self, "Analyzer", "Please click 'Best Fit' first.")
            return

        n = self._selected_fold()
        self._update_shock_parameters(n)
        self.status.showMessage(f"Updated using Newkirk {n}-fold model.", 3000)
        self._emit_session_changed()

    def _update_shock_parameters(self, n):
        # Your updated n-fold formulas
        denom = n * 3.385
        drift_vals = self._drift_vals
        drift_errs = self._drift_errs

        shock_speed = (13853221.38 * np.abs(drift_vals)) / (
                self.freq * (np.log(self.freq ** 2 / denom) ** 2)
        )
        R_p = 4.32 * np.log(10) / np.log(self.freq ** 2 / denom)

        # Starting frequency (same logic you already use)
        percentile = 90
        start_freq = np.percentile(self.freq, percentile)
        if self.harmonic:
            start_freq = start_freq / 2

        idx = np.abs(self.freq - start_freq).argmin()
        f0 = self.freq[idx]
        drift_err0 = drift_errs[idx]

        start_shock_speed = shock_speed[idx]
        start_height = R_p[idx]

        shock_speed_err = (13853221.38 * drift_err0) / (
                f0 * (np.log(f0 ** 2 / denom) ** 2)
        )

        # Error propagation for R_p based on your n-fold expression
        g0 = np.log(f0 ** 2 / denom)
        dRp_df = 8.64 * np.log(10) / (f0 * (g0 ** 2))
        Rp_err = np.abs(dRp_df * self.freq_err)

        # Averages (drift and freq do not depend on n, speeds/heights do)
        avg_freq = np.mean(self.freq)
        avg_freq_err = np.std(self.freq) / np.sqrt(len(self.freq))
        avg_drift = np.mean(drift_vals)
        avg_drift_err = np.std(drift_vals) / np.sqrt(len(drift_vals))

        avg_speed = np.mean(shock_speed)
        avg_speed_err = np.std(shock_speed) / np.sqrt(len(shock_speed))
        avg_height = np.mean(R_p)
        avg_height_err = np.std(R_p) / np.sqrt(len(R_p))

        # Store arrays for extra plots
        self.shock_speed = shock_speed
        self.R_p = R_p
        self.start_freq = start_freq
        self.start_height = start_height

        self._shock_summary = {
            "avg_freq_mhz": float(avg_freq),
            "avg_freq_err_mhz": float(avg_freq_err),
            "avg_drift_mhz_s": float(avg_drift),
            "avg_drift_err_mhz_s": float(avg_drift_err),
            "start_freq_mhz": float(start_freq),
            "start_freq_err_mhz": float(self.freq_err),
            "initial_shock_speed_km_s": float(start_shock_speed),
            "initial_shock_speed_err_km_s": float(shock_speed_err),
            "initial_shock_height_rs": float(start_height),
            "initial_shock_height_err_rs": float(Rp_err),
            "avg_shock_speed_km_s": float(avg_speed),
            "avg_shock_speed_err_km_s": float(avg_speed_err),
            "avg_shock_height_rs": float(avg_height),
            "avg_shock_height_err_rs": float(avg_height_err),
            "fold": int(n),
            "fundamental": bool(self.fundamental),
            "harmonic": bool(self.harmonic),
        }

        self._set_summary_labels_from_dict(self._shock_summary)

    def save_graph(self):
        plot_name = getattr(self, "current_plot_title", None) or f"{self.filename}_Plot"

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        file_path, ext = pick_export_path(
            self,
            "Export Figure",
            plot_name,
            formats,
            default_filter="PNG (*.png)"
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        try:
            # ✅ If user didn't type an extension, add the one from ext
            root, current_ext = os.path.splitext(file_path)
            if current_ext == "":
                ext = ext.lower().lstrip(".")
                file_path = f"{file_path}.{ext}"
            else:
                ext = current_ext.lower().lstrip(".")

            self.canvas.figure.savefig(
                file_path,
                dpi=300,
                bbox_inches="tight",
                format=ext
            )
            QMessageBox.information(self, "Export Complete", f"Plot saved:\n{file_path}")
            self.status.showMessage("Export successful!", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not save file:\n{e}")
            self.status.showMessage("Export failed!", 3000)

    def save_data(self):

        # All known e-Callisto station names
        station_list = [
            'ALASKA-ANCHORAGE', 'ALASKA-COHOE', 'ALASKA-HAARP', 'ALGERIA-CRAAG', 'ALMATY',
            'Arecibo-observatory', 'AUSTRIA-Krumbach', 'AUSTRIA-MICHELBACH', 'AUSTRIA-OE3FLB',
            'AUSTRIA-UNIGRAZ', 'Australia-ASSA', 'BRAZIL', 'BIR', 'Croatia-Visnjan', 'DENMARK',
            'EGYPT-Alexandria', 'EGYPT-SpaceAgency', 'ETHIOPIA', 'FINLAND-Siuntio', 'FINLAND-Kempele',
            'GERMANY-ESSEN', 'GERMANY-DLR', 'GLASGOW', 'GREENLAND', 'HUMAIN', 'HURBANOVO',
            'INDIA-GAURI', 'INDIA-Nashik', 'INDIA-OOTY', 'INDIA-UDAIPUR', 'INDONESIA',
            'ITALY-Strassolt', 'JAPAN-IBARAKI', 'KASI', 'KRIM', 'MEXART',
            'MEXICO-ENSENADA-UNAM', 'MEXICO-FCFM-UANL', 'MEXICO-FCFM-UNACH', 'MEXICO-LANCE-A',
            'MEXICO-LANCE-B', 'MEXICO-UANL-INFIERNILLO', 'MONGOLIA-UB', 'MRO', 'MRT1', 'MRT3',
            'Malaysia_Banting', 'NASA-GSFC', 'NORWAY-EGERSUND', 'NORWAY-NY-AALESUND', 'NORWAY-RANDABERG',
            'PARAGUAY', 'POLAND-BALDY', 'POLAND-Grotniki', 'ROMANIA', 'ROSWELL-NM', 'RWANDA',
            'SOUTHAFRICA-SANSA', 'SPAIN-ALCALA', 'SPAIN-PERALEJOS', 'SPAIN-SIGUENZA', 'SRI-Lanka',
            'SSRT', 'SWISS-CalU', 'SWISS-FM', 'SWISS-HB9SCT', 'SWISS-HEITERSWIL', 'SWISS-IRSOL',
            'SWISS-Landschlacht', 'SWISS-MUHEN', 'TAIWAN-NCU', 'THAILAND-Pathumthani', 'TRIEST',
            'TURKEY', 'UNAM', 'URUGUAY', 'USA-ARIZONA-ERAU', 'USA-BOSTON', 'UZBEKISTAN'
        ]

        # ✅ Extract Station
        station = "UNKNOWN"
        filename_lower = self.filename.lower()
        for s in station_list:
            if filename_lower.startswith(s.lower()):
                station = s
                break

        # ✅ Extract Date
        date_match = re.search(r'_(\d{4})(\d{2})(\d{2})_', self.filename)
        if date_match:
            date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
        else:
            date = "UNKNOWN"

        # ✅ Excel File Handling
        if self.existing_excel_checkbox.isChecked():
            path, _ = QFileDialog.getOpenFileName(self, "Select Existing Excel File", "", "Excel Files (*.xlsx)")
            if not path:
                return
            try:
                wb = load_workbook(path)
                ws = wb.active
            except Exception as e:
                QMessageBox.critical(self, "Load Error", f"Could not open Excel file:\n{str(e)}")
                return
        else:
            path, _ = QFileDialog.getSaveFileName(self, "Save as Excel", f"{self.filename}_data.xlsx",
                                                  "Excel Files (*.xlsx)")
            if not path:
                return
            try:
                wb = Workbook()
                ws = wb.active
                headers = [
                    "Date", "Station", "Best_fit", "R_sq", "RMSE",
                    "avg_freq", "avg_freq_err", "Avg_drift", "avg_drift_err",
                    "start_freq", "start_freq_err", "initial_shock_speed", "initial_shock_speed_err",
                    "initial_shock_height", "initial_shock_height_err", "avg_shock_speed", "avg_shock_speed_err",
                    "avg_shock_height", "avg_shock_height_err", "avg_drift_abs"
                ]
                ws.append(headers)
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Could not create Excel file:\n{str(e)}")
                return

        # ✅ Extract and clean text
        def extract_val_err(label):
            # Remove HTML tags
            clean_text = re.sub(r'<[^>]+>', '', label.text())
            # Remove units and stray characters
            clean_text = re.sub(r'(MHz|km/s|Rₛ|s|/)', '', clean_text)
            # Clean spaces
            clean_text = clean_text.strip()
            # Extract value ± error
            value_text = clean_text.split(":")[-1].strip()
            if "±" in value_text:
                value, err = value_text.split("±")
                return value.strip(), err.strip()
            else:
                return value_text.strip(), ""

        # ✅ Read values
        try:
            best_fit = re.sub(r'<[^>]+>', '', self.equation_display.text()).replace("<sup>", "^").replace("</sup>", "")
            r2 = self.r2_display.text().split("=")[-1].strip()
            rmse = self.rmse_display.text().split("=")[-1].strip()

            avg_freq, avg_freq_err = extract_val_err(self.avg_freq_display)
            avg_drift, avg_drift_err = extract_val_err(self.drift_display)

            try:
                avg_drift_abs = abs(float(avg_drift))
            except ValueError:
                avg_drift_abs = ""

            start_freq, start_freq_err = extract_val_err(self.start_freq_display)
            init_speed, init_speed_err = extract_val_err(self.initial_shock_speed_display)
            init_height, init_height_err = extract_val_err(self.initial_shock_height_display)
            avg_speed, avg_speed_err = extract_val_err(self.avg_shock_speed_display)
            avg_height, avg_height_err = extract_val_err(self.avg_shock_height_display)

            row = [
                date, station, best_fit, r2, rmse,
                avg_freq, avg_freq_err, avg_drift, avg_drift_err,
                start_freq, start_freq_err, init_speed, init_speed_err,
                init_height, init_height_err, avg_speed, avg_speed_err,
                avg_height, avg_height_err, avg_drift_abs
            ]

            ws.append(row)
            wb.save(path)
            self.status.showMessage("✅ Data saved to Excel successfully!", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Write Error", f"Could not write to Excel file:\n{str(e)}")
            self.status.showMessage("❌ Failed to save data to Excel.", 3000)

    def plot_extra(self):
        choice = self.extra_plot_combo.currentText()
        self.canvas.ax.clear()
        if choice == "Shock Speed vs Shock Height":
            self.canvas.ax.scatter(self.R_p, self.shock_speed, color='green', s=10)
            self.canvas.ax.set_xlabel("Shock Height (Rₛ)")
            self.canvas.ax.set_ylabel("Shock Speed (km/s)")
            self.canvas.ax.set_title(f"{self.filename}_Shock_Speed_vs_Shock_Height")
            self.current_plot_title = f"{self.filename}_Shock_Speed_vs_Shock_Height"
            self.status.showMessage("Shock Speed vs Shock Height plotted successfully!", 3000)

        elif choice == "Shock Speed vs Frequency":
            self.canvas.ax.scatter(self.freq, self.shock_speed, color='purple', s=10)
            self.canvas.ax.set_xlabel("Frequency (MHz)")
            self.canvas.ax.set_ylabel("Shock Speed (km/s)")
            self.canvas.ax.set_title(f"{self.filename}_Shock_Speed_vs_Frequency")
            self.current_plot_title = f"{self.filename}_Shock_Speed_vs_Frequency"
            self.status.showMessage("Shock Speed vs Frequency plotted successfully!", 3000)

        elif choice == "Shock Height vs Frequency":
            self.canvas.ax.scatter(self.R_p, self.freq, color='red', s=10)
            self.canvas.ax.set_xlabel("Shock Height (Rₛ)")
            self.canvas.ax.set_ylabel("Frequency (MHz)")
            self.canvas.ax.set_title(f"{self.filename}_Rs_vs_Freq")
            self.current_plot_title = f"{self.filename}_Rs_vs_Freq"
            self.status.showMessage("Shock Height vs Frequency plotted successfully!", 3000)
        self.canvas.ax.grid(True)
        self.canvas.draw()

    def closeEvent(self, event):
        self._emit_session_changed()
        super().closeEvent(event)
