"""
e-CALLISTO FITS Analyzer
Version 1.7.4
Sahan S Liyanage (sahanslst@gmail.com)
Astronomical and Space Science Unit, University of Colombo, Sri Lanka.
"""
import sys
import io
import os
import tempfile
import re
import gc
import requests
from PySide6.QtWidgets import (
    QMainWindow, QSlider, QDialog, QMenuBar, QMessageBox, QLabel, QFormLayout, QGroupBox,
    QStatusBar, QProgressBar, QApplication, QMenu, QCheckBox, QRadioButton, QButtonGroup, QComboBox, QToolBar,
    QLineEdit, QSpinBox, QScrollArea, QFrame, QVBoxLayout, QWidget, QFileDialog, QHBoxLayout, QSizePolicy
)
from PySide6.QtCore import QObject, QEvent
from PySide6.QtWidgets import QLayout
from PySide6.QtGui import QAction, QPixmap, QImage, QGuiApplication, QIcon, QFontDatabase
from PySide6.QtCore import Qt
from PySide6.QtCore import QTimer, QSize
from callisto_downloader import CallistoDownloaderApp
from goes_xrs_gui import MainWindow as GoesXrsWindow
#from soho_lasco_viewer import CMEViewer as CMEViewerWindow
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.widgets import LassoSelector
from matplotlib.path import Path
from astropy.io import fits
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.colors as mcolors
from mpl_toolkits.axes_grid1 import make_axes_locatable
from matplotlib.ticker import FuncFormatter, ScalarFormatter
import csv
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

from PySide6.QtCore import QObject, QEvent
from PySide6.QtWidgets import QLayout

#LINUX Specific Fixes for messageboxes

IS_LINUX = sys.platform.startswith("linux")

_linux_msgbox_fixer = None

class _LinuxMessageBoxFixer(QObject):
    def eventFilter(self, obj, event):
        if IS_LINUX and event.type() == QEvent.Show and isinstance(obj, QMessageBox):
            # Make the main text label wrap and give it room
            label = obj.findChild(QLabel, "qt_msgbox_label")
            if label:
                label.setWordWrap(True)
                label.setTextInteractionFlags(Qt.TextSelectableByMouse)
                label.setMinimumWidth(520)

            # Allow the dialog to grow to its content
            obj.setSizeGripEnabled(True)
            obj.setMinimumWidth(560)

            lay = obj.layout()
            if lay:
                lay.setSizeConstraint(QLayout.SetMinimumSize)

            QTimer.singleShot(0, obj.adjustSize)

        return super().eventFilter(obj, event)

def _install_linux_msgbox_fixer():
    global _linux_msgbox_fixer
    if not IS_LINUX:
        return
    app = QApplication.instance()
    if app is None or _linux_msgbox_fixer is not None:
        return
    _linux_msgbox_fixer = _LinuxMessageBoxFixer(app)
    app.installEventFilter(_linux_msgbox_fixer)




def start_combine(self):
    QTimer.singleShot(100, self.combine_files)  # delays execution and avoids UI freeze

#Uncomment for windows build
"""
def resource_path(relative_path: str) -> str:
    if hasattr(sys, "_MEIPASS"):
        # Packaged app
        return os.path.join(sys._MEIPASS, relative_path)
    # Development mode
    base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)
"""

#Uncomment for Linux build
def resource_path(relative_path: str) -> str:
    # PyInstaller sets sys._MEIPASS
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
        return os.path.join(base_path, relative_path)

    # Development
    base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

#Uncomment for Linux build
"""
def resource_path(relative_path: str) -> str:
   # py2app
    if getattr(sys, "frozen", False):
        base_path = os.path.abspath(
            os.path.join(os.path.dirname(sys.executable), "..", "Resources")
        )
        return os.path.join(base_path, relative_path)
    # Development
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)
"""

#Fix for figure saving issue on Linux
def _ext_from_filter(name_filter: str) -> str:
    m = re.search(r"\*\.(\w+)", name_filter or "")
    return m.group(1).lower() if m else ""

def pick_export_path(parent, caption: str, default_name: str, filters: str, default_filter: str = None):
    """
    Returns (path, ext).
    Linux uses a QFileDialog instance (non-native) so selectedNameFilter is reliable.
    Windows/macOS keep using getSaveFileName.
    """
    if IS_LINUX:
        dlg = QFileDialog(parent, caption)
        dlg.setAcceptMode(QFileDialog.AcceptSave)
        dlg.setNameFilters(filters.split(";;"))
        if default_filter:
            dlg.selectNameFilter(default_filter)
        dlg.selectFile(default_name)

        # Important for Linux reliability
        dlg.setOption(QFileDialog.DontUseNativeDialog, True)

        if not dlg.exec():
            return "", ""
        path = dlg.selectedFiles()[0]
        chosen_filter = dlg.selectedNameFilter()
    else:
        path, chosen_filter = QFileDialog.getSaveFileName(parent, caption, default_name, filters)
        if not path:
            return "", ""

    ext = os.path.splitext(path)[1].lstrip(".").lower()

    # If user didn’t type an extension, take it from the selected filter
    if not ext:
        ext = _ext_from_filter(chosen_filter) or "png"
        path = f"{path}.{ext}"

    return path, ext


class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=10, height=6, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.updateGeometry()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        #Linux Messagebox Fix
        _install_linux_msgbox_fixer()

        self.setWindowTitle("e-CALLISTO FITS Analyzer 1.7.4")
        #self.resize(1000, 700)
        self.setMinimumSize(1000, 700)

        self.use_utc = False
        self.ut_start_sec = None
        self.use_db = False  # False = Digits (default), True = dB

        # --- Undo / Redo ---
        self._undo_stack = []
        self._redo_stack = []
        self._max_undo = 30  # prevent memory blow-up

        # --- Graph Properties (non-colormap) ---
        self.graph_title_override = ""  # empty = use default "{filename} - {title}"
        self.graph_font_family = ""  # empty = use Matplotlib default

        self.tick_font_px = 11
        self.axis_label_font_px = 12
        self.title_font_px = 14

        self._colorbar_label_text = ""

        self.current_cmap_name = "Custom"
        self.lasso_active = False

        self.noise_vmin = None
        self.noise_vmax = None

        self.current_display_data = None

        # --- Graph Properties: style flags ---
        self.title_bold = False
        self.title_italic = False

        self.axis_bold = False
        self.axis_italic = False

        self.ticks_bold = False
        self.ticks_italic = False

        self.remove_titles = False

        self._build_toolbar()

        # Debounce timer for smooth slider updates
        self.noise_smooth_timer = QTimer()
        self.noise_smooth_timer.setInterval(40)
        self.noise_smooth_timer.setSingleShot(True)
        self.noise_smooth_timer.timeout.connect(self.update_noise_live)

        # Canvas
        self.canvas = MplCanvas(self, width=10, height=6)
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        self.canvas.mpl_connect("scroll_event", self.on_scroll_zoom)
        self._cid_press = self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self._cid_release = self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self._cid_motion = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
        self._cid_motion_status = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_motion_status)

        self._panning = False
        self._last_pan_xy = None

        # Colorbar
        self.current_colorbar = None
        self.current_cax = None

        # Statusbar
        self.setStatusBar(QStatusBar())
        # Permanent label on right side for cursor coordinates
        self.cursor_label = QLabel("")
        self.cursor_label.setStyleSheet("padding-right: 8px;")
        self.statusBar().addPermanentWidget(self.cursor_label)

        # =========================
        # LEFT SIDEBAR (CROSS-PLATFORM SAFE)
        # Put this whole block where you currently build:
        #   slider_group, units_group_box, graph_group, main_layout, container
        #
        # Required imports (add if missing):
        #   from PySide6.QtWidgets import QScrollArea, QFrame
        # =========================

        # -------------------------
        # Noise clipping sliders
        # -------------------------
        self.lower_slider = QSlider(Qt.Horizontal)
        self.lower_slider.setRange(-100, 100)
        self.lower_slider.setValue(0)

        self.upper_slider = QSlider(Qt.Horizontal)
        self.upper_slider.setRange(-100, 100)
        self.upper_slider.setValue(0)

        slider_group = QGroupBox("Noise Clipping Thresholds")
        slider_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        slider_layout = QVBoxLayout(slider_group)
        slider_layout.setContentsMargins(10, 10, 10, 10)
        slider_layout.setSpacing(6)

        lbl_low = QLabel("Lower Threshold")
        lbl_low.setAlignment(Qt.AlignLeft)
        slider_layout.addWidget(lbl_low)
        slider_layout.addWidget(self.lower_slider)

        lbl_high = QLabel("Upper Threshold")
        lbl_high.setAlignment(Qt.AlignLeft)
        slider_layout.addWidget(lbl_high)
        slider_layout.addWidget(self.upper_slider)

        # -------------------------
        # Units Group (Intensity + Time in one row)
        # -------------------------
        self.units_group_box = QGroupBox("Units")
        self.units_group_box.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        units_layout = QVBoxLayout(self.units_group_box)
        units_layout.setContentsMargins(10, 10, 10, 10)
        units_layout.setSpacing(8)

        # ---- Horizontal container for Intensity + Time ----
        units_row = QHBoxLayout()
        units_row.setSpacing(20)

        # ===== Intensity column =====
        intensity_col = QVBoxLayout()
        intensity_col.setSpacing(8)

        intensity_label = QLabel("Intensity")
        intensity_label.setProperty("section", True)

        self.units_digits_radio = QRadioButton("Digits")
        self.units_db_radio = QRadioButton("dB")
        self.units_digits_radio.setChecked(True)

        self.units_group = QButtonGroup(self)
        self.units_group.addButton(self.units_digits_radio)
        self.units_group.addButton(self.units_db_radio)

        intensity_col.addWidget(intensity_label)
        intensity_col.addWidget(self.units_digits_radio)
        intensity_col.addWidget(self.units_db_radio)
        intensity_col.addStretch(1)

        # ===== Time column =====
        time_col = QVBoxLayout()
        time_col.setSpacing(6)

        time_label = QLabel("Time")
        time_label.setProperty("section", True)

        self.time_sec_radio = QRadioButton("Seconds")
        self.time_ut_radio = QRadioButton("UT")
        self.time_sec_radio.setChecked(True)

        self.time_group = QButtonGroup(self)
        self.time_group.addButton(self.time_sec_radio)
        self.time_group.addButton(self.time_ut_radio)

        time_col.addWidget(time_label)
        time_col.addWidget(self.time_sec_radio)
        time_col.addWidget(self.time_ut_radio)
        time_col.addStretch(1)

        # ---- Add both columns to the row ----
        units_row.addLayout(intensity_col, 1)
        units_row.addLayout(time_col, 1)

        # ---- Add row to Units group ----
        units_layout.addLayout(units_row)
        units_layout.addStretch(1)

        # -------------------------
        # Graph Properties Group
        # -------------------------
        def _section_label(text: str) -> QLabel:
            lbl = QLabel(text)
            lbl.setObjectName("SectionLabel")
            return lbl

        def _spin_row(label_text: str, spin: QSpinBox) -> QWidget:
            w = QWidget()
            row = QHBoxLayout(w)
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)

            label = QLabel(label_text)
            label.setWordWrap(False)
            label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

            spin.setMinimumWidth(90)
            spin.setMaximumWidth(110)
            spin.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            row.addWidget(label, 1)
            row.addWidget(spin, 0, Qt.AlignRight)
            return w

        def _style_row(label_text: str, cb_bold: QCheckBox, cb_italic: QCheckBox) -> QWidget:
            w = QWidget()
            row = QHBoxLayout(w)
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)

            label = QLabel(label_text)
            label.setWordWrap(False)
            label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

            row.addWidget(label, 1)
            row.addWidget(cb_bold, 0)
            row.addWidget(cb_italic, 0)
            return w

        self.graph_group = QGroupBox("Graph Properties")
        self.graph_group.setEnabled(False)
        self.graph_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        graph_layout = QVBoxLayout(self.graph_group)
        graph_layout.setContentsMargins(10, 10, 10, 10)
        graph_layout.setSpacing(6)

        # Appearance
        graph_layout.addWidget(_section_label("Appearance"))

        self.cmap_combo = QComboBox()
        self.cmap_combo.addItems([
            "Custom", "viridis", "plasma", "inferno", "magma",
            "cividis", "turbo", "RdYlBu", "jet", "cubehelix",
        ])
        graph_layout.addWidget(QLabel("Colormap"))
        graph_layout.addWidget(self.cmap_combo)

        self.font_combo = QComboBox()
        self.font_combo.addItem("Default")
        for f in sorted(QFontDatabase.families()):
            self.font_combo.addItem(f)

        graph_layout.addWidget(QLabel("Font family"))
        graph_layout.addWidget(self.font_combo)

        # Text
        graph_layout.addWidget(_section_label("Text"))

        self.title_edit = QLineEdit()
        self.title_edit.setPlaceholderText("Custom title (leave empty for default)")
        self.title_edit.setMinimumHeight(20)
        graph_layout.addWidget(QLabel("Graph title"))
        graph_layout.addWidget(self.title_edit)

        self.remove_titles_chk = QCheckBox("Remove Titles")
        graph_layout.addWidget(self.remove_titles_chk)

        # Font sizes
        graph_layout.addWidget(_section_label("Font sizes"))

        self.tick_font_spin = QSpinBox()
        self.tick_font_spin.setRange(6, 60)
        self.tick_font_spin.setValue(self.tick_font_px)
        graph_layout.addWidget(_spin_row("Tick labels (px)", self.tick_font_spin))

        self.axis_font_spin = QSpinBox()
        self.axis_font_spin.setRange(6, 60)
        self.axis_font_spin.setValue(self.axis_label_font_px)
        graph_layout.addWidget(_spin_row("Axis labels (px)", self.axis_font_spin))

        self.title_font_spin = QSpinBox()
        self.title_font_spin.setRange(6, 80)
        self.title_font_spin.setValue(self.title_font_px)
        graph_layout.addWidget(_spin_row("Title (px)", self.title_font_spin))

        # Text style
        graph_layout.addWidget(_section_label("Text style"))

        self.title_bold_chk = QCheckBox("Bold")
        self.title_italic_chk = QCheckBox("Italic")
        graph_layout.addWidget(_style_row("Title", self.title_bold_chk, self.title_italic_chk))

        self.axis_bold_chk = QCheckBox("Bold")
        self.axis_italic_chk = QCheckBox("Italic")
        graph_layout.addWidget(_style_row("Axis labels", self.axis_bold_chk, self.axis_italic_chk))

        self.ticks_bold_chk = QCheckBox("Bold")
        self.ticks_italic_chk = QCheckBox("Italic")
        graph_layout.addWidget(_style_row("Tick labels", self.ticks_bold_chk, self.ticks_italic_chk))

        graph_layout.addStretch(1)

        # -------------------------
        # Build the LEFT PANEL as a widget, then put it in a ScrollArea
        # This is the key fix for Windows (no overlaps, no clipping).
        # -------------------------
        side_panel_widget = QWidget()
        side_panel_layout = QVBoxLayout(side_panel_widget)
        side_panel_layout.setContentsMargins(8, 8, 8, 8)
        side_panel_layout.setSpacing(10)

        side_panel_layout.addWidget(slider_group)
        side_panel_layout.addWidget(self.units_group_box)
        side_panel_layout.addWidget(self.graph_group)
        side_panel_layout.addStretch(1)

        # Consistent width for all groups (better on Windows DPI scaling)
        SIDEBAR_W = 250
        slider_group.setMaximumWidth(SIDEBAR_W)
        self.units_group_box.setMaximumWidth(SIDEBAR_W)
        self.graph_group.setMaximumWidth(SIDEBAR_W)

        side_scroll = QScrollArea()
        side_scroll.setWidgetResizable(True)
        side_scroll.setFrameShape(QFrame.NoFrame)
        side_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        side_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        side_scroll.setMinimumWidth(SIDEBAR_W + 16)  # room for scrollbar
        side_scroll.setMaximumWidth(SIDEBAR_W + 28)
        side_scroll.setWidget(side_panel_widget)

        # -------------------------
        # Style (safe sizes, no tiny max-heights)
        # -------------------------
        sidebar_style = """
        QGroupBox {
            font-weight: bold;
        }
        QLabel {
            font-size: 12px;
        }
        QLabel[section="true"] {
            font-weight: bold;
            color: #444;
            margin-top: 6px;
        }
        QLabel#SectionLabel {
            font-weight: bold;
            color: #555;
            margin-top: 8px;
            margin-bottom: 4px;
        }
        
        QLineEdit, QComboBox {
            min-height: 24px;
            padding: 4px 6px;
            font-size: 12px;
        }

        QSpinBox {
            min-height: 32px;      /* REQUIRED for Windows */
            min-width: 90px;
            padding-left: 6px;
            font-size: 12px;
        }

        QCheckBox {
            spacing: 6px;
            font-size: 12px;
            margin-top: 10px;
            margin-bottom: 10px;
            margin-right: 10px;
            
        }
        """
        slider_group.setStyleSheet(sidebar_style)
        self.units_group_box.setStyleSheet(sidebar_style)
        self.graph_group.setStyleSheet(sidebar_style)

        # -------------------------
        # Main layout with scrollable sidebar + canvas
        # -------------------------
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(10)

        main_layout.addWidget(side_scroll, 0)
        main_layout.addWidget(self.canvas, 1)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        # ----- Menu Bar -----
        menubar = self.menuBar()

        # File Menu
        file_menu = menubar.addMenu("File")

        # --- Open ---
        self.open_action = QAction("Open", self)
        file_menu.addAction(self.open_action)

        # --- Save As (disabled for main window) ---
        self.save_action = QAction("Save As", self)
        self.save_action.setEnabled(False)
        file_menu.addAction(self.save_action)

        # --- Export As ---
        self.export_action = QAction("Export As", self)
        file_menu.addAction(self.export_action)
        self.export_action.triggered.connect(self.export_figure)

        # Edit Menu
        edit_menu = menubar.addMenu("Edit")

        self.undo_action = QAction("Undo", self)
        self.undo_action.setShortcut("Ctrl+Z")
        edit_menu.addAction(self.undo_action)

        self.redo_action = QAction("Redo", self)
        self.redo_action.setShortcut("Ctrl+Shift+Z")
        edit_menu.addAction(self.redo_action)

        edit_menu.addSeparator()

        reset_action = QAction("Reset All", self)
        edit_menu.addAction(reset_action)
        reset_action.triggered.connect(self.reset_all)

        self.undo_action.triggered.connect(self.undo)
        self.redo_action.triggered.connect(self.redo)

        # Download Menu
        download_menu = menubar.addMenu("Download")

        # --- Launch Downloader ---
        launch_downloader_action = QAction("Launch FITS Downloader", self)
        download_menu.addAction(launch_downloader_action)
        launch_downloader_action.triggered.connect(self.launch_downloader)

        # Combine Menu
        combine_menu = menubar.addMenu("Combine FITS")

        combine_freq_action = QAction("Combine Frequency", self)
        combine_freq_action.triggered.connect(self.open_combine_freq_window)
        combine_menu.addAction(combine_freq_action)

        combine_time_action = QAction("Combine Time", self)
        combine_time_action.triggered.connect(self.open_combine_time_window)
        combine_menu.addAction(combine_time_action)

        # Graph Menu
        graph_menu = menubar.addMenu("Graph")

        xaxis_unit_menu = QMenu("x-axis units", self)
        self.xaxis_sec_action = QAction("Seconds (s)", self, checkable=True)
        self.xaxis_ut_action = QAction("Universal Time (UT)", self, checkable=True)

        # Make "Seconds" default
        self.xaxis_sec_action.setChecked(True)

        xaxis_unit_menu.addAction(self.xaxis_sec_action)
        xaxis_unit_menu.addAction(self.xaxis_ut_action)
        graph_menu.addMenu(xaxis_unit_menu)

        # Toggle logic
        self.xaxis_sec_action.triggered.connect(self.set_axis_to_seconds)
        self.xaxis_ut_action.triggered.connect(self.set_axis_to_utc)

        # CMEs
        cmes_menu = self.menuBar().addMenu("CME")
        soho_lasco_action = QAction("SOHO/LASCO CME Catalog", self)
        soho_lasco_action.triggered.connect(self.open_cme_viewer)
        cmes_menu.addAction(soho_lasco_action)

        # Flares
        flares_menu = self.menuBar().addMenu("Flares")
        goes_flux_action = QAction("GOES X-Ray Flux", self)
        goes_flux_action.triggered.connect(self.open_goes_xrs_window)
        flares_menu.addAction(goes_flux_action)

        # About Menu
        about_menu = menubar.addMenu("About")
        about_action = QAction("About", self)
        about_action.setMenuRole(QAction.NoRole)
        about_menu.addAction(about_action)
        about_action.triggered.connect(self.show_about_dialog)

        # (OPTIONAL) Connect them later like:
        # open_action.triggered.connect(self.open_file)

        self.setCentralWidget(container)

        # Signals
        self.lower_slider.valueChanged.connect(self.schedule_noise_update)
        self.upper_slider.valueChanged.connect(self.schedule_noise_update)
        self.cmap_combo.currentTextChanged.connect(self.change_cmap)
        self.open_action.triggered.connect(self.load_file)
        self.units_digits_radio.toggled.connect(self.update_units)
        self.units_db_radio.toggled.connect(self.update_units)

        self.time_sec_radio.toggled.connect(
            lambda checked: checked and self.set_axis_to_seconds()
        )
        self.time_ut_radio.toggled.connect(
            lambda checked: checked and self.set_axis_to_utc()
        )

        # Keep existing colormap live behavior
        self.cmap_combo.currentTextChanged.connect(self.change_cmap)

        # Real-time graph properties (non-colormap)
        self.title_edit.textChanged.connect(self.apply_graph_properties_live)
        self.font_combo.currentTextChanged.connect(self.apply_graph_properties_live)
        self.tick_font_spin.valueChanged.connect(self.apply_graph_properties_live)
        self.axis_font_spin.valueChanged.connect(self.apply_graph_properties_live)
        self.title_font_spin.valueChanged.connect(self.apply_graph_properties_live)

        # Real-time style toggles
        self.remove_titles_chk.toggled.connect(self.apply_graph_properties_live)

        self.title_bold_chk.toggled.connect(self.apply_graph_properties_live)
        self.title_italic_chk.toggled.connect(self.apply_graph_properties_live)

        self.axis_bold_chk.toggled.connect(self.apply_graph_properties_live)
        self.axis_italic_chk.toggled.connect(self.apply_graph_properties_live)

        self.ticks_bold_chk.toggled.connect(self.apply_graph_properties_live)
        self.ticks_italic_chk.toggled.connect(self.apply_graph_properties_live)

        # Data placeholders
        self.raw_data = None
        self.freqs = None
        self.time = None
        self.filename = ""
        self.current_plot_type = "Raw"  # or "NoiseReduced" or "Isolated"

        self.lasso = None
        self.lasso_mask = None
        self.noise_reduced_data = None

        self.setStyleSheet("""
            QLabel {
                font-size: 13px;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
            }
        """)

        self.noise_reduced_original = None  # backup before lasso

    def _icon(self, filename: str) -> QIcon:
        icon_path = resource_path(os.path.join("assets", "icons", filename))

        if os.path.exists(icon_path):
            return QIcon(icon_path)

        print(f"⚠️ Icon not found: {icon_path}")
        return QIcon()

    def _build_toolbar(self):
        tb = QToolBar("Main Toolbar", self)
        tb.setMovable(False)
        tb.setIconSize(QSize(36, 36))
        self.addToolBar(tb)

        # --- Actions (toolbar) ---
        self.tb_open = QAction(self._icon("open.svg"), "Open / Load", self)
        self.tb_open.setShortcut("Ctrl+O")
        self.tb_open.triggered.connect(self.load_file)
        tb.addAction(self.tb_open)

        self.tb_export = QAction(self._icon("export.svg"), "Export", self)
        self.tb_export.setShortcut("Ctrl+E")
        self.tb_export.triggered.connect(self.export_figure)
        tb.addAction(self.tb_export)

        tb.addSeparator()

        self.tb_undo = QAction(self._icon("undo.svg"), "Undo", self)
        self.tb_undo.setShortcut("Ctrl+Z")
        self.tb_undo.triggered.connect(self.undo)
        tb.addAction(self.tb_undo)

        self.tb_redo = QAction(self._icon("redo.svg"), "Redo", self)
        self.tb_redo.setShortcut("Ctrl+Shift+Z")
        self.tb_redo.triggered.connect(self.redo)
        tb.addAction(self.tb_redo)

        tb.addSeparator()

        self.tb_download = QAction(self._icon("download.svg"), "Download", self)
        self.tb_download.triggered.connect(self.launch_downloader)
        tb.addAction(self.tb_download)

        tb.addSeparator()

        self.tb_drift = QAction(self._icon("drift.svg"), "Estimate Drift Rate", self)
        self.tb_drift.triggered.connect(self.activate_drift_tool)
        tb.addAction(self.tb_drift)

        self.tb_isolate = QAction(self._icon("isolate.svg"), "Isolate Burst", self)
        self.tb_isolate.triggered.connect(self.activate_lasso)
        tb.addAction(self.tb_isolate)

        self.tb_max = QAction(self._icon("max.svg"), "Plot Maximum Intensities", self)
        self.tb_max.triggered.connect(self.plot_max_intensities)
        tb.addAction(self.tb_max)

        tb.addSeparator()

        self.tb_reset_sel = QAction(self._icon("reset_selection.svg"), "Reset Selection", self)
        self.tb_reset_sel.triggered.connect(self.reset_selection)
        tb.addAction(self.tb_reset_sel)

        self.tb_reset_all = QAction(self._icon("reset_all.svg"), "Reset All", self)
        self.tb_reset_all.triggered.connect(self.reset_all)
        tb.addAction(self.tb_reset_all)

        # Initial enable/disable states
        self._sync_toolbar_enabled_states()

    def _sync_toolbar_enabled_states(self):
        has_file = getattr(self, "raw_data", None) is not None
        has_noise = getattr(self, "noise_reduced_data", None) is not None
        has_undo = len(getattr(self, "_undo_stack", [])) > 0
        has_redo = len(getattr(self, "_redo_stack", [])) > 0
        filename = getattr(self, "filename", "")

        # Always allowed
        self.tb_open.setEnabled(True)
        self.tb_download.setEnabled(True)

        # Needs a plot / filename
        self.tb_export.setEnabled(bool(filename))

        # Undo/redo availability
        self.tb_undo.setEnabled(has_undo)
        self.tb_redo.setEnabled(has_redo)

        # Tools that require processed data
        self.tb_drift.setEnabled(has_noise)
        self.tb_isolate.setEnabled(has_noise)
        self.tb_max.setEnabled(has_noise)
        self.tb_reset_sel.setEnabled(has_noise)
        self.tb_reset_all.setEnabled(has_file)

    def apply_graph_properties_live(self, *_):
        """
        Apply graph styling changes (title, font family, font sizes, bold/italic flags)
        to the CURRENT plot without rebuilding the image.
        """

        ax = getattr(self.canvas, "ax", None)
        if ax is None:
            return

        # Must have an image already
        if not ax.images or len(ax.images) == 0:
            return

        # -----------------------------
        # 1) READ UI STATE FIRST
        # -----------------------------
        self.remove_titles = self.remove_titles_chk.isChecked()

        self.title_bold = self.title_bold_chk.isChecked()
        self.title_italic = self.title_italic_chk.isChecked()

        self.axis_bold = self.axis_bold_chk.isChecked()
        self.axis_italic = self.axis_italic_chk.isChecked()

        self.ticks_bold = self.ticks_bold_chk.isChecked()
        self.ticks_italic = self.ticks_italic_chk.isChecked()

        self.graph_title_override = self.title_edit.text().strip()

        font_choice = self.font_combo.currentText()
        self.graph_font_family = "" if font_choice == "Default" else font_choice

        self.tick_font_px = int(self.tick_font_spin.value())
        self.axis_label_font_px = int(self.axis_font_spin.value())
        self.title_font_px = int(self.title_font_spin.value())

        # Helpers
        def _weight(bold: bool) -> str:
            return "bold" if bold else "normal"

        def _style(italic: bool) -> str:
            return "italic" if italic else "normal"

        fontfam = self.graph_font_family if self.graph_font_family else None

        # -----------------------------
        # 2) APPLY TITLES / LABELS
        # -----------------------------
        if self.remove_titles:
            ax.set_title("")
            ax.set_xlabel("")
            ax.set_ylabel("")
        else:
            # Default title if custom is empty
            title_text = self.graph_title_override or f"{self.filename} - {self.current_plot_type}"

            ax.set_title(
                title_text,
                fontsize=self.title_font_px,
                fontfamily=fontfam,
                fontweight=_weight(self.title_bold),
                fontstyle=_style(self.title_italic),
            )

            # Always set axis labels explicitly so toggles apply immediately
            xlab = "Time [UT]" if (self.use_utc and self.ut_start_sec is not None) else "Time [s]"
            ylab = "Frequency [MHz]"

            ax.set_xlabel(
                xlab,
                fontsize=self.axis_label_font_px,
                fontfamily=fontfam,
                fontweight=_weight(self.axis_bold),
                fontstyle=_style(self.axis_italic),
            )
            ax.set_ylabel(
                ylab,
                fontsize=self.axis_label_font_px,
                fontfamily=fontfam,
                fontweight=_weight(self.axis_bold),
                fontstyle=_style(self.axis_italic),
            )

        # -----------------------------
        # 3) APPLY TICK LABEL STYLE
        # -----------------------------
        for lbl in ax.get_xticklabels():
            lbl.set_fontsize(self.tick_font_px)
            lbl.set_fontweight(_weight(self.ticks_bold))
            lbl.set_fontstyle(_style(self.ticks_italic))
            if fontfam:
                lbl.set_fontfamily(fontfam)

        for lbl in ax.get_yticklabels():
            lbl.set_fontsize(self.tick_font_px)
            lbl.set_fontweight(_weight(self.ticks_bold))
            lbl.set_fontstyle(_style(self.ticks_italic))
            if fontfam:
                lbl.set_fontfamily(fontfam)

        # Colorbar style (ticks + title)
        if self.current_colorbar is not None:
            try:
                cax = self.current_colorbar.ax
                fontfam = (self.graph_font_family if self.graph_font_family else None)

                # ---- ticks ----
                cax.tick_params(labelsize=self.tick_font_px)
                for lbl in cax.get_yticklabels():
                    lbl.set_fontsize(self.tick_font_px)
                    lbl.set_fontweight(_weight(self.ticks_bold))
                    lbl.set_fontstyle(_style(self.ticks_italic))
                    if fontfam:
                        lbl.set_fontfamily(fontfam)

                # ---- title/label ----
                if self.remove_titles:
                    # Hide the colorbar label, but DO NOT erase the stored text
                    self.current_colorbar.set_label("")
                    cax.set_ylabel("")  # extra safety for some backends
                    cax.yaxis.label.set_text("")
                else:
                    # Restore label text + style
                    label_text = getattr(self, "_colorbar_label_text", "")
                    if not label_text:
                        label_text = cax.get_ylabel()  # fallback

                    self.current_colorbar.set_label(
                        label_text,
                        fontsize=self.axis_label_font_px,
                        fontfamily=fontfam,
                        fontweight=_weight(self.axis_bold),
                        fontstyle=_style(self.axis_italic),
                    )

                    # Force styling on the underlying label object too
                    ylab = cax.yaxis.label
                    ylab.set_fontsize(self.axis_label_font_px)
                    ylab.set_fontweight(_weight(self.axis_bold))
                    ylab.set_fontstyle(_style(self.axis_italic))
                    if fontfam:
                        ylab.set_fontfamily(fontfam)

            except Exception:
                pass

        # -----------------------------
        # 4) UI: disable title inputs when Remove Titles is on
        # -----------------------------
        disable = self.remove_titles
        self.title_edit.setEnabled(not disable)
        self.title_bold_chk.setEnabled(not disable)
        self.title_italic_chk.setEnabled(not disable)
        self.axis_bold_chk.setEnabled(not disable)
        self.axis_italic_chk.setEnabled(not disable)

        self.canvas.draw_idle()

    def load_file(self):
        initial_dir = os.path.dirname(self.filename) if self.filename else ""
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFiles)
        dialog.setNameFilter("FITS files (*.fit *.fit.gz)")
        dialog.setOption(QFileDialog.DontUseNativeDialog, True)

        if dialog.exec():
            file_paths = dialog.selectedFiles()
        else:
            return

        if not file_paths:
            return

        if len(file_paths) == 1:
            file_path = file_paths[0]
            self.filename = os.path.basename(file_path)

            hdul = fits.open(file_path)
            self.raw_data = hdul[0].data
            self.freqs = hdul[1].data['frequency'][0]
            self.time = hdul[1].data['time'][0]

            # UT start
            hdr = hdul[0].header
            hh, mm, ss = hdr['TIME-OBS'].split(":")
            self.ut_start_sec = int(hh) * 3600 + int(mm) * 60 + float(ss)
            hdul.close()

            self.plot_data(self.raw_data, title="Raw Data")
            return

        basenames = [os.path.basename(p) for p in file_paths]

        pattern = r"(.*)_(\d{8})_(\d{6})_(\d+)\.fit(?:\.gz)?"

        meta = []
        for name in basenames:
            m = re.match(pattern, name)
            if not m:
                QMessageBox.warning(self, "Invalid File",
                                    f"Invalid CALLISTO filename format:\n{name}")
                return
            meta.append(m.groups())

        stations = [m[0] for m in meta]
        dates = [m[1] for m in meta]
        times = [m[2] for m in meta]
        ids = [int(m[3]) for m in meta]

        same_station = len(set(stations)) == 1
        same_date = len(set(dates)) == 1
        same_time = len(set(times)) == 1

        if same_station and same_date and same_time:
            combine_type = "frequency"

        elif same_station and len(set(ids)) == 1:
            combine_type = "time"
        else:
            error_msg = (

                "The selected FITS files cannot be combined.\n\n"

                "Valid combinations are:\n"

                "1. Frequency Combine:\n"

                "   • Same station\n"

                "   • Same date\n"

                "   • Same timestamp (HHMMSS)\n"

                "   • Different receiver IDs\n\n"

                "2. Time Combine:\n"

                "   • Same station\n"

                "   • Same receiver ID\n"

                "   • Same date\n"

                "   • Different timestamps (continuous time segments)\n\n"

                "Your selection does not match either rule.\n"

                "Please choose files that follow one of the above patterns."

            )

            QMessageBox.warning(self, "Invalid Combination Selection", error_msg)

            return

        if combine_type == "frequency":
            combined_data, combined_freqs, combined_time = self.combine_frequency_files(file_paths)

        elif combine_type == "time":
            combined_data, combined_freqs, combined_time = self.combine_time_files(file_paths)

        self.raw_data = combined_data
        self.freqs = combined_freqs
        self.time = combined_time

        # Extract UT start from FIRST FITS file
        try:
            hdul = fits.open(file_paths[0])
            hdr = hdul[0].header
            hh, mm, ss = hdr["TIME-OBS"].split(":")
            self.ut_start_sec = int(hh) * 3600 + int(mm) * 60 + float(ss)
            hdul.close()
        except Exception as e:
            print("⚠️ UT extraction failed:", e)
            self.ut_start_sec = None

        original_name = os.path.basename(file_paths[0])
        self.filename = original_name

        self.plot_data(self.raw_data, title="Raw Data")
        self.statusBar().showMessage(f"Loaded {len(file_paths)} files (combined)", 5000)

    def update_units(self):
        if self.units_db_radio.isChecked():
            self.use_db = True
        else:
            self.use_db = False

        if self.raw_data is None:
            return

        # Choose which data to replot
        data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data

        # Replot with new unit selection
        self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def combine_frequency_files(self, file_paths):
        file_paths = sorted(file_paths)
        data_list = []
        freq_list = []

        for path in file_paths:
            hdul = fits.open(path)
            data_list.append(hdul[0].data)
            freq_list.append(hdul[1].data['frequency'][0])
            time_array = hdul[1].data['time'][0]
            hdul.close()

        combined_data = np.concatenate(data_list, axis=0)
        combined_freqs = np.concatenate(freq_list)
        combined_time = time_array

        return combined_data, combined_freqs, combined_time

    def combine_time_files(self, file_paths):
        file_paths = sorted(file_paths)
        data_list = []
        time_list = []

        for path in file_paths:
            hdul = fits.open(path)
            data_list.append(hdul[0].data)
            time_list.append(hdul[1].data['time'][0])
            freqs = hdul[1].data['frequency'][0]
            hdul.close()

        combined_data = np.concatenate(data_list, axis=1)

        fixed_time = []
        offset = 0

        for t in time_list:
            t = np.array(t)
            fixed_time.append(t + offset)
            offset += t[-1]

        combined_time = np.concatenate(fixed_time)

        combined_freqs = freqs
        return combined_data, combined_freqs, combined_time

    def load_fits_into_main(self, file_path):
        hdul = fits.open(file_path)
        self.raw_data = hdul[0].data
        self.freqs = hdul[1].data['frequency'][0]
        self.time = hdul[1].data['time'][0]
        self.filename = os.path.basename(file_path)

        try:
            hdr = hdul[0].header
            hh, mm, ss = hdr['TIME-OBS'].split(":")
            hh = int(hh)
            mm = int(mm)
            ss = float(ss)
            self.ut_start_sec = hh * 3600 + mm * 60 + ss
        except Exception:
            self.ut_start_sec = None

        hdul.close()

        self.plot_data(self.raw_data, title="Raw Data")

    def load_combined_into_main(self, combined):
        self.raw_data = combined["data"]
        self.freqs = combined["freqs"]
        self.time = combined["time"]
        self.filename = combined.get("filename", "Combined")
        self.ut_start_sec = combined.get("ut_start_sec", None)
        self.plot_data(self.raw_data, title="Combined Data")

    def schedule_noise_update(self):
        if self.raw_data is None:
            return
        self.noise_smooth_timer.start()

    def update_noise_live(self):
        self._push_undo_state()
        if self.raw_data is None:
            return

        low = self.lower_slider.value()
        high = self.upper_slider.value()

        data = self.raw_data.copy()
        data = data - data.mean(axis=1, keepdims=True)
        data = np.clip(data, low, high)
        data = data * 2500.0 / 255.0 / 25.4

        self.noise_reduced_data = data
        self.noise_reduced_original = data.copy()

        self.noise_vmin = data.min()
        self.noise_vmax = data.max()

        self.plot_data(data, title="Noise Reduced")

        # enable tools
        self._sync_toolbar_enabled_states()

    def plot_data(self, data, title="Dynamic Spectrum", keep_view=False):
        view = self._capture_view() if keep_view else None
        QTimer.singleShot(0, lambda: self._plot_data_internal(data, title, view))

    def _capture_view(self):
        """Save current zoom/pan limits (only if a plot exists)."""
        try:
            ax = self.canvas.ax
            if ax is None or ax.images is None or len(ax.images) == 0:
                return None
            return {
                "xlim": ax.get_xlim(),
                "ylim": ax.get_ylim(),
            }
        except Exception:
            return None

    def _restore_view(self, view):
        """Restore zoom/pan limits safely."""
        if not view:
            return
        try:
            ax = self.canvas.ax
            ax.set_xlim(view["xlim"])
            ax.set_ylim(view["ylim"])
        except Exception:
            pass

    def _plot_data_internal(self, data, title="Dynamic Spectrum", view=None):
        if self.time is None or self.freqs is None:
            print("Time or frequency data not loaded. Skipping plot.")
            return

        if not hasattr(self.canvas, 'ax') or self.canvas.ax is None:
            print("Canvas not ready yet")
            return

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        # Remove old colorbar axis safely
        try:
            if self.current_cax:
                self.current_cax.remove()
                self.current_cax = None
            if self.current_colorbar:
                self.current_colorbar.remove()
                self.current_colorbar = None
        except Exception as e:
            print("Error removing previous colorbar:", e)

        # Define colormap
        # Choose cmap
        if self.current_cmap_name == "Custom":
            colors = [(0.0, 'blue'), (0.5, 'red'), (1.0, 'yellow')]
            cmap = mcolors.LinearSegmentedColormap.from_list('custom_RdYlBu', colors)
        else:
            cmap = plt.get_cmap(self.current_cmap_name)

        # x-axis always in seconds, UT formatting handled separately
        extent = [0, self.time[-1], self.freqs[-1], self.freqs[0]]

        # Prepare colorbar axis
        divider = make_axes_locatable(self.canvas.ax)
        cax = divider.append_axes("right", size="5%", pad=0.1)
        self.current_cax = cax

        # Show image
        # Convert units if needed
        if self.use_db:
            display_data = data * 2500.0 / 255.0 / 25.4
        else:
            display_data = data

        self.current_display_data = display_data

        im = self.canvas.ax.imshow(display_data, aspect='auto', extent=extent, cmap=cmap)

        self.current_colorbar = self.canvas.figure.colorbar(im, cax=cax)

        label = "Intensity [Digits]" if not self.use_db else "Intensity [dB]"
        self.current_colorbar.set_label(label)

        # Store label string so apply_graph_properties_live can re-style it
        self._colorbar_label_text = label

        self.canvas.ax.set_ylabel("Frequency [MHz]")
        self.canvas.ax.set_title(f"{self.filename} - {title}", fontsize=14)

        self.format_axes()  # Format x-axis based on user selection (seconds/UT)
        self._restore_view(view)

        # Apply graph properties (title/font/sizes) after plot rebuild
        self.apply_graph_properties_live()
        self.graph_group.setEnabled(True)

        self.canvas.draw()

        self.current_plot_type = title
        self._sync_toolbar_enabled_states()
        self.statusBar().showMessage(f"Loaded: {self.filename}", 5000)

    def on_mouse_motion_status(self, event):
        """Show time, frequency and intensity under cursor in status bar."""
        # Cursor not over axes or no data
        if event.inaxes != self.canvas.ax:
            # Cursor outside plot → show zeros
            self.cursor_label.setText("t = 0.00   |   f = 0.00 MHz   |   I = 0.00")
            return

        if self.current_display_data is None or self.time is None or self.freqs is None:
            return
        if event.xdata is None or event.ydata is None:
            return

        x = float(event.xdata)
        y = float(event.ydata)

        # Convert x to nearest time index
        time_arr = np.array(self.time)
        freq_arr = np.array(self.freqs)

        # Safety guard
        if time_arr.size == 0 or freq_arr.size == 0:
            return

        # Find nearest indices in time and frequency
        idx_x = int(np.argmin(np.abs(time_arr - x)))
        idx_y = int(np.argmin(np.abs(freq_arr - y)))

        ny, nx = self.current_display_data.shape
        if idx_x < 0 or idx_x >= nx or idx_y < 0 or idx_y >= ny:
            return

        t_val = time_arr[idx_x]
        f_val = freq_arr[idx_y]
        intensity = self.current_display_data[idx_y, idx_x]

        # Format time string in seconds or UT
        if self.use_utc and self.ut_start_sec is not None:
            total_seconds = self.ut_start_sec + t_val
            hours = int(total_seconds // 3600) % 24
            minutes = int((total_seconds % 3600) // 60)
            seconds = int(total_seconds % 60)
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d} UT"
        else:
            time_str = f"{t_val:.2f} s"

        unit_label = "dB" if self.use_db else "Digits"

        msg = (
            f"t = {time_str}   |   "
            f"f = {f_val:.2f} MHz   |   "
            f"I = {intensity:.2f} {unit_label}"
        )
        self.cursor_label.setText(msg)

    def change_cmap(self, name):
        self._push_undo_state()
        self.current_cmap_name = name
        if self.raw_data is None:
            return
        data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
        self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def get_current_cmap(self):
        if self.current_cmap_name == "Custom":
            colors = [(0.0, 'blue'), (0.5, 'red'), (1.0, 'yellow')]
            return mcolors.LinearSegmentedColormap.from_list('custom_RdYlBu', colors)
        else:
            return plt.get_cmap(self.current_cmap_name)

    def on_scroll_zoom(self, event):
        """Smooth zoom using mouse scroll wheel."""
        if self.lasso_active:
            return

        ax = self.canvas.ax

        # Mouse pointer must be inside the plot
        if event.inaxes != ax:
            return

        # Zoom factor
        base_scale = 1.15  # smooth and gentle zoom

        # Zoom IN
        if event.button == "up":
            scale_factor = 1 / base_scale
        # Zoom OUT
        elif event.button == "down":
            scale_factor = base_scale
        else:
            return

        # Current axis limits
        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()

        x_range = (x_max - x_min)
        y_range = (y_max - y_min)

        # Mouse location in data coords
        xdata = event.xdata
        ydata = event.ydata

        # Compute new ranges
        new_x_range = x_range * scale_factor
        new_y_range = y_range * scale_factor

        # Shift around mouse cursor
        x_shift = (xdata - x_min) * (1 - scale_factor)
        y_shift = (ydata - y_min) * (1 - scale_factor)

        new_x_min = x_min + x_shift
        new_x_max = new_x_min + new_x_range

        new_y_min = y_min + y_shift
        new_y_max = new_y_min + new_y_range

        # Apply limits
        ax.set_xlim(new_x_min, new_x_max)
        ax.set_ylim(new_y_min, new_y_max)

        self.canvas.draw_idle()

    def on_mouse_press(self, event):
        """
        Start panning with LEFT mouse button inside the main axes.
        (No modifier keys needed.)
        """
        if self.lasso_active:
            return

        # Only react if we click inside the image axes
        if event.inaxes != self.canvas.ax:
            return

        # Left button = start pan
        if event.button == 1 and event.xdata is not None and event.ydata is not None:
            self._panning = True
            self._last_pan_xy = (event.xdata, event.ydata)

    def on_mouse_move(self, event):
        """
        Perform the pan movement while the left mouse button is held.
        """
        if self.lasso_active:
            return

        # Only act if we are in panning mode and pointer is on data
        if not self._panning or event.xdata is None or event.ydata is None:
            return

        ax = self.canvas.ax

        # Previous mouse position in data coordinates
        x_prev, y_prev = self._last_pan_xy

        # Current mouse position
        x_curr, y_curr = event.xdata, event.ydata

        # Difference
        dx = x_prev - x_curr
        dy = y_prev - y_curr

        # Current limits
        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()

        # Shift both axes
        ax.set_xlim(x_min + dx, x_max + dx)
        ax.set_ylim(y_min + dy, y_max + dy)

        self.canvas.draw_idle()

        # Update last mouse position
        self._last_pan_xy = (x_curr, y_curr)

    def on_mouse_release(self, event):
        """
        Stop panning when the mouse button is released.
        """
        self._panning = False
        self._last_pan_xy = None


    def activate_drift_tool(self):
        self.statusBar().showMessage("Click multiple points along the burst. Right-click or double-click to finish.",
                                     8000)
        self.drift_points = []
        self.drift_click_cid = self.canvas.mpl_connect("button_press_event", self.on_drift_point_click)

    def on_drift_point_click(self, event):
        if not event.inaxes:
            return

        # Right-click or double-click to finish
        if event.button == 3 or event.dblclick:
            self.finish_drift_estimation()
            return

        self.drift_points.append((event.xdata, event.ydata))
        self.canvas.ax.plot(event.xdata, event.ydata, 'w*')
        self.canvas.draw()

    def finish_drift_estimation(self):
        self.canvas.mpl_disconnect(self.drift_click_cid)

        if len(self.drift_points) < 2:
            self.statusBar().showMessage("Need at least two points to estimate drift.", 4000)
            return

        drift_rates = []
        for i in range(len(self.drift_points) - 1):
            x1, y1 = self.drift_points[i]
            x2, y2 = self.drift_points[i + 1]
            drift = (y2 - y1) / (x2 - x1)
            drift_rates.append(drift)
            # Draw line between points
            self.canvas.ax.plot([x1, x2], [y1, y2], linestyle='--', color='lime')

        avg_drift = np.mean(drift_rates)
        self.canvas.ax.legend(["Drift Segments"])
        self.canvas.draw()

        self.statusBar().showMessage(
            f"Average Drift Rate: {avg_drift:.4f} MHz/s, Start Frequency: {y1: .3f}, End Frequency: {y2: .3f}, Duration: {x2 - x1: .3f} s",
            0)

    def activate_lasso(self):
        if self.noise_reduced_data is None:
            QMessageBox.warning(self, "Error", "Please apply noise reduction before isolating a burst.")
            return

        self.canvas.mpl_disconnect(self._cid_press)
        self.canvas.mpl_disconnect(self._cid_motion)
        self.canvas.mpl_disconnect(self._cid_release)

        self.lasso_active = True

        # Disconnect old lasso
        if self.lasso:
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None

        self.canvas.ax.set_title("Draw around the burst")
        self.canvas.draw()

        self.lasso = LassoSelector(self.canvas.ax, onselect=self.on_lasso_select)

    def on_lasso_select(self, verts):

        if self.noise_reduced_data is None:
            print("Lasso used before data was prepared. Ignoring.")
            return

        path = Path(verts)

        ny, nx = self.noise_reduced_data.shape
        y = np.linspace(self.freqs[0], self.freqs[-1], ny)
        x = np.linspace(0, self.time[-1], nx)
        X, Y = np.meshgrid(x, y)

        coords = np.column_stack((X.flatten(), Y.flatten()))
        mask = path.contains_points(coords).reshape(ny, nx)

        self.lasso_mask = mask  # store for use later

        # Safely disconnect the lasso tool
        if self.lasso:
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None

        self._cid_press = self.canvas.mpl_connect("button_press_event", self.on_mouse_press)
        self._cid_motion = self.canvas.mpl_connect("motion_notify_event", self.on_mouse_move)
        self._cid_release = self.canvas.mpl_connect("button_release_event", self.on_mouse_release)
        self.lasso_active = False
        # Defer drawing to avoid crash during event handling
        QTimer.singleShot(0, lambda: self._plot_isolated_burst(mask))

    def _plot_isolated_burst(self, mask):
        self._push_undo_state()
        # Build isolated data array
        burst_isolated = np.zeros_like(self.noise_reduced_data)
        burst_isolated[mask] = self.noise_reduced_data[mask]

        # Clear figure
        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        # Remove old colorbar safely
        try:
            if self.current_colorbar:
                self.current_colorbar.remove()
        except Exception:
            pass
        self.current_colorbar = None

        try:
            if self.current_cax:
                self.current_cax.remove()
        except Exception:
            pass
        self.current_cax = None

        # Restore extent
        extent = [0, self.time[-1], self.freqs[-1], self.freqs[0]]

        # Get SAME colormap user selected
        cmap = self.get_current_cmap()

        # === CREATE COLORBAR AXIS HERE (must exist before colorbar) ===
        divider = make_axes_locatable(self.canvas.ax)
        cax = divider.append_axes("right", size="5%", pad=0.1)
        self.current_cax = cax

        # Plot with fixed vmin/vmax from noise reduction
        im = self.canvas.ax.imshow(
            burst_isolated,
            aspect='auto',
            extent=extent,
            cmap=cmap,
            vmin=self.noise_vmin,
            vmax=self.noise_vmax,
        )

        self.current_display_data = burst_isolated

        # Create new colorbar
        self.current_colorbar = self.canvas.figure.colorbar(im, cax=cax)
        self._colorbar_label_text = "Intensity [Digits]" if not self.use_db else "Intensity [dB]"
        self.current_colorbar.set_label(self._colorbar_label_text)

        # Labels
        self.canvas.ax.set_title("Isolated Burst")
        self.canvas.ax.set_ylabel("Frequency [MHz]")
        self.format_axes()

        self.canvas.draw()

        # Replace display data with isolated data
        self.noise_reduced_data = burst_isolated

        self.statusBar().showMessage("Burst isolated using lasso", 4000)

    def plot_max_intensities(self):
        # Ensure any active lasso from the main plot is fully disconnected
        if getattr(self, "lasso", None):
            try:
                self.lasso.disconnect_events()
            except Exception:
                pass
            self.lasso = None

        if self.noise_reduced_data is None:
            print("No burst-isolated data available.")
            return

        # Diagnostics
        print("🖥️ Screens:", QGuiApplication.screens())
        print("🎯 Creating MaxIntensityPlotDialog...")

        try:
            data = self.noise_reduced_data
            ny, nx = data.shape
            time_channel_number = np.linspace(0, nx, nx)
            max_intensity_freqs = self.freqs[np.argmax(data, axis=0)]

            # Safely create the dialog
            dialog = MaxIntensityPlotDialog(time_channel_number, max_intensity_freqs, self.filename, self)
            dialog.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)

            # Connect to GC after close
            dialog.finished.connect(lambda: gc.collect())

            dialog.exec()
            gc.collect()

        except Exception as e:
            print(f"❌ Error showing MaxIntensityPlotDialog: {e}")

    def export_figure(self):

        if not self.filename:
            QMessageBox.warning(self, "No File Loaded", "Load a FITS file before exporting.")
            return

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        base_name = self.filename.split(".")[0]
        suffix = self.current_plot_type.replace(" ", "")
        default_name = f"{base_name}_{suffix}"

        file_path, ext = pick_export_path(
            self,
            "Export Figure",
            default_name,
            formats,
            default_filter="PNG (*.png)"
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        def normalize_ext(ext_value: str) -> str:
            """
            Accepts values like: 'png', '.png', 'PNG (*.png)'
            Returns: 'png'
            """
            if not ext_value:
                return "png"
            s = str(ext_value).strip().lower()
            if s.startswith("."):
                s = s[1:]
            m = re.search(r"\*\.(\w+)", s)
            if m:
                return m.group(1).lower()
            return s

        try:
            root, current_ext = os.path.splitext(file_path)

            # If user did not type an extension, add one based on returned ext
            if current_ext == "":
                ext_final = normalize_ext(ext)
                file_path = f"{file_path}.{ext_final}"
            else:
                ext_final = current_ext.lower().lstrip(".")

            self.canvas.figure.savefig(
                file_path,
                dpi=300,
                bbox_inches="tight",
                format=ext_final
            )

            QMessageBox.information(self, "Export Complete", f"Figure saved:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"An error occurred:\n{e}")

    def reset_all(self):
        # Safely remove colorbar
        try:
            if self.current_colorbar and self.current_colorbar.ax:
                self.current_colorbar.remove()
        except Exception:
            pass
        self.current_colorbar = None

        try:
            if self.current_cax:
                self.current_cax.remove()
        except Exception:
            pass
        self.current_cax = None

        # Clear canvas
        self.canvas.ax.clear()
        self.canvas.draw()

        # Clear data
        self.raw_data = None
        self.freqs = None
        self.time = None
        self.filename = ""
        self.noise_reduced_data = None
        self.noise_reduced_original = None
        self.lasso_mask = None
        self.current_plot_type = "Raw"

        # Reset GUI
        self.statusBar().showMessage("All reset", 4000)

        # Tool bar
        self._sync_toolbar_enabled_states()
        self.graph_group.setEnabled(False)

        if self.canvas.ax:
            self.canvas.ax.set_xlim(0, 1)
            self.canvas.ax.set_ylim(1, 0)

        print("Application reset to initial state.")

    def show_about_dialog(self):
        QMessageBox.information(
            self,
            "About e-Callisto FITS Analyzer",
            "e-CALLISTO FITS Analyzer version 1.7.4.\n\n"
            "Developed by Sahan S Liyanage\n\n"
            "Astronomical and Space Science Unit\n"
            "University of Colombo, Sri Lanka\n\n"
            "2025©Copyright, All Rights Reserved."
        )

    def reset_selection(self):
        self._push_undo_state()
        if self.noise_reduced_original is not None:
            self.noise_reduced_data = self.noise_reduced_original.copy()
            if self.time is not None and self.freqs is not None:
                self.plot_data(self.noise_reduced_data, title="Noise Reduced")
            self.lasso_mask = None
            self.lasso = None
            self.statusBar().showMessage("Selection Reset", 4000)
            print("Lasso selection reset. Original noise-reduced data restored.")
        else:
            print("No noise-reduced backup found. Reset skipped.")

    def open_combine_freq_window(self):
        dialog = CombineFrequencyDialog(self)
        dialog.exec()

    def open_combine_time_window(self):
        dialog = CombineTimeDialog(self)
        dialog.exec()

    def set_axis_to_seconds(self):
        self._push_undo_state()
        self.use_utc = False
        self.xaxis_sec_action.setChecked(True)
        self.xaxis_ut_action.setChecked(False)

        if self.raw_data is not None:
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def set_axis_to_utc(self):
        self._push_undo_state()
        self.use_utc = True
        self.xaxis_sec_action.setChecked(False)
        self.xaxis_ut_action.setChecked(True)

        if self.raw_data is not None:
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, keep_view=True)

    def format_axes(self):
        if self.use_utc and self.ut_start_sec is not None:
            def format_func(x, pos):
                total_seconds = self.ut_start_sec + x
                hours = int(total_seconds // 3600) % 24
                minutes = int((total_seconds % 3600) // 60)
                seconds = int(total_seconds % 60)
                return f"{hours:02d}:{minutes:02d}"

            self.canvas.ax.xaxis.set_major_formatter(FuncFormatter(format_func))
            self.canvas.ax.set_xlabel("Time [UT]")
        else:
            self.canvas.ax.xaxis.set_major_formatter(ScalarFormatter())
            self.canvas.ax.set_xlabel("Time [s]")

        self.canvas.ax.figure.canvas.draw()

    def process_imported_files(self, urls):
        if not urls:
            QMessageBox.warning(self, "No Files", "No files were received from the downloader.")
            return

        local_files = []

        try:
            for url in urls:
                r = requests.get(url, timeout=20)
                r.raise_for_status()

                original_name = url.split("/")[-1]

                temp_dir = tempfile.gettempdir()
                local_path = os.path.join(temp_dir, original_name)

                with open(local_path, "wb") as f:
                    f.write(r.content)

                local_files.append(local_path)

        except Exception as e:
            QMessageBox.critical(self, "Download Error",
                                 f"Failed to download one or more FITS files:\n{e}")
            return

        if len(local_files) == 1:
            self.load_fits_into_main(local_files[0])
            self.downloader_dialog.import_success.emit()
            return

        from burst_processor import (
            are_time_combinable,
            are_frequency_combinable,
            combine_time,
            combine_frequency,
        )

        try:
            if are_time_combinable(local_files):
                combined = combine_time(local_files)
                self.load_combined_into_main(combined)
                self.downloader_dialog.import_success.emit()
                return

            if are_frequency_combinable(local_files):
                combined = combine_frequency(local_files)
                self.load_combined_into_main(combined)
                self.downloader_dialog.import_success.emit()
                return

            # Extract UT from first downloaded file
            try:
                hdul = fits.open(local_files[0])
                hdr = hdul[0].header
                hh, mm, ss = hdr["TIME-OBS"].split(":")
                self.ut_start_sec = int(hh) * 3600 + int(mm) * 60 + float(ss)
                hdul.close()
            except Exception:
                self.ut_start_sec = None


        except Exception as e:
            QMessageBox.critical(self, "Combine Error", f"An error occurred while combining files:\n{e}")
            return

        QMessageBox.warning(
            self,
            "Invalid Selection",
            "Selected files cannot be time-combined or frequency-combined.\n"
            "Please ensure they are consecutive in time or adjacent in frequency."
        )

    def launch_downloader(self):
        self.downloader_dialog = CallistoDownloaderApp()
        self.downloader_dialog.import_request.connect(self.process_imported_files)

        self.import_success_signal = lambda: self.downloader_dialog.accept()
        self.downloader_dialog.import_success.connect(self.import_success_signal)

        self.downloader_dialog.exec()

    def open_goes_xrs_window(self):
        self.goes_window = GoesXrsWindow()
        self.goes_window.show()

    def open_soho_lasco_window(self):
        self.open_cme_viewer()

    def _capture_state(self):
        """Capture the current application state for Undo/Redo."""
        state = {
            "raw_data": None if self.raw_data is None else self.raw_data.copy(),
            "noise_reduced_data": None if self.noise_reduced_data is None else self.noise_reduced_data.copy(),
            "noise_reduced_original": None if self.noise_reduced_original is None else self.noise_reduced_original.copy(),
            "lasso_mask": None if self.lasso_mask is None else self.lasso_mask.copy(),
            "freqs": None if self.freqs is None else self.freqs.copy(),
            "time": None if self.time is None else self.time.copy(),
            "filename": self.filename,
            "current_plot_type": self.current_plot_type,
            "lower_slider": self.lower_slider.value(),
            "upper_slider": self.upper_slider.value(),
            "use_db": self.use_db,
            "use_utc": self.use_utc,
            "cmap": self.current_cmap_name,
            "view": self._capture_view(),
        }
        return state

    def _push_undo_state(self):
        self._undo_stack.append(self._capture_state())
        if len(self._undo_stack) > self._max_undo:
            self._undo_stack.pop(0)
        self._redo_stack.clear()

    def _restore_state(self, state):
        """Restore a previously captured application state."""
        self.raw_data = state["raw_data"]
        self.noise_reduced_data = state["noise_reduced_data"]
        self.noise_reduced_original = state["noise_reduced_original"]
        self.lasso_mask = state["lasso_mask"]
        self.freqs = state["freqs"]
        self.time = state["time"]
        self.filename = state["filename"]
        self.current_plot_type = state["current_plot_type"]
        self.use_db = state["use_db"]
        self.use_utc = state["use_utc"]
        self.current_cmap_name = state["cmap"]

        self.lower_slider.blockSignals(True)
        self.upper_slider.blockSignals(True)
        self.lower_slider.setValue(state["lower_slider"])
        self.upper_slider.setValue(state["upper_slider"])
        self.lower_slider.blockSignals(False)
        self.upper_slider.blockSignals(False)

        if self.raw_data is not None:
            data = self.noise_reduced_data if self.noise_reduced_data is not None else self.raw_data
            self.plot_data(data, title=self.current_plot_type, keep_view=False)
            self._restore_view(state["view"])

    def undo(self):
        if not self._undo_stack:
            self.statusBar().showMessage("Nothing to undo", 2000)
            return

        current = self._capture_state()
        self._redo_stack.append(current)

        state = self._undo_stack.pop()
        self._restore_state(state)
        self.statusBar().showMessage("Undo", 2000)

    def redo(self):
        if not self._redo_stack:
            self.statusBar().showMessage("Nothing to redo", 2000)
            return

        current = self._capture_state()
        self._undo_stack.append(current)

        state = self._redo_stack.pop()
        self._restore_state(state)
        self.statusBar().showMessage("Redo", 2000)

    def open_cme_viewer(self):
        from soho_lasco_viewer import CMEViewer  # import here, not at top
        self._cme_viewer = CMEViewer(parent=self)
        self._cme_viewer.show()


class MaxIntensityPlotDialog(QDialog):
    def __init__(self, time_channels, max_freqs, filename, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Maximum Intensities for Each Time Channel")
        self.resize(1000, 700)
        self.filename = filename
        self.current_plot_type = "MaxIntensityPlot"

        # Data
        self.time_channels = np.array(time_channels)
        self.freqs = np.array(max_freqs)
        self.selected_mask = np.zeros_like(self.time_channels, dtype=bool)
        self.lasso = None

        # Canvas
        self.canvas = MplCanvas(self, width=10, height=6)
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)
        self.canvas.ax.scatter(self.time_channels, self.freqs, marker="o", s=5, color='red')
        self.canvas.ax.set_xlabel("Time Channel Number")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.set_title("Maximum Intensity for Each Time Channel")
        self.canvas.draw()

        # Buttons
        self.select_button = QPushButton("Select Outliers")
        self.remove_button = QPushButton("Remove Outliers")

        self.fundamental_radio = QRadioButton("Fundamental")
        self.harmonic_radio = QRadioButton("Harmonic")
        self.fundamental_radio.setChecked(True)

        self.mode_group = QButtonGroup(self)
        self.mode_group.addButton(self.fundamental_radio)
        self.mode_group.addButton(self.harmonic_radio)

        self.analyze_button = QPushButton("Analyze Burst")
        self.select_button.setToolTip("Use Lasso tool to select points to remove")
        self.remove_button.setToolTip("Remove previously selected outliers")
        self.select_button.setMinimumWidth(150)
        self.remove_button.setMinimumWidth(150)
        self.analyze_button.setMinimumWidth(150)
        self.select_button.clicked.connect(self.activate_lasso)
        self.remove_button.clicked.connect(self.remove_selected_outliers)

        self.analyze_button.clicked.connect(lambda: self.open_analyze_window(
            fundamental=self.fundamental_radio.isChecked(),
            harmonic=self.harmonic_radio.isChecked()
        ))

        # Layouts
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.select_button)
        button_layout.addWidget(self.remove_button)
        button_layout.addWidget(self.fundamental_radio)
        button_layout.addWidget(self.harmonic_radio)
        button_layout.addWidget(self.analyze_button)
        button_layout.addStretch()

        # Status bar
        self.status = QStatusBar()
        self.status.showMessage("Ready")

        # Menubar
        menubar = QMenuBar(self)
        file_menu = menubar.addMenu("File")
        self.save_action = QAction("Save As", self)
        self.export_action = QAction("Export As", self)
        file_menu.addAction(self.save_action)
        file_menu.addAction(self.export_action)
        self.save_action.triggered.connect(self.save_as_csv)
        self.export_action.triggered.connect(self.export_figure)

        edit_menu = menubar.addMenu("Edit")
        reset_action = QAction("Reset All", self)
        edit_menu.addAction(reset_action)
        reset_action.triggered.connect(self.reset_all)

        analyze_menu = menubar.addMenu("Analyze")
        analyze_action = QAction("Open Analyzer", self)
        analyze_menu.addAction(analyze_action)
        analyze_action.triggered.connect(self.open_analyze_window)

        about_menu = menubar.addMenu("About")
        about_action = QAction("About", self)
        about_action.setMenuRole(QAction.NoRole)
        about_menu.addAction(about_action)
        about_action.triggered.connect(self.show_about_dialog)

        # Main Layout
        layout = QVBoxLayout()
        layout.setMenuBar(menubar)
        layout.addLayout(button_layout)
        layout.addWidget(self.canvas)
        layout.addWidget(self.status)
        self.setLayout(layout)

        # Styling
        self.setStyleSheet("""
            QPushButton { font-size: 13px; padding: 6px 12px; }
            QLabel { font-size: 13px; }
        """)

    def activate_lasso(self):
        self.canvas.ax.set_title("Draw around outliers to remove")
        self.canvas.draw()

        if self.lasso:
            self.lasso.disconnect_events()

        self.lasso = LassoSelector(self.canvas.ax, onselect=self.on_lasso_select)

    def on_lasso_select(self, verts):
        path = Path(verts)
        points = np.column_stack((self.time_channels, self.freqs))
        self.selected_mask = path.contains_points(points)
        if self.lasso:
            self.lasso.disconnect_events()
            self.lasso = None
        self.status.showMessage(f"{np.sum(self.selected_mask)} points selected", 3000)

    def remove_selected_outliers(self):
        if not np.any(self.selected_mask):
            self.status.showMessage("No points selected for removal", 3000)
            return

        self.time_channels = self.time_channels[~self.selected_mask]
        self.freqs = self.freqs[~self.selected_mask]
        self.selected_mask = np.zeros_like(self.time_channels, dtype=bool)

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        self.canvas.ax.scatter(self.time_channels, self.freqs, marker="o", s=5, color='red')
        self.canvas.ax.set_xlabel("Time Channel Number")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.set_title("Filtered Max Intensities")
        self.canvas.draw()

        self.status.showMessage("Selected outliers removed", 3000)

    def reset_all(self):
        # Clear canvas
        self.canvas.ax.clear()
        self.canvas.draw()

        # Clear internal variables
        self.raw_data = None
        self.freqs = None
        self.time = None
        self.filename = ""
        self.noise_reduced_data = None
        self.lasso_mask = None
        self.current_plot_type = "Raw"

        self.statusBar().showMessage("All reset", 4000)

        print("Application reset to initial state.")

    def save_as_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save CSV File", "", "CSV files (*.csv)")
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Time Channel", "Frequency (MHz)"])
                for t, fval in zip(self.time_channels * 0.25, self.freqs):
                    writer.writerow([t, fval])
            self.status.showMessage(f"Saved to {file_path}", 3000)
        except Exception as e:
            self.status.showMessage(f"Error: {e}", 3000)

    def export_figure(self):

        if not self.filename:
            QMessageBox.warning(self, "No File Loaded", "Load a FITS file before exporting.")
            return

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        base_name = self.filename.split(".")[0]
        suffix = self.current_plot_type.replace(" ", "")
        default_name = f"{base_name}_{suffix}"

        file_path, ext = pick_export_path(
            self,
            "Export Figure",
            default_name,
            formats,
            default_filter="PNG (*.png)"
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        try:
            # ✅ If user didn't type an extension, add the one from ext
            root, current_ext = os.path.splitext(file_path)
            if current_ext == "":
                ext = ext.lower().lstrip(".")  # ext should be like "png"
                file_path = f"{file_path}.{ext}"
            else:
                ext = current_ext.lower().lstrip(".")  # use what user typed

            self.canvas.figure.savefig(
                file_path,
                dpi=300,
                bbox_inches="tight",
                format=ext
            )

            QMessageBox.information(self, "Export Complete", f"Figure saved:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"An error occurred:\n{e}")

    def show_about_dialog(self):
        QMessageBox.information(
            self,
            "About e-Callisto FITS Analyzer",
            "e-CALLISTO FITS Analyzer version 1.7.4.\n\n"
            "Developed by Sahan S Liyanage\n\n"
            "Astronomical and Space Science Unit\n"
            "University of Colombo, Sri Lanka\n\n"
            "2025©Copyright, All Rights Reserved."
        )

    def open_analyze_window(self, fundamental=True, harmonic=False):
        dialog = AnalyzeDialog(self.time_channels, self.freqs, self.filename, fundamental=fundamental,
                               harmonic=harmonic, parent=self)
        dialog.exec()

    def closeEvent(self, event):
        try:
            if hasattr(self.canvas, 'ax'):
                self.canvas.ax.clear()
            self.canvas.figure.clf()
            self.canvas.deleteLater()

            if self.lasso:
                self.lasso.disconnect_events()
                self.lasso = None
        except Exception as e:
            print(f"Cleanup error: {e}")
        event.accept()

from PySide6.QtWidgets import (
    QDialog, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QFileDialog, QComboBox, QScrollArea, QWidget, QSizePolicy
)

from matplotlib.figure import Figure
import numpy as np
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score, mean_squared_error



class AnalyzeDialog(QDialog):
    def __init__(self, time_channels, freqs, filename, fundamental=True, harmonic=False, parent=None):
        super().__init__(parent)
        self.fundamental = fundamental
        self.harmonic = harmonic

        self.setWindowTitle("Analyzer")
        self.resize(1100, 700)

        self.time = np.array(time_channels) * 0.25
        self.freq = np.array(freqs)
        self.filename = filename.split(".")[0]
        self.current_plot_title = f"{self.filename}_Best_Fit"

        # Canvas
        self.canvas = MplCanvas(self, width=8, height=5)
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        # Buttons
        self.max_button = QPushButton("Maximum Intensities")
        self.fit_button = QPushButton("Best Fit")
        self.save_plot_button = QPushButton("Save Graph")
        self.save_data_button = QPushButton("Save Data")
        self.existing_excel_checkbox = QCheckBox("Existing Excel File")

        self.extra_plot_label = QLabel("Extra Plots:")
        self.extra_plot_combo = QComboBox()
        self.extra_plot_combo.addItems([
            "Shock Speed vs Shock Height",
            "Shock Speed vs Frequency",
            "Shock Height vs Frequency"
        ])
        self.extra_plot_button = QPushButton("Plot")

        self.max_button.clicked.connect(self.plot_max)
        self.fit_button.clicked.connect(self.plot_fit)
        self.save_plot_button.clicked.connect(self.save_graph)
        self.save_data_button.clicked.connect(self.save_data)
        self.extra_plot_button.clicked.connect(self.plot_extra)

        # Plot control layout
        plot_button_layout = QHBoxLayout()
        plot_button_layout.addWidget(self.max_button)
        plot_button_layout.addWidget(self.fit_button)

        left_layout = QVBoxLayout()
        left_layout.addLayout(plot_button_layout)
        left_layout.addWidget(self.canvas)

        # === Info Panel ===
        self.equation_label = QLabel("Best Fit Equation:")
        self.equation_display = QLabel("")
        self.equation_display.setTextFormat(Qt.RichText)
        self.equation_display.setStyleSheet("font-size: 16px; padding: 4px;")

        self.stats_header = QLabel("<b>Fit Metrics:</b>")
        self.r2_display = QLabel("R² = ")
        self.rmse_display = QLabel("RMSE = ")

        self.shock_header = QLabel("<b>Shock Parameters:</b>")
        self.avg_freq_display = QLabel("")
        self.drift_display = QLabel("")
        self.start_freq_display = QLabel("")
        self.initial_shock_speed_display = QLabel("")
        self.initial_shock_height_display = QLabel("")
        self.avg_shock_speed_display = QLabel("")
        self.avg_shock_height_display = QLabel("")

        self.labels = [
            self.equation_label, self.equation_display,
            self.stats_header, self.r2_display, self.rmse_display,
            self.shock_header,
            self.avg_freq_display, self.drift_display, self.start_freq_display,
            self.initial_shock_speed_display, self.initial_shock_height_display,
            self.avg_shock_speed_display, self.avg_shock_height_display,
            self.save_plot_button, self.save_data_button, self.existing_excel_checkbox,
            self.extra_plot_label, self.extra_plot_combo, self.extra_plot_button
        ]

        right_inner = QVBoxLayout()
        for widget in self.labels:
            right_inner.addWidget(widget)
        right_inner.addStretch()

        right_widget = QWidget()
        right_widget.setLayout(right_inner)
        right_scroll = QScrollArea()
        right_scroll.setWidget(right_widget)
        right_scroll.setWidgetResizable(True)
        right_scroll.setMinimumWidth(300)

        self.status = QStatusBar()
        self.status.showMessage("Ready")

        # Main layout
        main_layout = QHBoxLayout()
        main_layout.addLayout(left_layout, stretch=3)
        main_layout.addWidget(right_scroll, stretch=1)
        main_with_status = QVBoxLayout()
        main_with_status.addLayout(main_layout)
        main_with_status.addWidget(self.status)
        self.setLayout(main_with_status)

        # Styling
        self.setStyleSheet("""
            QLabel {
                font-size: 13px;
                padding: 2px;
            }
            QLabel#value {
                font-weight: bold;
            }
            QPushButton {
                padding: 6px;
            }
        """)

    def plot_max(self):
        self.canvas.ax.clear()
        self.canvas.ax.scatter(self.time, self.freq, s=10, color='blue')
        self.canvas.ax.set_title(f"{self.filename}_Maximum_Intensity")
        self.canvas.ax.set_xlabel("Time (s)")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.grid(True)
        self.canvas.draw()
        self.equation_display.setText("")
        self.status.showMessage("Max intensities plotted successfully!", 3000)

    def plot_fit(self):
        def model_func(t, a, b): return a * t ** (b)

        def drift_rate(t, a_, b_): return a_ * b_ * t ** (b_ - 1)

        params, cov = curve_fit(model_func, self.time, self.freq, maxfev=10000)
        a, b = params
        std_errs = np.sqrt(np.diag(cov))

        time_fit = np.linspace(self.time.min(), self.time.max(), 400)
        freq_fit = model_func(time_fit, a, b)

        self.canvas.ax.clear()
        self.canvas.figure.clf()
        self.canvas.ax = self.canvas.figure.add_subplot(111)

        self.canvas.ax.scatter(self.time, self.freq, s=10, color='blue', label="Original Data")
        self.canvas.ax.plot(time_fit, freq_fit, color='red', label=fr"Best Fit: $f = {a:.2f} \cdot t^{{{b:.2f}}}$")
        self.canvas.ax.set_title(f"{self.filename}_Best_Fit")
        self.canvas.ax.set_xlabel("Time (s)")
        self.canvas.ax.set_ylabel("Frequency (MHz)")
        self.canvas.ax.legend()
        self.canvas.ax.grid(True)
        self.canvas.draw()
        self.current_plot_title = f"{self.filename}_Best_Fit"

        predicted = model_func(self.time, a, b)
        r2 = r2_score(self.freq, predicted)
        rmse = np.sqrt(mean_squared_error(self.freq, predicted))

        self.equation_display.setText(f"<b>f(t) = {a:.2f} · t<sup>{b:.2f}</sup></b>")
        self.r2_display.setText(f"R² = {r2:.4f}")
        self.rmse_display.setText(f"RMSE = {rmse:.4f}")

        drift_vals = drift_rate(self.time, a, b)
        residuals = self.freq - predicted
        freq_err = np.std(residuals)
        drift_errs = np.abs(drift_vals) * np.sqrt((std_errs[0] / a) ** 2 + (std_errs[1] / b) ** 2)

        shock_speed = (13853221.38 * np.abs(drift_vals)) / (self.freq * (np.log(self.freq ** 2 / 3.385)) ** 2)
        R_p = 4.32 * np.log(10) / np.log(self.freq ** 2 / 3.385)

        percentile = 90
        start_freq = np.percentile(self.freq, percentile)

        if self.harmonic:
            start_freq = start_freq / 2

        idx = np.abs(self.freq - start_freq).argmin()
        f0 = self.freq[idx]
        start_shock_speed = shock_speed[idx]
        start_height = R_p[idx]
        drift0 = drift_vals[idx]
        drift_err0 = drift_errs[idx]

        shock_speed_err = (13853221.38 * drift_err0) / (f0 * (np.log(f0 ** 2 / 3.385)) ** 2)
        dRp_df = (8.64 / f0) / np.log(10) / np.log(f0 ** 2 / 3.385)
        Rp_err = np.abs(dRp_df * freq_err)

        avg_freq = np.mean(self.freq)
        avg_freq_err = np.std(self.freq) / np.sqrt(len(self.freq))
        avg_drift = np.mean(drift_vals)
        avg_drift_err = np.std(drift_vals) / np.sqrt(len(drift_vals))
        avg_speed = np.mean(shock_speed)
        avg_speed_err = np.std(shock_speed) / np.sqrt(len(shock_speed))
        avg_height = np.mean(R_p)
        avg_height_err = np.std(R_p) / np.sqrt(len(R_p))

        self.shock_speed = shock_speed
        self.R_p = R_p
        self.freq_err = freq_err
        self.start_freq = start_freq
        self.start_height = start_height

        self.status.showMessage("Best fit plotted successfully!", 3000)

        # Display values
        self.avg_freq_display.setText(f"Average Frequency: <b>{avg_freq:.2f} ± {avg_freq_err:.2f}</b> MHz")
        self.drift_display.setText(f"Average Drift Rate: <b>{avg_drift:.4f} ± {avg_drift_err:.4f}</b> MHz/s")
        self.start_freq_display.setText(f"Starting Frequency: <b>{start_freq:.2f} ± {freq_err:.2f}</b> MHz")
        self.initial_shock_speed_display.setText(
            f"Initial Shock Speed: <b>{start_shock_speed:.2f} ± {shock_speed_err:.2f}</b> km/s")
        self.initial_shock_height_display.setText(f"Initial Shock Height: <b>{start_height:.3f} ± {Rp_err:.3f}</b> Rₛ")
        self.avg_shock_speed_display.setText(f"Average Shock Speed: <b>{avg_speed:.2f} ± {avg_speed_err:.2f}</b> km/s")
        self.avg_shock_height_display.setText(
            f"Average Shock Height: <b>{avg_height:.3f} ± {avg_height_err:.3f}</b> Rₛ")

    def save_graph(self):
        plot_name = getattr(self, "current_plot_title", None) or f"{self.filename}_Plot"

        formats = "PNG (*.png);;PDF (*.pdf);;EPS (*.eps);;SVG (*.svg);;TIFF (*.tiff)"

        file_path, ext = pick_export_path(
            self,
            "Export Figure",
            plot_name,
            formats,
            default_filter="PNG (*.png)"
        )

        if not file_path:
            return

        if sys.platform.startswith("win") and file_path.lower().startswith("c:\\program files"):
            QMessageBox.warning(
                self,
                "Permission Denied",
                "Windows does not allow saving files inside Program Files.\n"
                "Please choose another folder such as Documents or Desktop."
            )
            return

        try:
            # ✅ If user didn't type an extension, add the one from ext
            root, current_ext = os.path.splitext(file_path)
            if current_ext == "":
                ext = ext.lower().lstrip(".")
                file_path = f"{file_path}.{ext}"
            else:
                ext = current_ext.lower().lstrip(".")

            self.canvas.figure.savefig(
                file_path,
                dpi=300,
                bbox_inches="tight",
                format=ext
            )
            QMessageBox.information(self, "Export Complete", f"Plot saved:\n{file_path}")
            self.status.showMessage("Export successful!", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not save file:\n{e}")
            self.status.showMessage("Export failed!", 3000)

    def save_data(self):

        # All known e-Callisto station names
        station_list = [
            'ALASKA-ANCHORAGE', 'ALASKA-COHOE', 'ALASKA-HAARP', 'ALGERIA-CRAAG', 'ALMATY',
            'Arecibo-observatory', 'AUSTRIA-Krumbach', 'AUSTRIA-MICHELBACH', 'AUSTRIA-OE3FLB',
            'AUSTRIA-UNIGRAZ', 'Australia-ASSA', 'BRAZIL', 'BIR', 'Croatia-Visnjan', 'DENMARK',
            'EGYPT-Alexandria', 'EGYPT-SpaceAgency', 'ETHIOPIA', 'FINLAND-Siuntio', 'FINLAND-Kempele',
            'GERMANY-ESSEN', 'GERMANY-DLR', 'GLASGOW', 'GREENLAND', 'HUMAIN', 'HURBANOVO',
            'INDIA-GAURI', 'INDIA-Nashik', 'INDIA-OOTY', 'INDIA-UDAIPUR', 'INDONESIA',
            'ITALY-Strassolt', 'JAPAN-IBARAKI', 'KASI', 'KRIM', 'MEXART',
            'MEXICO-ENSENADA-UNAM', 'MEXICO-FCFM-UANL', 'MEXICO-FCFM-UNACH', 'MEXICO-LANCE-A',
            'MEXICO-LANCE-B', 'MEXICO-UANL-INFIERNILLO', 'MONGOLIA-UB', 'MRO', 'MRT1', 'MRT3',
            'Malaysia_Banting', 'NASA-GSFC', 'NORWAY-EGERSUND', 'NORWAY-NY-AALESUND', 'NORWAY-RANDABERG',
            'PARAGUAY', 'POLAND-BALDY', 'POLAND-Grotniki', 'ROMANIA', 'ROSWELL-NM', 'RWANDA',
            'SOUTHAFRICA-SANSA', 'SPAIN-ALCALA', 'SPAIN-PERALEJOS', 'SPAIN-SIGUENZA', 'SRI-Lanka',
            'SSRT', 'SWISS-CalU', 'SWISS-FM', 'SWISS-HB9SCT', 'SWISS-HEITERSWIL', 'SWISS-IRSOL',
            'SWISS-Landschlacht', 'SWISS-MUHEN', 'TAIWAN-NCU', 'THAILAND-Pathumthani', 'TRIEST',
            'TURKEY', 'UNAM', 'URUGUAY', 'USA-ARIZONA-ERAU', 'USA-BOSTON', 'UZBEKISTAN'
        ]

        # ✅ Extract Station
        station = "UNKNOWN"
        filename_lower = self.filename.lower()
        for s in station_list:
            if filename_lower.startswith(s.lower()):
                station = s
                break

        # ✅ Extract Date
        date_match = re.search(r'_(\d{4})(\d{2})(\d{2})_', self.filename)
        if date_match:
            date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
        else:
            date = "UNKNOWN"

        # ✅ Excel File Handling
        if self.existing_excel_checkbox.isChecked():
            path, _ = QFileDialog.getOpenFileName(self, "Select Existing Excel File", "", "Excel Files (*.xlsx)")
            if not path:
                return
            try:
                wb = load_workbook(path)
                ws = wb.active
            except Exception as e:
                QMessageBox.critical(self, "Load Error", f"Could not open Excel file:\n{str(e)}")
                return
        else:
            path, _ = QFileDialog.getSaveFileName(self, "Save as Excel", f"{self.filename}_data.xlsx",
                                                  "Excel Files (*.xlsx)")
            if not path:
                return
            try:
                wb = Workbook()
                ws = wb.active
                headers = [
                    "Date", "Station", "Best_fit", "R_sq", "RMSE",
                    "avg_freq", "avg_freq_err", "Avg_drift", "avg_drift_err",
                    "start_freq", "start_freq_err", "initial_shock_speed", "initial_shock_speed_err",
                    "initial_shock_height", "initial_shock_height_err", "avg_shock_speed", "avg_shock_speed_err",
                    "avg_shock_height", "avg_shock_height_err", "avg_drift_abs"
                ]
                ws.append(headers)
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Could not create Excel file:\n{str(e)}")
                return

        # ✅ Extract and clean text
        def extract_val_err(label):
            # Remove HTML tags
            clean_text = re.sub(r'<[^>]+>', '', label.text())
            # Remove units and stray characters
            clean_text = re.sub(r'(MHz|km/s|Rₛ|s|/)', '', clean_text)
            # Clean spaces
            clean_text = clean_text.strip()
            # Extract value ± error
            value_text = clean_text.split(":")[-1].strip()
            if "±" in value_text:
                value, err = value_text.split("±")
                return value.strip(), err.strip()
            else:
                return value_text.strip(), ""

        # ✅ Read values
        try:
            best_fit = re.sub(r'<[^>]+>', '', self.equation_display.text()).replace("<sup>", "^").replace("</sup>", "")
            r2 = self.r2_display.text().split("=")[-1].strip()
            rmse = self.rmse_display.text().split("=")[-1].strip()

            avg_freq, avg_freq_err = extract_val_err(self.avg_freq_display)
            avg_drift, avg_drift_err = extract_val_err(self.drift_display)

            try:
                avg_drift_abs = abs(float(avg_drift))
            except ValueError:
                avg_drift_abs = ""

            start_freq, start_freq_err = extract_val_err(self.start_freq_display)
            init_speed, init_speed_err = extract_val_err(self.initial_shock_speed_display)
            init_height, init_height_err = extract_val_err(self.initial_shock_height_display)
            avg_speed, avg_speed_err = extract_val_err(self.avg_shock_speed_display)
            avg_height, avg_height_err = extract_val_err(self.avg_shock_height_display)

            row = [
                date, station, best_fit, r2, rmse,
                avg_freq, avg_freq_err, avg_drift, avg_drift_err,
                start_freq, start_freq_err, init_speed, init_speed_err,
                init_height, init_height_err, avg_speed, avg_speed_err,
                avg_height, avg_height_err, avg_drift_abs
            ]

            ws.append(row)
            wb.save(path)
            self.status.showMessage("✅ Data saved to Excel successfully!", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Write Error", f"Could not write to Excel file:\n{str(e)}")
            self.status.showMessage("❌ Failed to save data to Excel.", 3000)

    def plot_extra(self):
        choice = self.extra_plot_combo.currentText()
        self.canvas.ax.clear()
        if choice == "Shock Speed vs Shock Height":
            self.canvas.ax.scatter(self.R_p, self.shock_speed, color='green', s=10)
            self.canvas.ax.set_xlabel("Shock Height (Rₛ)")
            self.canvas.ax.set_ylabel("Shock Speed (km/s)")
            self.canvas.ax.set_title(f"{self.filename}_Shock_Speed_vs_Shock_Height")
            self.current_plot_title = f"{self.filename}_Shock_Speed_vs_Shock_Height"
            self.status.showMessage("Shock Speed vs Shock Height plotted successfully!", 3000)

        elif choice == "Shock Speed vs Frequency":
            self.canvas.ax.scatter(self.freq, self.shock_speed, color='purple', s=10)
            self.canvas.ax.set_xlabel("Frequency (MHz)")
            self.canvas.ax.set_ylabel("Shock Speed (km/s)")
            self.canvas.ax.set_title(f"{self.filename}_Shock_Speed_vs_Frequency")
            self.current_plot_title = f"{self.filename}_Shock_Speed_vs_Frequency"
            self.status.showMessage("Shock Speed vs Frequency plotted successfully!", 3000)

        elif choice == "Shock Height vs Frequency":
            self.canvas.ax.scatter(self.R_p, self.freq, color='red', s=10)
            self.canvas.ax.set_xlabel("Shock Height (Rₛ)")
            self.canvas.ax.set_ylabel("Frequency (MHz)")
            self.canvas.ax.set_title(f"{self.filename}_Rs_vs_Freq")
            self.current_plot_title = f"{self.filename}_Rs_vs_Freq"
            self.status.showMessage("Shock Height vs Frequency plotted successfully!", 3000)
        self.canvas.ax.grid(True)
        self.canvas.draw()


class CombineFrequencyDialog(QDialog):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.setWindowTitle("Combine Frequency Ranges")
        self.setMinimumWidth(600)

        self.file_paths = []

        self.load_button = QPushButton("Import FITS Files")
        self.load_button.clicked.connect(self.load_files)

        self.combine_button = QPushButton("Combine")
        self.combine_button.clicked.connect(self.combine_files)
        self.combine_button.setEnabled(False)

        self.import_button = QPushButton("Import to Analyzer")
        self.import_button.clicked.connect(self.import_to_main)
        self.import_button.setEnabled(False)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)

        self.image_label = QLabel("Combined output will appear here.")
        self.image_label.setAlignment(Qt.AlignCenter)

        layout = QVBoxLayout()
        layout.addWidget(self.load_button)
        layout.addWidget(self.combine_button)
        layout.addWidget(self.import_button)
        layout.addWidget(self.image_label)
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)

        self.combined_data = None
        self.combined_freqs = None
        self.combined_time = None
        self.combined_filename = "Combined_Frequency"

    def load_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select FITS Files to Combine",
            "",
            "FITS files (*.fit *.fit.gz)"
        )
        if len(files) != 2:
            QMessageBox.warning(self, "Error", "Please select exactly TWO files.")
            return

        station1 = files[0].split("/")[-1].split("_")[0]
        station2 = files[1].split("/")[-1].split("_")[0]

        if station1 != station2:
            QMessageBox.critical(self, "Error",
                                 "You must select consecutive frequency data files from the same station!")
            return

        self.file_paths = files
        self.combine_button.setEnabled(True)

    def combine_files(self):
        from astropy.io import fits
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(10)
        QApplication.processEvents()

        try:
            hdul1 = fits.open(self.file_paths[0])
            data1 = hdul1[0].data
            freqs1 = hdul1[1].data['frequency'][0]
            time1 = hdul1[1].data['time'][0]
            hdul1.close()
            self.progress_bar.setValue(30)
            QApplication.processEvents()

            hdul2 = fits.open(self.file_paths[1])
            data2 = hdul2[0].data
            freqs2 = hdul2[1].data['frequency'][0]
            time2 = hdul2[1].data['time'][0]
            hdul2.close()
            self.progress_bar.setValue(60)
            QApplication.processEvents()

            if not np.allclose(time1, time2, rtol=1e-2):
                QMessageBox.critical(self, "Error", "Time arrays must match to combine frequencies.")
                self.progress_bar.setVisible(False)
                return

            self.combined_data = np.vstack([data1, data2])
            self.combined_freqs = np.concatenate([freqs1, freqs2])
            self.combined_time = time1
            self.progress_bar.setValue(80)
            QApplication.processEvents()

            # Plot image
            fig, ax = plt.subplots(figsize=(6, 4))
            extent = [0, self.combined_time[-1], self.combined_freqs[-1], self.combined_freqs[0]]
            cmap = mcolors.LinearSegmentedColormap.from_list("custom", [(0.0, 'blue'), (0.5, 'red'), (1.0, 'yellow')])
            ax.imshow(self.combined_data, aspect='auto', extent=extent, cmap=cmap)
            ax.set_xlabel("Time [s]")
            ax.set_ylabel("Frequency [MHz]")
            # Extract base filenames (e.g., 'BIR_20240720_123000_123000_00.fit.gz')
            fname1 = os.path.basename(self.file_paths[0])
            fname2 = os.path.basename(self.file_paths[1])

            # Extract focus codes (last 2 digits before .fit.gz, assuming filename ends with _00.fit.gz or _01.fit.gz etc.)
            focus1 = fname1.split("_")[-1].split(".")[0]
            focus2 = fname2.split("_")[-1].split(".")[0]

            # Extract common base (e.g., remove focus code and extension)
            base_name = "_".join(fname1.split("_")[:-1])

            # Set title with base + both focus codes
            ax.set_title(f"{base_name}_{focus1}+{focus2} (Combined Frequency)")

            self.combined_title = f"{base_name}_{focus1}+{focus2} (Combined Frequency)"
            ax.set_title(self.combined_title)

            buf = io.BytesIO()
            fig.savefig(buf, format='png')
            buf.seek(0)
            img = QImage()
            img.loadFromData(buf.read())
            self.image_label.setPixmap(QPixmap.fromImage(img).scaledToWidth(550))
            buf.close()
            plt.close(fig)

            self.progress_bar.setValue(100)
            QApplication.processEvents()
            self.import_button.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            self.progress_bar.setVisible(False)

    def import_to_main(self):
        if self.combined_data is None or self.combined_freqs is None or self.combined_time is None:
            QMessageBox.warning(self, "No Data", "Please combine the files first.")
            return
        self.main_window.raw_data = self.combined_data
        self.main_window.freqs = self.combined_freqs
        self.main_window.time = self.combined_time
        self.main_window.filename = self.combined_title  # ✅ update filename as the title
        self.main_window.plot_data(self.combined_data, title="Raw Data (Combined Frequency)")
        self.close()


class CombineTimeDialog(QDialog):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.setWindowTitle("Combine Time Ranges")
        self.setMinimumWidth(600)

        self.file_paths = []
        self.combined_data = None

        # Buttons
        self.load_button = QPushButton("Import FITS Files")
        self.load_button.clicked.connect(self.load_files)

        self.combine_button = QPushButton("Combine")
        self.combine_button.clicked.connect(self.combine_files)
        self.combine_button.setEnabled(False)

        self.import_button = QPushButton("Import to Analyzer")
        self.import_button.clicked.connect(self.import_to_main)
        self.import_button.setEnabled(False)

        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)

        # Output Image Preview
        self.image_label = QLabel("Combined output will appear here.")
        self.image_label.setAlignment(Qt.AlignCenter)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.load_button)
        layout.addWidget(self.combine_button)
        layout.addWidget(self.import_button)
        layout.addWidget(self.image_label)
        layout.addWidget(self.progress_bar)
        self.setLayout(layout)

    def load_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select FITS Files to Combine",
            "",
            "FITS files (*.fit *.fit.gz)"
        )

        if len(files) < 2:
            QMessageBox.warning(self, "Error", "Please select at least 2 FITS files.")
            return

        try:
            from datetime import datetime

            # Sort files by timestamp
            self.file_paths = sorted(files, key=lambda f: os.path.basename(f).split("_")[2])

            # Check station/date and time continuity
            parts_ref = os.path.basename(self.file_paths[0]).split("_")
            t_ref = datetime.strptime(parts_ref[2], "%H%M%S")

            for f in self.file_paths[1:]:
                parts = os.path.basename(f).split("_")
                if parts[0] != parts_ref[0] or parts[1] != parts_ref[1]:
                    raise ValueError("Different station or date")

                t_next = datetime.strptime(parts[2], "%H%M%S")
                diff = abs((t_next - t_ref).total_seconds())
                if not (800 <= diff <= 1000):  # ~15min ±1.5min
                    raise ValueError(f"File {f} is not consecutive")
                t_ref = t_next

            self.combine_button.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Invalid Selection", f"Error while validating files:\n{str(e)}")

    def combine_files(self):
        if len(self.file_paths) < 2:
            QMessageBox.warning(self, "Error", "Please load at least 2 valid FITS files to combine.")
            return

        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(10)

        try:
            combined_data = None
            combined_time = None
            reference_freqs = None

            for idx, file_path in enumerate(self.file_paths):
                hdul = fits.open(file_path)
                data = hdul[0].data
                freqs = hdul[1].data["frequency"][0]
                time = hdul[1].data["time"][0]

                if reference_freqs is None:
                    reference_freqs = freqs
                elif not np.allclose(freqs, reference_freqs):
                    raise ValueError("Frequency mismatch in file: " + os.path.basename(file_path))

                # Compute dt and shift time
                if idx == 0:
                    dt = time[1] - time[0]
                    adjusted_time = time
                    combined_data = data
                    combined_time = adjusted_time
                else:
                    dt = time[1] - time[0]
                    shift = combined_time[-1] + dt
                    adjusted_time = time + shift
                    combined_data = np.concatenate((combined_data, data), axis=1)
                    combined_time = np.concatenate((combined_time, adjusted_time))

                hdul.close()

            self.combined_data = combined_data
            self.combined_time = combined_time
            self.main_window.freqs = reference_freqs
            self.main_window.time = combined_time

            self.progress_bar.setValue(80)

            # Plot preview
            fig, ax = plt.subplots(figsize=(6, 4))
            extent = [combined_time[0], combined_time[-1], reference_freqs[-1], reference_freqs[0]]
            cmap = LinearSegmentedColormap.from_list('custom_cmap', [(0, 'darkblue'), (1, 'orange')])
            im = ax.imshow(combined_data, aspect='auto', extent=extent, cmap=cmap)
            ax.set_xlabel("Time [s]")
            ax.set_ylabel("Frequency [MHz]")
            ax.set_title("Combined Time Plot")
            fig.tight_layout()

            temp_dir = tempfile.gettempdir()
            preview_path = os.path.join(temp_dir, "preview_combined_time.png")
            fig.savefig(preview_path, dpi=100)
            plt.close(fig)

            self.image_label.setPixmap(QPixmap(preview_path).scaled(550, 350, Qt.KeepAspectRatio))
            self.progress_bar.setValue(100)
            self.import_button.setEnabled(True)

            # Set filename
            base1 = os.path.basename(self.file_paths[0]).split(".")[0]
            self.main_window.filename = base1 + "_combined_time"

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to combine:\n{str(e)}")

    def import_to_main(self):
        if self.combined_data is not None and self.combined_time is not None:
            self.main_window.raw_data = self.combined_data
            self.main_window.freqs = self.main_window.freqs  # already set earlier
            self.main_window.time = self.combined_time
            self.main_window.filename = self.main_window.filename  # already set earlier

            # Calculate UT start from FITS header of first file
            try:
                hdul = fits.open(self.file_paths[0])
                hdr = hdul[0].header
                hh, mm, ss = hdr['TIME-OBS'].split(":")
                hh = int(hh)
                mm = int(mm)
                ss = float(ss)
                self.main_window.ut_start_sec = hh * 3600 + mm * 60 + ss
                hdul.close()
            except Exception as e:
                print("⚠️ Could not extract UT time from first file:", e)
                self.main_window.ut_start_sec = None

            self.main_window.plot_data(self.combined_data, title="Combined Time")
            self.close()
